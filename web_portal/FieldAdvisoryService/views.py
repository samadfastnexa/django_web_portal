from rest_framework import viewsets,permissions,filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from django_filters.rest_framework import DjangoFilterBackend
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from accounts.hierarchy_filters import HierarchyFilterMixin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Dealer, MeetingSchedule, SalesOrder
from .serializers import (
    DealerSerializer, DealerRequestSerializer, MeetingScheduleSerializer, SalesOrderSerializer, SalesOrderFormSerializer,
    CompanySerializer, RegionSerializer, ZoneSerializer, TerritorySerializer, DealerRequestSerializer,
    CompanyNestedSerializer, RegionNestedSerializer, ZoneNestedSerializer, TerritoryNestedSerializer
)
from .models import DealerRequest
from .models import Company, Region, Zone, Territory
from drf_yasg.utils import swagger_auto_schema, no_body
from .permissions import IsAdminOrReadOnly
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from accounts.permissions import HasRolePermission
from rest_framework import viewsets

from drf_yasg import openapi


class MeetingScheduleViewSet(HierarchyFilterMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing Meeting Schedules.
    Supports list, retrieve, create, update, partial update, and delete.
    Filters data based on user's position in reporting hierarchy.
    """
    queryset = MeetingSchedule.objects.select_related(
        'region', 
        'zone', 
        'territory', 
        'staff',
        'region__company',
        'zone__company',
        'zone__region',
        'territory__company',
        'territory__zone',
        'territory__zone__region'
    ).prefetch_related('attendees').all()
    serializer_class = MeetingScheduleSerializer
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    permission_classes = [IsAuthenticated, HasRolePermission]
    hierarchy_field = 'staff'  # Filter by staff who created the meeting
    filterset_fields = ["fsm_name", "region", "zone", "territory", "location", "presence_of_zm", "presence_of_rsm", "staff"]
    search_fields = ["fsm_name", "region__name", "zone__name", "territory__name", "location", "key_topics_discussed"]
    ordering_fields = ["date", "fsm_name", "region__name", "zone__name", "territory__name", "total_attendees"]
    ordering = ["-id"]
    common_parameters = [
        openapi.Parameter('fsm_name', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True, description='Field Sales Manager name'),
        openapi.Parameter('territory_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Territory ID'),
        openapi.Parameter('zone_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Zone ID'),
        openapi.Parameter('region_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Region ID'),
        openapi.Parameter('date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=True, description='Meeting date (YYYY-MM-DD)'),
        openapi.Parameter('location', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True, description='Meeting location'),
        openapi.Parameter('total_attendees', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Total number of attendees'),
        openapi.Parameter('key_topics_discussed', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Key topics discussed'),
        openapi.Parameter('presence_of_zm', openapi.IN_FORM, type=openapi.TYPE_BOOLEAN, required=False, description='Presence of Zone Manager'),
        openapi.Parameter('presence_of_rsm', openapi.IN_FORM, type=openapi.TYPE_BOOLEAN, required=False, description='Presence of Regional Sales Manager'),
        openapi.Parameter('feedback_from_attendees', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Feedback from attendees'),
        openapi.Parameter('suggestions_for_future', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Suggestions for future'),
        openapi.Parameter('attendee_name', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING), required=False, description='List of farmer names'),
        openapi.Parameter('attendee_contact', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING), required=False, description='List of contact numbers'),
        openapi.Parameter('attendee_acreage', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER), required=False, description='List of acreage values'),
        openapi.Parameter('attendee_crop', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING), required=False, description='List of crops'),
        openapi.Parameter('attendee_farmer_id', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING), required=False, description='List of farmer IDs to link existing farmers'),
    ]

    @swagger_auto_schema(
        operation_description="Retrieve a list of meeting schedules with FSM name, region, zone, territory, date, location, and attendee summary. Supports filtering by staff ID (to see specific user's records).",
        manual_parameters=[
            openapi.Parameter('staff', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Staff ID (User ID). Use to see specific user records.'),
            openapi.Parameter('fsm_name', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by Field Sales Manager name'),
            openapi.Parameter('region', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Region ID'),
            openapi.Parameter('zone', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Zone ID'),
            openapi.Parameter('territory', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Territory ID'),
            openapi.Parameter('location', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by Location'),
            openapi.Parameter('date', openapi.IN_QUERY, type=openapi.TYPE_STRING, format='date', required=False, description='Filter by Date (YYYY-MM-DD)'),
            openapi.Parameter('key_topics_discussed', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by key topics'),
            openapi.Parameter('presence_of_zm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by presence of Zone Manager'),
            openapi.Parameter('presence_of_rsm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by presence of Regional Sales Manager'),
        ],
        responses={
            200: openapi.Response(
                description='List of meeting schedules',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'meeting_id': 'FAS01',
                            'fsm_name': 'Ahmed Ali',
                            'company_name': 'Tarzan Crop Sciences',
                            'region_id': 1,
                            'region_name': 'Punjab',
                            'zone_id': 1,
                            'zone_name': 'Central Punjab',
                            'territory_id': 1,
                            'territory_name': 'Lahore',
                            'date': '2024-01-15',
                            'location': 'Community Center',
                            'total_attendees': 25,
                            'key_topics_discussed': 'Crop rotation, pest management',
                            'presence_of_zm': True,
                            'presence_of_rsm': False,
                            'attendees': [
                                {
                                    'id': 1,
                                    'farmer_name': 'Farmer John',
                                    'contact_number': '+92-300-1234567',
                                    'acreage': 5.5,
                                    'crop': 'Wheat'
                                }
                            ]
                        }
                    ]
                }
            )
        },
        tags=["14. MeetingSchedules"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Get detailed information of a specific meeting schedule including attendees.",
        responses={
            200: openapi.Response(
                description='Meeting schedule details',
                examples={
                    'application/json': {
                        'id': 1,
                        'fsm_name': 'Ahmed Ali',
                        'company_name': 'Tarzan Crop Sciences',
                        'region_id': 1,
                        'region_name': 'Punjab',
                        'zone_id': 1,
                        'zone_name': 'Central Punjab',
                        'territory_id': 1,
                        'territory_name': 'Lahore',
                        'date': '2024-01-15',
                        'location': 'Community Center',
                        'total_attendees': 25,
                        'key_topics_discussed': 'Crop rotation, pest management, irrigation techniques',
                        'presence_of_zm': True,
                        'presence_of_rsm': True,
                        'feedback_from_attendees': 'Informative session',
                        'suggestions_for_future': 'More demonstrations',
                        'attendees': [
                            {
                                'id': 1,
                                'farmer_name': 'Farmer John',
                                'contact_number': '+92-300-1234567',
                                'acreage': 5.5,
                                'crop': 'Wheat'
                            }
                        ]
                    }
                }
            ),
            404: 'Meeting schedule not found'
        },
        tags=["14. MeetingSchedules"]
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Export meeting schedules to Excel. Returns an Excel file with all filtered meeting schedules and their attendee details. Use 'ids' parameter to download specific records (comma-separated) or 'all' combined with 'staff' parameter to download all records for a specific user. Respects all query parameters used for filtering.",
        manual_parameters=[
            openapi.Parameter('ids', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Download specific meetings by IDs (comma-separated, e.g., 123,456,789) or "all" to download all records (use with "staff" parameter to specify user)'),
            openapi.Parameter('staff', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Staff ID (User ID)'),
            openapi.Parameter('fsm_name', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by Field Sales Manager name'),
            openapi.Parameter('region', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Region ID'),
            openapi.Parameter('zone', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Zone ID'),
            openapi.Parameter('territory', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by Territory ID'),
            openapi.Parameter('location', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by Location'),
            openapi.Parameter('presence_of_zm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by presence of Zone Manager'),
            openapi.Parameter('presence_of_rsm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by presence of Regional Sales Manager'),
            openapi.Parameter('search', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Search in FSM name, region, zone, territory, location, topics'),
            openapi.Parameter('ordering', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Order by field (e.g., date, -date, fsm_name, -id)'),
        ],
        responses={
            200: openapi.Response(
                description='Excel file download',
                schema=openapi.Schema(
                    type=openapi.TYPE_FILE,
                    format='binary'
                )
            )
        },
        tags=["14. MeetingSchedules"]
    )
    @action(detail=False, methods=['get'], url_path='export-to-excel')
    def export_to_excel(self, request):
        """Export filtered meeting schedules to Excel"""
        from datetime import datetime
        queryset = self.filter_queryset(self.get_queryset())
        
        # Handle multiple IDs if provided
        ids_param = request.query_params.get('ids', None)
        if ids_param:
            # Check if user wants all records for a specific user
            if ids_param.strip().lower() == 'all':
                # If 'staff' parameter is provided, get all records for that user
                staff_id = request.query_params.get('staff', None)
                if staff_id:
                    queryset = queryset.filter(staff_id=staff_id)
                # Otherwise, queryset is already filtered by get_queryset() and filters
            else:
                # Split comma-separated IDs and filter
                id_list = [int(id.strip()) for id in ids_param.split(',') if id.strip().isdigit()]
                queryset = queryset.filter(id__in=id_list)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Meeting Schedules"
        
        # Add export date/time at the top
        export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"Field Advisory Meeting Schedule Export - {export_time}"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:D1')
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        info_label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        info_label_font = Font(bold=True, size=10, color="1F4E78")
        
        odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        data_alignment = Alignment(vertical="top", wrap_text=True)
        
        current_row = 3
        
        for meeting in queryset:
            # Meeting Information Section
            info_fields = [
                ('Meeting ID', meeting.meeting_id),
                ('FSM Name', meeting.fsm_name),
                ('Date', meeting.date.strftime('%Y-%m-%d %H:%M') if meeting.date else ''),
                ('Region', meeting.region.name if meeting.region else ''),
                ('Zone', meeting.zone.name if meeting.zone else ''),
                ('Territory', meeting.territory.name if meeting.territory else ''),
                ('Location', meeting.location),
                ('Total Attendees', meeting.total_attendees),
                ('Confirmed Attendees', meeting.confirmed_attendees),
                ('Min Farmers Required', meeting.min_farmers_required),
                ('ZM Present', 'Yes' if meeting.presence_of_zm else 'No'),
                ('RSM Present', 'Yes' if meeting.presence_of_rsm else 'No'),
                ('Key Topics Discussed', meeting.key_topics_discussed),
                ('Feedback', meeting.feedback_from_attendees or ''),
                ('Suggestions', meeting.suggestions_for_future or '')
            ]
            
            # Write meeting information as key-value pairs
            for label, value in info_fields:
                label_cell = ws.cell(row=current_row, column=1, value=label)
                label_cell.fill = info_label_fill
                label_cell.font = info_label_font
                label_cell.border = thin_border
                label_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                value_cell = ws.cell(row=current_row, column=2, value=value)
                value_cell.border = thin_border
                value_cell.alignment = data_alignment
                ws.merge_cells(f'B{current_row}:D{current_row}')
                
                current_row += 1
            
            current_row += 1  # Empty row
            
            # Attendee table headers
            attendee_headers = ['Attendee Name', 'Contact Number', 'Acreage', 'Crop']
            for col_num, header in enumerate(attendee_headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            current_row += 1
            
            # Write attendees
            attendees = meeting.attendees.all()
            if attendees.exists():
                for idx, attendee in enumerate(attendees):
                    is_odd_row = idx % 2 == 0
                    row_fill = odd_row_fill if is_odd_row else even_row_fill
                    
                    attendee_data = [
                        attendee.farmer_name,
                        attendee.contact_number,
                        attendee.acreage,
                        attendee.crop
                    ]
                    
                    for col_num, value in enumerate(attendee_data, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=value)
                        cell.fill = row_fill
                        cell.border = thin_border
                        cell.alignment = data_alignment
                    
                    current_row += 1
            else:
                # No attendees message
                no_attendee_cell = ws.cell(row=current_row, column=1, value="No attendees")
                no_attendee_cell.font = Font(italic=True, color="999999")
                ws.merge_cells(f'A{current_row}:D{current_row}')
                current_row += 1
            
            current_row += 2  # Space before next meeting
        
        # Adjust column widths
        from openpyxl.cell.cell import MergedCell
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                if col_letter is None:
                    col_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if col_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        ws.row_dimensions[1].height = 25
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=meeting_schedules.xlsx'
        wb.save(response)
        return response

    @swagger_auto_schema(
        operation_description="Create a meeting schedule with FSM name, region, zone, territory, and attendees.",
        manual_parameters=common_parameters,
        request_body=None,
        responses={
            201: 'Meeting schedule created successfully',
            400: 'Bad Request - Invalid data provided'
        },
        tags=["14. MeetingSchedules"]
    )
    def create(self, request, *args, **kwargs):
        # Auto-assign the logged-in user as staff if not provided (though it's read-only now)
        # Note: In standard perform_create we usually do serializer.save(staff=self.request.user)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        if not user or not user.is_authenticated:
            raise PermissionDenied("Authentication required to create meeting schedule.")
        
        # Use fsm_name from request if provided, otherwise fallback to user's name
        fsm_name = serializer.validated_data.get('fsm_name')
        
        if not fsm_name:
            # Fallback to logged-in user's name if fsm_name not provided
            if hasattr(user, "get_full_name") and callable(getattr(user, "get_full_name", None)):
                fsm_name = user.get_full_name() or getattr(user, "username", "") or getattr(user, "email", "")
            elif hasattr(user, "full_name"):
                fsm_name = getattr(user, "full_name") or getattr(user, "username", "") or getattr(user, "email", "")
            elif hasattr(user, "first_name") or hasattr(user, "last_name"):
                first = getattr(user, "first_name", "") or ""
                last = getattr(user, "last_name", "") or ""
                combined = f"{first} {last}".strip()
                fsm_name = combined or getattr(user, "username", "") or getattr(user, "email", "")
            else:
                fsm_name = getattr(user, "username", "") or getattr(user, "email", "") or str(getattr(user, "pk", ""))

        serializer.save(staff=user, fsm_name=fsm_name)

    @swagger_auto_schema(
        operation_description="Update all details of an existing meeting schedule (full update).",
        manual_parameters=common_parameters,
        request_body=None,
        responses={
            200: 'Meeting schedule updated successfully',
            404: 'Meeting schedule not found',
            400: 'Bad Request - Invalid data provided'
        },
        tags=["14. MeetingSchedules"]
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Update specific fields of a meeting schedule (partial update).",
        manual_parameters=common_parameters,
        request_body=None,
        responses={
            200: 'Meeting schedule updated successfully',
            404: 'Meeting schedule not found',
            400: 'Bad Request - Invalid data provided'
        },
        tags=["14. MeetingSchedules"]
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Cancel and permanently delete a meeting schedule.",
        responses={
            204: 'Meeting schedule deleted successfully',
            404: 'Meeting schedule not found'
        },
        tags=["14. MeetingSchedules"]
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)


class SalesOrderViewSet(HierarchyFilterMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing Sales Orders.
    Supports list, retrieve, create, update, partial update, and delete.
    Accepts form-data format only.
    Filters data based on user's position in reporting hierarchy.
    """
    queryset = SalesOrder.objects.select_related(
        'customer',
        'staff',
        'staff__sales_profile',
        'territory',
        'territory__zone',
        'territory__zone__region',
        'zone',
        'zone__region',
        'region'
    ).prefetch_related(
        'lines',
        'lines__item',
        'attachments'
    ).all()
    serializer_class = SalesOrderSerializer
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated, HasRolePermission]
    ordering = ['-id']
    hierarchy_field = 'staff'  # Filter by staff who created the order
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return SalesOrderFormSerializer
        return SalesOrderSerializer

    @swagger_auto_schema(
        operation_description="Retrieve a list of all sales orders with their status, dealer, and meeting schedule information.",
        responses={
            200: openapi.Response(
                description='List of sales orders',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'portal_order_id': 'SO001',
                            'schedule': 1,
                            'staff': 1,
                            'dealer': 1,
                            'status': 'pending',
                            'card_code': 'C001',
                            'card_name': 'ABC Company',
                            'created_at': '2024-01-15T10:30:00Z'
                        },
                        {
                            'id': 2,
                            'portal_order_id': 'SA002',
                            'schedule': 2,
                            'staff': 1,
                            'dealer': 2,
                            'status': 'entertained',
                            'card_code': 'C002',
                            'card_name': 'XYZ Company',
                            'created_at': '2024-01-16T14:20:00Z'
                        }
                    ]
                }
            )
        },
        tags=["15. SalesOrders"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Get detailed information of a specific sales order including all related data.",
        responses={
            200: openapi.Response(
                description='Sales order details',
                examples={
                    'application/json': {
                        'id': 1,
                        'portal_order_id': 'SO001',
                        'schedule': 1,
                        'staff': 1,
                        'dealer': 1,
                        'status': 'entertained',
                        'card_code': 'C001',
                        'card_name': 'ABC Company',
                        'created_at': '2024-01-15T10:30:00Z'
                    }
                }
            ),
            404: 'Sales order not found'
        },
        tags=["15. SalesOrders"]
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="""
        Create a new sales order with header fields and line items using form-data.
        
        **To add multiple line items:**
        - Provide arrays for item fields (item_code, quantity, unit_price, etc.)
        - Each array index represents one line item
        - Example: item_code[0]="SEED-001", quantity[0]=10, item_code[1]="FERT-001", quantity[1]=5
        
        The arrays will be automatically zipped together to create multiple line items.
        """,
        request_body=no_body,
        manual_parameters=[
            # Header fields
            openapi.Parameter('staff', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Staff user ID'),
            openapi.Parameter('dealer', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Dealer ID'),
            openapi.Parameter('status', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Order status'),
            openapi.Parameter('series', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='SAP series'),
            openapi.Parameter('doc_type', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Document type'),
            openapi.Parameter('doc_date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=False, description='Document date (YYYY-MM-DD)'),
            openapi.Parameter('doc_due_date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=False, description='Document due date (YYYY-MM-DD)'),
            openapi.Parameter('tax_date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=False, description='Tax date (YYYY-MM-DD)'),
            openapi.Parameter('card_code', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Customer card code'),
            openapi.Parameter('card_name', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Customer name'),
            openapi.Parameter('contact_person_code', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Contact person code'),
            openapi.Parameter('federal_tax_id', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Federal tax ID'),
            openapi.Parameter('address', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Customer address'),
            openapi.Parameter('doc_currency', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Document currency'),
            openapi.Parameter('doc_rate', openapi.IN_FORM, type=openapi.TYPE_NUMBER, required=False, description='Document exchange rate'),
            openapi.Parameter('comments', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Order comments'),
            openapi.Parameter('u_sotyp', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Sales order type UDF'),
            openapi.Parameter('u_usid', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='User ID UDF'),
            openapi.Parameter('u_s_card_code', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Secondary card code UDF'),
            openapi.Parameter('u_s_card_name', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Secondary card name UDF'),
            
            # Line item array fields - add multiple items
            openapi.Parameter('item_code', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING), 
                            required=False, description='Array of item codes. Example: ["SEED-001", "FERT-001"]'),
            openapi.Parameter('item_description', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of item descriptions. Example: ["Corn Seed", "Fertilizer"]'),
            openapi.Parameter('quantity', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of quantities. Example: [10, 5]'),
            openapi.Parameter('unit_price', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of unit prices. Example: [2500, 1500]'),
            openapi.Parameter('discount_percent', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of discount percentages. Example: [0, 5]'),
            openapi.Parameter('warehouse_code', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of warehouse codes. Example: ["WH01", "WH01"]'),
            openapi.Parameter('vat_group', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of VAT groups. Example: ["AT1", "AT1"]'),
            openapi.Parameter('u_crop', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of crop codes. Example: ["CORN", "WHEAT"]'),
            openapi.Parameter('tax_percentage_per_row', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of tax percentages per line. Example: [0, 17]'),
            openapi.Parameter('units_of_measurment', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of UoM conversion factors. Example: [1, 1]'),
            openapi.Parameter('uom_entry', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_INTEGER),
                            required=False, description='Array of UoM entry IDs'),
            openapi.Parameter('measure_unit', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of measure unit names'),
            openapi.Parameter('uom_code', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of UoM codes'),
            openapi.Parameter('project_code', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of project codes'),
            openapi.Parameter('u_sd', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of special discount percentages'),
            openapi.Parameter('u_ad', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of additional discount percentages'),
            openapi.Parameter('u_exd', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of extra discount percentages'),
            openapi.Parameter('u_zerop', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of phase discount percentages'),
            openapi.Parameter('u_pl', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_INTEGER),
                            required=False, description='Array of Policy Link IDs'),
            openapi.Parameter('u_bp', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_NUMBER),
                            required=False, description='Array of Project Balance values'),
            openapi.Parameter('u_policy', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of Policy Codes'),
            openapi.Parameter('u_focitem', openapi.IN_FORM, type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING),
                            required=False, description='Array of FOC flags (Yes/No). Example: ["No", "Yes"]'),
        ],
        responses={
            201: openapi.Response(
                description='Sales order created successfully',
                examples={
                    'application/json': {
                        'id': 1,
                        'portal_order_id': 'SO001',
                        'staff': 1,
                        'dealer': 2,
                        'status': 'pending',
                        'card_code': 'C20000',
                        'card_name': 'ABC Traders',
                        'created_at': '2024-01-15T10:30:00Z',
                        'document_lines': [
                            {
                                'id': 1,
                                'line_num': 1,
                                'item_code': 'ITEM-001',
                                'item_description': 'Hybrid Seed',
                                'quantity': 10,
                                'unit_price': 2500,
                                'discount_percent': 0,
                                'warehouse_code': 'WH01',
                                'vat_group': 'SE',
                                'tax_percentage_per_row': 0,
                                'units_of_measurment': 1,
                                'uom_entry': 1,
                                'measure_unit': 'KG',
                                'uom_code': 'KG',
                                'project_code': None,
                                'u_sd': 0,
                                'u_ad': 0,
                                'u_exd': 0,
                                'u_zerop': 0,
                                'u_pl': 123,
                                'u_bp': 5000.0,
                                'u_policy': 'POL-001',
                                'u_focitem': 'No',
                                'u_crop': 'CORN'
                            }
                        ]
                    }
                }
            ),
            400: 'Bad Request - Invalid data provided'
        },
        tags=["15. SalesOrders"]
    )
    def create(self, request, *args, **kwargs):
        # Use SalesOrderFormSerializer for input validation
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        
        # Use SalesOrderSerializer for output to include document_lines
        output_serializer = SalesOrderSerializer(instance)
        headers = self.get_success_headers(output_serializer.data)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @swagger_auto_schema(
        operation_description="Update all details of an existing sales order (full update) using form-data.",
        request_body=no_body,
        manual_parameters=[
            openapi.Parameter('staff', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Staff user ID'),
            openapi.Parameter('dealer', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Dealer ID'),
            openapi.Parameter('status', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Order status'),
            openapi.Parameter('document_lines', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='JSON string array of line items'),
        ],
        responses={
            200: 'Sales order updated successfully',
            404: 'Sales order not found',
            400: 'Bad Request - Invalid data provided'
        },
        tags=["15. SalesOrders"]
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="""
        Update specific fields of a sales order (PATCH method) using form-data.
        
        **All fields are optional** - send only the fields you want to update.
        Typically used for status changes, adding comments, or updating SAP posting information.
        """,
        request_body=no_body,
        manual_parameters=[
            openapi.Parameter('staff', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Staff user ID'),
            openapi.Parameter('dealer', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False, description='Dealer ID'),
            openapi.Parameter('status', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='Order status'),
            openapi.Parameter('document_lines', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False, description='JSON string array of line items'),
        ],
        responses={
            200: 'Sales order updated successfully',
            404: 'Sales order not found',
            400: 'Bad Request - Invalid data'
        },
        tags=["15. SalesOrders"]
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Permanently delete a sales order from the system.",
        responses={
            204: 'Sales order deleted successfully',
            404: 'Sales order not found'
        },
        tags=["15. SalesOrders"]
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    @swagger_auto_schema(
        method='get',
        operation_description="""
        Get all sales orders created by a specific user/staff member.
        
        Pass the user ID as a URL parameter to filter sales orders by the staff who created them.
        Returns all sales orders with their details including status, customer info, and SAP posting status.
        
        Example: `/api/sales-orders/by-user/5/` returns all orders created by user with ID 5
        """,
        responses={
            200: openapi.Response(
                description='List of sales orders created by the specified user',
                schema=SalesOrderSerializer(many=True),
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'staff': 5,
                            'dealer': 2,
                            'card_code': 'C20000',
                            'card_name': 'ABC Traders',
                            'status': 'pending',
                            'is_posted_to_sap': False,
                            'created_at': '2024-01-15T10:30:00Z'
                        },
                        {
                            'id': 2,
                            'staff': 5,
                            'dealer': 3,
                            'card_code': 'C20001',
                            'card_name': 'XYZ Company',
                            'status': 'entertained',
                            'is_posted_to_sap': True,
                            'sap_doc_num': 123456,
                            'created_at': '2024-01-16T14:20:00Z'
                        }
                    ]
                }
            ),
            404: 'User not found or no orders exist for this user'
        },
        tags=["15. SalesOrders"]
    )
    @action(detail=False, methods=['get'], url_path='by-user/(?P<user_id>[^/.]+)')
    def by_user(self, request, user_id=None):
        """Get all sales orders created by a specific user"""
        try:
            # Validate user exists
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            try:
                user = User.objects.get(pk=user_id)
                # Handle different user models - some may not have get_full_name
                if hasattr(user, 'get_full_name') and callable(user.get_full_name):
                    user_name = user.get_full_name() or user.username
                elif hasattr(user, 'full_name'):
                    user_name = user.full_name or user.username
                elif hasattr(user, 'first_name') and hasattr(user, 'last_name'):
                    user_name = f"{user.first_name} {user.last_name}".strip() or user.username
                else:
                    user_name = user.username
            except User.DoesNotExist:
                return Response(
                    {'error': f'User with ID {user_id} not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Get all sales orders for this user
            orders = SalesOrder.objects.filter(staff=user).order_by('-created_at')
            
            if not orders.exists():
                return Response(
                    {
                        'message': f'No sales orders found for user {user_name} (ID: {user_id})',
                        'user_id': user_id,
                        'user_name': user_name,
                        'count': 0,
                        'orders': []
                    },
                    status=status.HTTP_200_OK
                )
            
            serializer = self.get_serializer(orders, many=True)
            
            return Response({
                'user_id': user_id,
                'user_name': user_name,
                'count': orders.count(),
                'orders': serializer.data
            }, status=status.HTTP_200_OK)
            
        except ValueError:
            return Response(
                {'error': 'Invalid user ID format. Must be a valid integer.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': f'Error retrieving sales orders: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DealerViewSet(viewsets.ModelViewSet):
    queryset = Dealer.objects.all()
    serializer_class = DealerSerializer
    ordering = ['-id']

    @swagger_auto_schema(
        operation_description="List all dealers with their user credentials",
        responses={200: DealerSerializer(many=True)},
        tags=["16. Dealers"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Retrieve a single dealer by ID with user account details",
        responses={200: DealerSerializer},
        tags=["16. Dealers"]
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Create a new dealer with user login credentials. Either provide an existing user_id or provide username/email/password to create a new user account for the dealer. CNIC images (cnic_front_image, cnic_back_image) are optional and can be uploaded later.",
        request_body=DealerSerializer,
        responses={201: DealerSerializer},
        tags=["16. Dealers"]
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Update an existing dealer and optionally update user account credentials",
        responses={200: DealerSerializer},
        tags=["16. Dealers"]
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Partially update an existing dealer",
        responses={200: DealerSerializer},
        tags=["16. Dealers"]
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Delete a dealer (also deletes associated user account if not used elsewhere)",
        responses={204: 'No Content'},
        tags=["16. Dealers"]
    )
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)
    
class DealerRequestViewSet(viewsets.ModelViewSet):
    serializer_class = DealerRequestSerializer
    permission_classes = [IsAuthenticated, HasRolePermission]
    parser_classes = [MultiPartParser, FormParser]  # Accept form-data
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'filer_status', 'card_type', 'is_posted_to_sap', 'requested_by']
    ordering = ['-id']

    def get_queryset(self):
        user = self.request.user

        # Prevent error during swagger schema generation
        if getattr(self, 'swagger_fake_view', False):
            return DealerRequest.objects.none()

        if not user or not user.is_authenticated:
            # Return empty queryset for anonymous users or unauthenticated requests
            return DealerRequest.objects.none()

        base_qs = DealerRequest.objects.all() if (user.is_superuser or (hasattr(user, 'role') and user.role.name == 'Admin')) else DealerRequest.objects.filter(requested_by=user)

        # Support alias 'created_by' for filtering requested_by
        created_by = self.request.query_params.get('created_by')
        if created_by:
            try:
                base_qs = base_qs.filter(requested_by=int(created_by))
            except ValueError:
                base_qs = base_qs.none()

        return base_qs

    @swagger_auto_schema(
        operation_description="Submit a new dealer registration request with comprehensive business partner information including contact details, address, tax information, and SAP configuration. All fields are optional except requested_by (auto-filled).",
        request_body=DealerRequestSerializer,
        responses={
            201: openapi.Response(
                description='Dealer request submitted successfully',
                examples={
                    'application/json': {
                        'id': 1,
                        'business_name': 'Khan Agro Store',
                        'owner_name': 'Ahmed Ali Khan',
                        'contact_number': '+92-300-1234567',
                        'mobile_phone': '+92-300-1234567',
                        'email': 'ahmed@example.com',
                        'address': 'Main Bazaar, Village ABC',
                        'city': 'Lahore',
                        'state': 'Punjab',
                        'country': 'PK',
                        'cnic_number': '12345-6789012-3',
                        'federal_tax_id': 'NTN123456',
                        'filer_status': '01',
                        'status': 'draft',
                        'card_type': 'cCustomer',
                        'group_code': 100,
                        'sap_series': 70,
                        'minimum_investment': 500000,
                        'is_posted_to_sap': False,
                        'created_at': '2024-01-15T10:30:00Z'
                    }
                }
            ),
            400: 'Bad Request - Invalid data or missing required fields'
        },
        tags=["17. Dealer Requests"]
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Retrieve a list of dealer requests. Admins can see all requests; users only see their own. Supports filtering by status, filer_status, card_type, is_posted_to_sap, requested_by, and alias created_by.",
        manual_parameters=[
            openapi.Parameter('status', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Filter by status (draft, pending, approved, rejected, posted_to_sap)'),
            openapi.Parameter('filer_status', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Filter by filer_status (01 for Filer, 02 for Non-Filer)'),
            openapi.Parameter('card_type', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Filter by card_type (cCustomer, cSupplier, cLid)'),
            openapi.Parameter('is_posted_to_sap', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, description='Filter by SAP posting status'),
            openapi.Parameter('requested_by', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Filter by requesting user ID'),
            openapi.Parameter('created_by', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Alias for requested_by; filters by creator user ID'),
        ],
        responses={
            200: openapi.Response(
                description='List of dealer requests',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'business_name': 'Khan Agro Store',
                            'owner_name': 'Ahmed Ali Khan',
                            'contact_number': '+92-300-1234567',
                            'mobile_phone': '+92-300-1234567',
                            'email': 'ahmed@example.com',
                            'city': 'Lahore',
                            'state': 'Punjab',
                            'status': 'pending',
                            'filer_status': '01',
                            'card_type': 'cCustomer',
                            'is_posted_to_sap': False,
                            'sap_card_code': None,
                            'minimum_investment': 500000,
                            'created_at': '2024-01-15T10:30:00Z'
                        },
                        {
                            'id': 2,
                            'business_name': 'Modern Seeds Shop',
                            'owner_name': 'Bilal Hassan',
                            'contact_number': '+92-321-9876543',
                            'status': 'posted_to_sap',
                            'filer_status': '02',
                            'is_posted_to_sap': True,
                            'sap_card_code': 'C000123',
                            'posted_at': '2024-01-20T15:45:00Z',
                            'created_at': '2024-01-18T09:15:00Z'
                        }
                    ]
                }
            )
        },
        tags=["17. Dealer Requests"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Retrieve detailed information of a specific dealer request including all business partner fields, SAP configuration, integration status, and submitted documents.",
        responses={
            200: openapi.Response(
                description='Dealer request details with comprehensive SAP Business Partner information',
                examples={
                    'application/json': {
                        'id': 1,
                        'requested_by': 5,
                        'status': 'posted_to_sap',
                        'reason': 'New dealer in territory 45',
                        'business_name': 'Khan Agro Store',
                        'owner_name': 'Ahmed Ali Khan',
                        'contact_number': '+92-300-1234567',
                        'mobile_phone': '+92-300-1234567',
                        'email': 'ahmed@khanagrostore.com',
                        'address': 'Main Bazaar, Village ABC, District XYZ',
                        'city': 'Lahore',
                        'state': 'Punjab',
                        'country': 'PK',
                        'cnic_number': '12345-6789012-3',
                        'federal_tax_id': 'NTN1234567',
                        'additional_id': 'STRN9876543',
                        'unified_federal_tax_id': 'UFTN1234567',
                        'filer_status': '01',
                        'govt_license_number': 'LIC-2024-001',
                        'license_expiry': '2025-12-31',
                        'u_leg': '17-5349',
                        'cnic_front': '/media/dealer_requests/cnic_front/image1.jpg',
                        'cnic_back': '/media/dealer_requests/cnic_back/image2.jpg',
                        'company': 1,
                        'region': 5,
                        'zone': 12,
                        'territory': 45,
                        'sap_series': 70,
                        'card_type': 'cCustomer',
                        'group_code': 100,
                        'debitor_account': 'A020301001',
                        'vat_group': 'AT1',
                        'vat_liable': 'vLiable',
                        'whatsapp_messages': 'YES',
                        'minimum_investment': 500000,
                        'is_posted_to_sap': True,
                        'sap_card_code': 'C000123',
                        'sap_doc_entry': 456789,
                        'sap_error': None,
                        'posted_at': '2024-01-20T15:45:00Z',
                        'created_at': '2024-01-15T10:30:00Z',
                        'updated_at': '2024-01-20T15:45:00Z',
                        'reviewed_at': '2024-01-20T14:30:00Z',
                        'reviewed_by': 2
                    }
                }
            ),
            404: 'Dealer request not found',
            403: 'Forbidden - You can only view your own requests'
        },
        tags=["17. Dealer Requests"]
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Update a dealer request with complete business partner information. Typically used by admins to approve/reject requests or modify SAP configuration. When status changes to 'approved' or 'posted_to_sap', the system automatically creates the Business Partner in SAP.",
        request_body=DealerRequestSerializer,
        responses={
            200: openapi.Response(
                description='Dealer request updated successfully. If approved, SAP Business Partner is created automatically.',
                examples={
                    'application/json': {
                        'id': 1,
                        'status': 'posted_to_sap',
                        'business_name': 'Khan Agro Store',
                        'is_posted_to_sap': True,
                        'sap_card_code': 'C000123',
                        'sap_doc_entry': 456789,
                        'posted_at': '2024-01-20T15:45:00Z',
                        'message': 'Dealer request approved and Business Partner created in SAP'
                    }
                }
            ),
            400: 'Bad Request - Invalid data',
            404: 'Dealer request not found',
            403: 'Forbidden - Insufficient permissions'
        },
        tags=["17. Dealer Requests"]
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Partially update specific fields of a dealer request. All fields are optional. Useful for updating contact info, address, tax details, or SAP configuration without sending the entire object. Status change to 'approved' or 'posted_to_sap' triggers SAP Business Partner creation.",
        request_body=DealerRequestSerializer,
        responses={
            200: openapi.Response(
                description='Dealer request fields updated successfully',
                examples={
                    'application/json': {
                        'id': 1,
                        'status': 'pending',
                        'email': 'newemail@example.com',
                        'mobile_phone': '+92-321-9999999',
                        'city': 'Karachi',
                        'updated_at': '2024-01-21T10:15:00Z'
                    }
                }
            ),
            400: 'Bad Request - Invalid field values',
            404: 'Dealer request not found',
            403: 'Forbidden - Insufficient permissions'
        },
        tags=["17. Dealer Requests"]
    )
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)


class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.only('id', 'Company_name', 'name', 'email', 'contact_number', 'is_active').all()
    serializer_class = CompanySerializer
    # permission_classes = [IsAdminOrReadOnly]
    permission_classes = [IsAuthenticated, HasRolePermission]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']
    

    @swagger_auto_schema(
        operation_description="Retrieve a list of all companies in the system.",
        responses={
            200: openapi.Response(
                description='List of companies',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'name': 'ABC Agriculture Ltd.',
                            'code': 'ABC001',
                            'address': '123 Business District, Lahore',
                            'contact_number': '+92-42-1234567',
                            'email': 'info@abcagriculture.com'
                        }
                    ]
                }
            )
        },
        tags=["18. Companies"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    @swagger_auto_schema(
        operation_description="Create a new company in the system with all required business information.",
        request_body=CompanySerializer,
        responses={
            201: 'Company created successfully',
            400: 'Bad Request - Invalid data or duplicate code'
        },
        tags=["18. Companies"]
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    @swagger_auto_schema(tags=["18. Company"])
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @swagger_auto_schema(tags=["18. Company"])
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    @swagger_auto_schema(tags=["18. Company"])
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)

    @swagger_auto_schema(tags=["18. Company"])
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.select_related('company').all()
    serializer_class = RegionSerializer
    # permission_classes = [IsAdminOrReadOnly]
    permission_classes = [IsAuthenticated, HasRolePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['company']
    search_fields = ['name']

    @swagger_auto_schema(
        operation_description="Retrieve a list of all regions, optionally filtered by company.",
        responses={
            200: openapi.Response(
                description='List of regions',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'name': 'Punjab Region',
                            'code': 'PUN001',
                            'company': 1,
                            'company_name': 'ABC Agriculture Ltd.'
                        }
                    ]
                }
            )
        },
        tags=["19. Regions"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Create a new region within a company's operational area.",
        request_body=RegionSerializer,
        responses={
            201: 'Region created successfully',
            400: 'Bad Request - Invalid data or duplicate code'
        },
        tags=["19. Regions"]
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)


class ZoneViewSet(viewsets.ModelViewSet):
    queryset = Zone.objects.select_related('company', 'region').all()
    serializer_class = ZoneSerializer
    # permission_classes = [IsAdminOrReadOnly]
    permission_classes = [IsAuthenticated, HasRolePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['region']
    search_fields = ['name']

    @swagger_auto_schema(
        operation_description="Retrieve a list of all zones, optionally filtered by region.",
        responses={
            200: openapi.Response(
                description='List of zones',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'name': 'Lahore Zone',
                            'code': 'LAH001',
                            'region': 1,
                            'region_name': 'Punjab Region'
                        }
                    ]
                }
            )
        },
        tags=["20. Zones"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        operation_description="Create a new zone within a region's operational area.",
        request_body=ZoneSerializer,
        responses={
            201: 'Zone created successfully',
            400: 'Bad Request - Invalid data or duplicate code'
        },
        tags=["20. Zones"]
    )
    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)


class TerritoryViewSet(viewsets.ModelViewSet):
    queryset = Territory.objects.select_related('company', 'zone', 'zone__region').all()
    serializer_class = TerritorySerializer
    # permission_classes = [IsAdminOrReadOnly]
    permission_classes = [IsAuthenticated, HasRolePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['zone', 'company']
    search_fields = ['name']

    @swagger_auto_schema(
        operation_description="Retrieve a list of all territories, optionally filtered by zone.",
        responses={
            200: openapi.Response(
                description='List of territories',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'name': 'Model Town Territory',
                            'code': 'MT001',
                            'zone': 1,
                            'zone_name': 'Lahore Zone'
                        }
                    ]
                }
            )
        },
        tags=["21. Territories"]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Territory.objects.none()
        qs = super().get_queryset()
        company_param = self.request.query_params.get('company')
        if company_param:
            try:
                cid = int(company_param)
                return qs.filter(company_id=cid)
            except Exception:
                pass
        db_key = self.request.session.get('selected_db', '4B-BIO')
        schema = '4B-BIO_APP' if db_key == '4B-BIO' else ('4B-ORANG_APP' if db_key == '4B-ORANG' else '4B-BIO_APP')
        try:
            from FieldAdvisoryService.models import Company
            comp = Company.objects.filter(name=schema).first() or Company.objects.filter(Company_name=schema).first()
            if comp:
                return qs.filter(company=comp)
        except Exception:
            pass
        return qs


class CompanyNestedViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CompanyNestedSerializer
    queryset = Company.objects.prefetch_related(
        'regions__zones__territories'
    ).all()

    @swagger_auto_schema(tags=["22. Companies (Nested View)"])
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
    @swagger_auto_schema(tags=["22. Companies (Nested View)"])
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

class RegionNestedViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = RegionNestedSerializer
    queryset = Region.objects.prefetch_related(
        'zones__territories'
    ).all()

    @swagger_auto_schema(tags=["23. Regions (Nested View)"])
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

class ZoneNestedViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ZoneNestedSerializer
    queryset = Zone.objects.prefetch_related(
        'territories'
    ).all()

    @swagger_auto_schema(tags=["24. Zones (Nested View)"])
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)
    
class TerritoryNestedViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = TerritoryNestedSerializer
    queryset = Territory.objects.select_related('company', 'zone__company', 'zone__region').all()

    @swagger_auto_schema(tags=["25. Territories (Nested View)"])
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)


# SAP LOV API endpoints for admin form dropdowns
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import csrf_exempt
from sap_integration import hana_connect
import os

def _load_env_file(path: str) -> None:
    """Load environment variables from .env file"""
    try:
        if os.path.isfile(path) and os.access(path, os.R_OK):
            with open(path, 'r', encoding='utf-8') as f:
                for line in f:
                    s = line.strip()
                    if s == '' or s.startswith('#') or '=' not in s:
                        continue
                    k, v = s.split('=', 1)
                    k = k.strip()
                    v = v.strip()
                    if v != '' and ((v[0] == '"' and v[-1] == '"') or (v[0] == "'" and v[-1] == "'")):
                        v = v[1:-1]
                    if k != '' and not os.environ.get(k):
                        os.environ[k] = v
    except Exception:
        pass

def get_hana_connection(request=None, selected_db_key=None):
    """Get HANA database connection honoring the selected DB (session/global dropdown)."""
    try:
        # Load .env file
        _load_env_file(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), '.env'))
        try:
            from django.conf import settings as _settings
            _load_env_file(os.path.join(str(_settings.BASE_DIR), '.env'))
            from pathlib import Path as _Path
            _load_env_file(os.path.join(str(_Path(_settings.BASE_DIR).parent), '.env'))
            _load_env_file(os.path.join(os.getcwd(), '.env'))
        except Exception:
            pass
        
        from hdbcli import dbapi
        from preferences.models import Setting

        if not selected_db_key and request and hasattr(request, 'session'):
            selected_db_key = request.session.get('selected_db')

        # Get database name from settings, preferring selected_db_key when provided
        try:
            db_setting = Setting.objects.filter(slug='SAP_COMPANY_DB').first()
            raw_value = getattr(db_setting, 'value', None) if db_setting else None
            schema = os.environ.get('HANA_SCHEMA') or os.environ.get('SAP_COMPANY_DB', '4B-BIO_APP')
            db_options = {}

            if isinstance(raw_value, dict):
                db_options = raw_value
            elif isinstance(raw_value, str):
                try:
                    import json
                    parsed = json.loads(raw_value)
                    if isinstance(parsed, dict):
                        db_options = parsed
                    else:
                        schema = str(parsed)
                except Exception:
                    schema = raw_value

            cleaned = {}
            for k, v in db_options.items():
                clean_key = str(k).strip().strip('"').strip("'")
                clean_val = str(v).strip().strip('"').strip("'")
                cleaned[clean_key] = clean_val
            db_options = cleaned

            if db_options:
                if selected_db_key and selected_db_key in db_options:
                    schema = db_options[selected_db_key]
                else:
                    schema = list(db_options.values())[0]
            elif raw_value and not isinstance(raw_value, dict):
                schema = str(raw_value).strip().strip('"').strip("'")
        except Exception as e:
            print(f"Error getting schema from settings: {e}")
            schema = os.environ.get('HANA_SCHEMA') or os.environ.get('SAP_COMPANY_DB', '4B-BIO_APP')
            
            # Try to get schema from Company model (dynamic, no hardcoding)
            if selected_db_key:
                try:
                    from FieldAdvisoryService.models import Company
                    # Try to find company by Company_name (display name) first
                    company = Company.objects.filter(Company_name=selected_db_key, is_active=True).first()
                    # If not found, try by schema name field
                    if not company:
                        company = Company.objects.filter(name=selected_db_key, is_active=True).first()
                    # If found, use its schema name
                    if company:
                        schema = company.name
                        print(f"[HANA] Found schema from Company model: {company.Company_name} -> {schema}")
                except Exception as company_err:
                    print(f"[HANA] Could not lookup company from model: {company_err}")

        # Strip quotes if present
        schema = schema.strip('"\'')
        
        # Connection parameters
        host = os.environ.get('HANA_HOST', '').strip()
        port = int(os.environ.get('HANA_PORT', 30015))
        user = os.environ.get('HANA_USER', '').strip()
        password = os.environ.get('HANA_PASSWORD', '').strip()
        encrypt = (os.environ.get('HANA_ENCRYPT') or '').strip().lower() in ('true','1','yes')
        ssl_validate = (os.environ.get('HANA_SSL_VALIDATE') or '').strip().lower() in ('true','1','yes')
        
        if not host or not user:
            return None
        
        # Connect
        kwargs = {'address': host, 'port': port, 'user': user, 'password': password}
        if encrypt:
            kwargs['encrypt'] = True
            kwargs['sslValidateCertificate'] = ssl_validate
        conn = dbapi.connect(**kwargs)
        
        # Set schema
        cursor = conn.cursor()
        cursor.execute(f'SET SCHEMA "{schema}"')
        cursor.close()
        
        return conn
    except Exception as e:
        print(f"Error connecting to HANA: {e}")
        return None

@require_http_methods(["GET"])
def api_warehouse_for_item(request):
    """Get warehouses for a specific item"""
    item_code = request.GET.get('item_code')
    if not item_code:
        return JsonResponse({'error': 'item_code parameter required'}, status=400)
    
    try:
        db = get_hana_connection(request)
        if not db:
            return JsonResponse({'error': 'Database connection failed'}, status=500)
        
        warehouses = hana_connect.warehouse_for_item(db, item_code)
        db.close()
        return JsonResponse({'warehouses': warehouses})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_customer_address(request):
    """Get address for a specific customer"""
    card_code = request.GET.get('card_code')
    if not card_code:
        return JsonResponse({'error': 'card_code parameter required'}, status=400)
    
    try:
        db = get_hana_connection(request)
        if not db:
            return JsonResponse({'error': 'Database connection failed'}, status=500)
        
        # Query for address
        query = f"""
        SELECT T0."Address", T0."Street" 
        FROM CRD1 T0 
        WHERE T0."CardCode" = '{card_code}'
        """
        cursor = db.cursor()
        cursor.execute(query)
        result = cursor.fetchone()
        
        db.close()
        
        if result:
            return JsonResponse({
                'address': result[0],
                'street': result[1]
            })
        else:
            return JsonResponse({'address': '', 'street': ''})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_policy_link(request):
    """Get policy link for a project"""
    project_code = request.GET.get('project_code')
    if not project_code:
        return JsonResponse({'error': 'project_code parameter required'}, status=400)
    
    try:
        db = get_hana_connection(request)
        if not db:
            return JsonResponse({'error': 'Database connection failed'}, status=500)
        
        query = f"""
        SELECT a."DocEntry" 
        FROM "@PL1" a 
        WHERE a."U_proj" = '{project_code}'
        """
        cursor = db.cursor()
        cursor.execute(query)
        result = cursor.fetchone()
        
        db.close()
        return JsonResponse({'policy_link': result[0] if result else None})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_discounts(request):
    """Get discounts (U_AD, U_EXD) for policy, item, and policy link"""
    policy = request.GET.get('policy')
    item_code = request.GET.get('item_code')
    pl = request.GET.get('pl')
    
    if not all([policy, item_code, pl]):
        return JsonResponse({'error': 'policy, item_code, and pl parameters required'}, status=400)
    
    try:
        db = get_hana_connection(request)
        if not db:
            return JsonResponse({'error': 'Database connection failed'}, status=500)
        
        # Get U_AD
        query_ad = f"""
        SELECT b."U_ad" 
        FROM "@PL1" a 
        INNER JOIN "@PLR4" b ON a."DocEntry" = b."DocEntry" 
        WHERE a."U_proj" = '{policy}' 
        AND b."U_itc" = '{item_code}' 
        AND a."DocEntry" = {pl}
        """
        
        # Get U_EXD
        query_exd = f"""
        SELECT b."U_ed" 
        FROM "@PL1" a 
        INNER JOIN "@PLR4" b ON a."DocEntry" = b."DocEntry" 
        WHERE a."U_proj" = '{policy}' 
        AND b."U_itc" = '{item_code}' 
        AND a."DocEntry" = {pl}
        """
        
        cursor = db.cursor()
        cursor.execute(query_ad)
        result_ad = cursor.fetchone()
        
        cursor.execute(query_exd)
        result_exd = cursor.fetchone()
        
        db.close()
        
        return JsonResponse({
            'u_ad': float(result_ad[0]) if result_ad and result_ad[0] else 0.0,
            'u_exd': float(result_exd[0]) if result_exd and result_exd[0] else 0.0
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def api_project_balance(request):
    """Get project balance (U_BP)"""
    project_code = request.GET.get('project_code')
    if not project_code:
        return JsonResponse({'error': 'project_code parameter required'}, status=400)
    
    try:
        db = get_hana_connection(request)
        if not db:
            return JsonResponse({'error': 'Database connection failed'}, status=500)
        
        query = f"""
        SELECT IFNULL(SUM(a."Debit" - a."Credit"), 0) 
        FROM jdt1 a 
        WHERE a."Project" = '{project_code}'
        """
        cursor = db.cursor()
        cursor.execute(query)
        result = cursor.fetchone()
        
        db.close()
        
        return JsonResponse({'u_bp': float(result[0]) if result else 0.0})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@staff_member_required
@require_http_methods(["GET"])
def api_child_customers(request):
    """Get child customers for a parent customer (FatherCard)"""
    import logging
    logger = logging.getLogger(__name__)
    
    father_card = request.GET.get('father_card')
    database = request.GET.get('database')  # Support explicit database parameter
    search = (request.GET.get('search') or '').strip()
    page_param = (request.GET.get('page') or '1').strip()
    page_size_param = (request.GET.get('page_size') or '').strip()
    logger.info(f"API called for child customers: father_card={father_card}, database={database}, search={search}, page={page_param}, page_size={page_size_param}")
    
    if not father_card:
        logger.error("No father_card provided")
        return JsonResponse({'error': 'father_card parameter required'}, status=400)
    
    try:
        # Get selected database from session for logging
        selected_db = request.session.get('selected_db') if hasattr(request, 'session') else None
        logger.info(f"Selected DB from session: {selected_db}, explicit database: {database}")
        
        # Use explicit database parameter if provided, otherwise session
        db_key = database or selected_db
        db = get_hana_connection(request, db_key)
        if not db:
            logger.error("Database connection failed")
            return JsonResponse({'error': 'Database connection failed - HANA service unavailable', 'children': []}, status=200)
        
        # Verify current schema
        try:
            cursor = db.cursor()
            cursor.execute('SELECT CURRENT_SCHEMA FROM DUMMY')
            current_schema = cursor.fetchone()[0]
            cursor.close()
            logger.info(f"Connected to HANA schema: {current_schema}, fetching child customers for {father_card}")
        except Exception as e:
            logger.warning(f"Could not verify current schema: {e}")
            logger.info(f"Database connected, fetching child customers for {father_card}")
        
        # Get child customers with optional search
        try:
            child_customers = hana_connect.child_card_code(db, father_card, search or None)
        except Exception as e:
            logger.error(f"Error calling child_card_code: {str(e)}")
            child_customers = []
        finally:
            try:
                db.close()
            except:
                pass

        # Return empty list if no children found
        if not child_customers:
            logger.info(f"No child customers found for {father_card}")
            return JsonResponse({
                'children': [],
                'page': 1,
                'page_size': 10,
                'num_pages': 0,
                'count': 0
            })

        # Pagination
        try:
            page_num = int(page_param) if page_param else 1
        except Exception:
            page_num = 1
        default_page_size = 10
        try:
            default_page_size = int(getattr(settings, 'REST_FRAMEWORK', {}).get('PAGE_SIZE', 10) or 10)
        except Exception:
            default_page_size = 10
        try:
            page_size = int(page_size_param) if page_size_param else default_page_size
        except Exception:
            page_size = default_page_size

        paginator = Paginator(child_customers or [], page_size)
        page_obj = paginator.get_page(page_num)
        logger.info(f"Found {paginator.count} child customers, returning page {page_obj.number}/{paginator.num_pages}")

        return JsonResponse({
            'children': list(page_obj.object_list),
            'page': page_obj.number,
            'page_size': page_size,
            'num_pages': paginator.num_pages,
            'count': paginator.count
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Error in api_child_customers: {str(e)}\n{error_trace}")
        return JsonResponse({'error': str(e), 'trace': error_trace, 'children': []}, status=200)

@staff_member_required
@require_http_methods(["GET"])
def api_customer_details(request):
    """Get full customer details (CardName, ContactPersonCode, FederalTaxID, PayToCode, Address)"""
    import logging
    logger = logging.getLogger(__name__)
    
    card_code = request.GET.get('card_code')
    database = request.GET.get('database')  # Support explicit database parameter
    logger.info(f"API called for customer details: card_code={card_code}, database={database}")
    
    if not card_code:
        logger.error("No card_code provided")
        return JsonResponse({'error': 'card_code parameter required'}, status=400)
    
    try:
        # Get selected database from session for logging
        selected_db = request.session.get('selected_db') if hasattr(request, 'session') else None
        logger.info(f"Selected DB from session: {selected_db}, explicit database: {database}")
        
        # Use explicit database parameter if provided, otherwise session
        db_key = database or selected_db
        db = get_hana_connection(request, db_key)
        if not db:
            logger.error("Database connection failed")
            return JsonResponse({'error': 'Database connection failed'}, status=500)
        
        # Verify current schema
        cursor = db.cursor()
        try:
            cursor.execute('SELECT CURRENT_SCHEMA FROM DUMMY')
            current_schema = cursor.fetchone()[0]
            logger.info(f"Connected to HANA schema: {current_schema}, fetching details for {card_code}")
        except Exception as e:
            logger.warning(f"Could not verify current schema: {e}")
            logger.info(f"Database connected, fetching details for {card_code}")
        
        # Get customer basic info
        customer_query = """
        SELECT 
            T0."CardName", 
            T0."CntctPrsn",
            T0."LicTradNum",
            T0."BillToDef",
            T0."Address"
        FROM OCRD T0 
        WHERE T0."CardCode" = ?
        """
        cursor.execute(customer_query, (card_code,))
        customer_result = cursor.fetchone()
        
        logger.info(f"Customer query result: {customer_result}")
        
        if not customer_result:
            cursor.close()
            db.close()
            logger.warning(f"Customer not found: {card_code}")
            return JsonResponse({'error': f'Customer not found: {card_code}'}, status=404)
        
        # Get contact person code from OCPR table
        contact_code = None
        if customer_result[1]:  # If CntctPrsn (contact name) exists
            contact_query = """
            SELECT T0."CntctCode"
            FROM OCPR T0
            WHERE T0."CardCode" = ? AND T0."Name" = ?
            """
            cursor.execute(contact_query, (card_code, customer_result[1]))
            contact_result = cursor.fetchone()
            if contact_result:
                contact_code = int(contact_result[0])
        
        # Get billing address from CRD1
        address = customer_result[4]  # Use Address from OCRD first
        if not address or address.strip() == '':
            # Try to get formatted address from CRD1
            address_query = """
            SELECT 
                T0."Street"||', '||T2."Name" AS "Address"
            FROM CRD1 T0 
            INNER JOIN OCRD T1 ON T0."CardCode" = T1."CardCode" AND T0."Address" = T1."BillToDef"
            INNER JOIN OCRY T2 ON T1."Country" = T2."Code"
            WHERE T1."CardCode" = ?
            """
            cursor.execute(address_query, (card_code,))
            address_result = cursor.fetchone()
            if address_result:
                address = address_result[0]
        
        logger.info(f"Address query result: {address}")
        
        cursor.close()
        db.close()
        
        # Parse the results safely
        card_name = customer_result[0] if customer_result[0] else ''
        federal_tax_id = customer_result[2] if customer_result[2] else ''
        
        # Handle pay_to_code (BillToDef) - might be string or int
        pay_to_code = None
        if customer_result[3]:
            try:
                pay_to_code = int(customer_result[3]) if customer_result[3] else None
            except (ValueError, TypeError):
                # If it's already a string address code, keep it
                pay_to_code = customer_result[3]
        
        response_data = {
            'card_name': card_name,
            'contact_person_code': contact_code if contact_code else '',
            'federal_tax_id': federal_tax_id,
            'pay_to_code': pay_to_code if pay_to_code else '',
            'address': address if address else ''
        }
        
        logger.info(f"Returning response: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Error in api_customer_details: {str(e)}\n{error_trace}")
        return JsonResponse({'error': str(e), 'trace': error_trace}, status=500)