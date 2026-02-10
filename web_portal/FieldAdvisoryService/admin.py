from django.contrib import admin
from web_portal.admin import admin_site
from django.contrib import messages
import json
import logging
import os
from .models import Dealer, MeetingSchedule, MeetingScheduleAttendance, SalesOrder, SalesOrderLine, SalesOrderAttachment
from .models import DealerRequest , Company, Region, Zone, Territory
from sap_integration.sap_client import SAPClient
from django import forms
from django.utils import timezone
from sap_integration import hana_connect
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db import models
from django.contrib.admin import widgets as admin_widgets
from datetime import datetime

# admin.site.register(Dealer)
class MeetingScheduleAttendanceInline(admin.TabularInline):
    model = MeetingScheduleAttendance
    extra = 1
    fields = ('farmer', 'farmer_name', 'contact_number', 'acreage', 'crop')
    readonly_fields = ()
    autocomplete_fields = ['farmer']
    
    class Media:
        js = ('admin/js/farmer_autocomplete.js',)

def export_meeting_schedule_to_excel(modeladmin, request, queryset):
    """Export selected Meeting Schedules to Excel with attendance details"""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Meeting Schedules"
    
    # Add export date/time at the top
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Field Advisory Meeting Schedule Export - {export_time}"
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells('A1:D1')
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    info_label_font = Font(bold=True, size=10, color="1F4E78")
    
    odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    current_row = 3
    
    for meeting in queryset:
        # Meeting Information Section
        info_fields = [
            ('Meeting ID', meeting.meeting_id),
            ('FSM Name', meeting.fsm_name),
            ('Date', meeting.date.strftime('%Y-%m-%d %H:%M') if meeting.date else ''),
            ('Region', meeting.region.name if meeting.region else ''),
            ('Zone', meeting.zone.name if meeting.zone else ''),
            ('Territory', meeting.territory.name if meeting.territory else ''),
            ('Location', meeting.location),
            ('Total Attendees', meeting.total_attendees),
            ('Confirmed Attendees', meeting.confirmed_attendees),
            ('Min Farmers Required', meeting.min_farmers_required),
            ('ZM Present', 'Yes' if meeting.presence_of_zm else 'No'),
            ('RSM Present', 'Yes' if meeting.presence_of_rsm else 'No'),
            ('Key Topics Discussed', meeting.key_topics_discussed),
            ('Feedback', meeting.feedback_from_attendees or ''),
            ('Suggestions', meeting.suggestions_for_future or '')
        ]
        
        # Write meeting information as key-value pairs
        for label, value in info_fields:
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.fill = info_label_fill
            label_cell.font = info_label_font
            label_cell.border = thin_border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            value_cell = ws.cell(row=current_row, column=2, value=value)
            value_cell.border = thin_border
            value_cell.alignment = data_alignment
            ws.merge_cells(f'B{current_row}:D{current_row}')
            
            current_row += 1
        
        current_row += 1  # Empty row
        
        # Attendee table headers
        attendee_headers = ['Attendee Name', 'Contact Number', 'Acreage', 'Crop']
        for col_num, header in enumerate(attendee_headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Write attendees
        attendees = meeting.attendees.all()
        if attendees.exists():
            for idx, attendee in enumerate(attendees):
                is_odd_row = idx % 2 == 0
                row_fill = odd_row_fill if is_odd_row else even_row_fill
                
                attendee_data = [
                    attendee.farmer_name,
                    attendee.contact_number,
                    attendee.acreage,
                    attendee.crop
                ]
                
                for col_num, value in enumerate(attendee_data, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = data_alignment
                
                current_row += 1
        else:
            # No attendees message
            no_attendee_cell = ws.cell(row=current_row, column=1, value="No attendees")
            no_attendee_cell.font = Font(italic=True, color="999999")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
        
        current_row += 2  # Space before next meeting
    
    # Adjust column widths
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if col_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=meeting_schedules.xlsx'
    wb.save(response)
    
    return response

export_meeting_schedule_to_excel.short_description = "Export selected to Excel"

@admin.register(MeetingSchedule, site=admin_site)
class MeetingScheduleAdmin(admin.ModelAdmin):
    inlines = [MeetingScheduleAttendanceInline]
    list_display = [
        'meeting_id',
        'fsm_name',
        'formatted_date',
        'region',
        'zone',
        'territory',
        'location',
        'total_attendees',
        'presence_of_zm',
        'presence_of_rsm'
    ]
    search_fields = [
        'meeting_id',
        'fsm_name',
        'region__name',
        'zone__name',
        'territory__name',
        'location',
        'key_topics_discussed'
    ]
    list_filter = [
        'region',
        'zone',
        'territory',
        ('date', admin.DateFieldListFilter),
        'presence_of_zm',
        'presence_of_rsm'
    ]
    ordering = ['-id']
    actions = [export_meeting_schedule_to_excel]
    readonly_fields = ['meeting_id']
    
    # Configure form to show datetime input with separate date and time fields
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'date':
            kwargs['widget'] = admin_widgets.AdminSplitDateTime()
        return super().formfield_for_dbfield(db_field, request, **kwargs)
    
    def formatted_date(self, obj):
        """Display date with time"""
        if obj.date:
            return obj.date.strftime('%Y-%m-%d %H:%M')
        return '-'
    formatted_date.short_description = 'Date & Time'
    formatted_date.admin_order_field = 'date'


def _load_env_file(path: str) -> None:
    """Load environment variables from .env file"""
    try:
        if os.path.isfile(path) and os.access(path, os.R_OK):
            with open(path, 'r', encoding='utf-8') as f:
                for line in f:
                    s = line.strip()
                    if s == '' or s.startswith('#') or '=' not in s:
                        continue
                    k, v = s.split('=', 1)
                    k = k.strip()
                    v = v.strip()
                    if v != '' and ((v[0] == '"' and v[-1] == '"') or (v[0] == "'" and v[-1] == "'")):
                        v = v[1:-1]
                    if k != '' and not os.environ.get(k):
                        os.environ[k] = v
    except Exception:
        pass


def get_hana_connection(selected_db_key=None):
    """Get HANA database connection honoring the selected DB (session/global dropdown)."""
    try:
        # Load .env file
        _load_env_file(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), '.env'))
        
        from hdbcli import dbapi
        from preferences.models import Setting

        # Resolve schema based on selected_db_key
        # Priority: 1. Direct lookup from Company model, 2. Database settings, 3. Fallback
        schema = None
        
        # FIRST: Try to find schema from Company model (most reliable and dynamic)
        if selected_db_key:
            try:
                # Try to find company by Company_name (display name) first
                company = Company.objects.filter(Company_name=selected_db_key, is_active=True).first()
                
                # If not found, try by schema name field
                if not company:
                    company = Company.objects.filter(name=selected_db_key, is_active=True).first()
                
                # If found, use its schema name
                if company:
                    schema = company.name
                    print(f"[HANA] Found company from model: {company.Company_name} -> schema: {schema}")
            except Exception as e:
                print(f"[HANA] Error looking up company from model: {e}")
        
        # SECOND: Try database settings if no schema found yet
        if not schema:
            try:
                db_setting = Setting.objects.filter(slug='SAP_COMPANY_DB').first()
                raw_value = getattr(db_setting, 'value', None) if db_setting else None
                db_options = {}

                # Extract mapping from stored value
                if isinstance(raw_value, dict):
                    db_options = raw_value
                elif isinstance(raw_value, str):
                    try:
                        import json
                        parsed = json.loads(raw_value)
                        if isinstance(parsed, dict):
                            db_options = parsed
                        else:
                            schema = str(parsed)
                    except Exception:
                        schema = raw_value

                cleaned = {}
                for k, v in db_options.items():
                    clean_key = str(k).strip().strip('"').strip("'")
                    clean_val = str(v).strip().strip('"').strip("'")
                    cleaned[clean_key] = clean_val
                db_options = cleaned

                if db_options and selected_db_key:
                    schema = db_options.get(selected_db_key)
                    print(f"[HANA] From settings: {selected_db_key} -> {schema}")
                elif db_options and not schema:
                    schema = list(db_options.values())[0]
                    print(f"[HANA] From settings (first): {schema}")
                elif raw_value and not isinstance(raw_value, dict):
                    schema = str(raw_value).strip().strip('"').strip("'")
            except Exception as e:
                print(f"[HANA] Error getting schema from settings: {e}")
        
        # LAST: Fallback to environment variable or default
        if not schema:
            schema = os.environ.get('SAP_COMPANY_DB', '4B-BIO_APP')
            print(f"[HANA] Fallback to env/default: {schema}")

        # Strip quotes if present
        schema = schema.strip("'\"")
        
        print(f"[HANA] FINAL SCHEMA TO USE: {schema}")
        
        # Connection parameters
        host = os.environ.get('HANA_HOST', '').strip()
        port = int(os.environ.get('HANA_PORT', 30015))
        user = os.environ.get('HANA_USER', '').strip()
        password = os.environ.get('HANA_PASSWORD', '').strip()
        
        if not host or not user:
            print(f"Missing HANA credentials: host={bool(host)}, user={bool(user)}")
            return None
        
        # Connect
        print(f"Connecting to HANA: {host}:{port} as {user}, schema={schema}")
        conn = dbapi.connect(address=host, port=port, user=user, password=password)

        # Set schema and verify
        cursor = conn.cursor()
        cursor.execute(f'SET SCHEMA "{schema}"')
        try:
            cursor.execute('SELECT CURRENT_SCHEMA FROM DUMMY')
            row = cursor.fetchone()
            current_schema = None
            if row:
                try:
                    current_schema = row[0]
                except Exception:
                    current_schema = None
            print(f"HANA current schema: {current_schema}")
        except Exception as e:
            print(f"Warning: could not verify current schema: {e}")
        finally:
            try:
                cursor.close()
            except Exception:
                pass

        print("HANA connection successful")
        return conn
    except Exception as e:
        print(f"Error connecting to HANA: {e}")
        import traceback
        traceback.print_exc()
        return None


class _CompanySessionResolver:
    def _get_selected_company(self, request):
        """Get the selected company dynamically from session without hardcoded mappings."""
        try:
            db_key = request.session.get('selected_db') or request.GET.get('company_db')
            if not db_key:
                # Fallback to first active company if no selection
                return Company.objects.filter(is_active=True).first()
            
            # Try to find company by Company_name (display name) first
            comp = Company.objects.filter(Company_name=db_key, is_active=True).first()
            
            # If not found, try by schema name field
            if not comp:
                comp = Company.objects.filter(name=db_key, is_active=True).first()
            
            # If still not found, try partial match on Company_name
            if not comp:
                comp = Company.objects.filter(Company_name__icontains=db_key, is_active=True).first()
            
            # If still not found, try partial match on schema name
            if not comp:
                comp = Company.objects.filter(name__icontains=db_key, is_active=True).first()
            
            return comp
        except Exception as e:
            print(f"Error getting selected company: {e}")
            return None


class SalesOrderForm(forms.ModelForm):
    """Custom form with LOV dropdowns for SAP data"""
    
    class Meta:
        model = SalesOrder
        fields = '__all__'
        widgets = {
            'card_code': forms.Select(attrs={'class': 'sap-customer-lov'}),
            'u_s_card_code': forms.Select(attrs={'class': 'sap-child-customer-lov', 'style': 'width: 400px;'}),
            'address': forms.Textarea(attrs={'rows': 3}),
            'comments': forms.Textarea(attrs={'rows': 3}),
            'sap_error': forms.Textarea(attrs={'rows': 3}),
            'sap_response_json': forms.Textarea(attrs={
                'rows': 18,
                'class': 'sap-json-viewer',
                'placeholder': 'SAP API response will appear here after posting...',
                'style': (
                    'height: 400px !important; '
                    'max-height: 400px !important; '
                    'overflow: scroll !important; '
                    'font-family: Consolas, Monaco, "Courier New", monospace !important; '
                    'font-size: 13px !important; '
                    'line-height: 1.6 !important; '
                    'background-color: #f8f9fa !important; '
                    'border: 2px solid #dee2e6 !important; '
                    'border-radius: 6px !important; '
                    'padding: 12px !important; '
                    'white-space: pre !important; '
                    'display: block !important; '
                    'width: 100% !important;'
                )
            }),
        }
    
    def __init__(self, *args, **kwargs):
        # Capture request to honor selected database
        self.request = kwargs.pop('request', None)
        selected_db_key = kwargs.pop('selected_db_key', None)
        
        # Try multiple sources to get the selected DB (in priority order)
        if not selected_db_key and self.request:
            # 1. Try URL parameter
            selected_db_key = self.request.GET.get('company_db')
            # 2. Try session
            if not selected_db_key and hasattr(self.request, 'session'):
                selected_db_key = self.request.session.get('selected_db')

        super().__init__(*args, **kwargs)
        
        print("=== FORM INITIALIZATION DEBUG ===")
        print(f"[FORM] Selected DB key: {selected_db_key}")
        print(f"[FORM] Request method: {getattr(self.request, 'method', 'UNKNOWN')}")
        print(f"[FORM] Request GET params: {dict(self.request.GET) if self.request else 'No request'}")
        print(f"[FORM] Available fields: {list(self.fields.keys())}")
        print(f"u_s_card_code field type: {type(self.fields.get('u_s_card_code'))}")
        print(f"u_s_card_code widget type: {type(self.fields.get('u_s_card_code').widget) if 'u_s_card_code' in self.fields else 'N/A'}")
        
        # Populate customer dropdown
        try:
            db = get_hana_connection(selected_db_key)
            if db:
                try:
                    customers = hana_connect.customer_codes_all(db, limit=2000)
                except Exception:
                    customers = hana_connect.customer_lov(db)
                # Log a preview of customer codes to confirm correct company
                try:
                    preview = [c.get('CardCode') for c in (customers or [])][:5]
                    print(f"Customer code preview: {preview}")
                except Exception:
                    pass
                customer_choices = [('', '--- Select Customer ---')] + [
                    (c['CardCode'], f"{c['CardCode']} - {c['CardName']}") 
                    for c in customers
                ]
                self.fields['card_code'].widget = forms.Select(choices=customer_choices, attrs={'class': 'sap-customer-lov', 'style': 'width: 400px;'})
                
                # If we have a card_code in the instance, load child customers
                if self.instance and self.instance.card_code:
                    print(f"Loading child customers for existing card_code: {self.instance.card_code}")
                    try:
                        child_customers = hana_connect.child_card_code(db, self.instance.card_code)
                        print(f"Found {len(child_customers)} child customers")
                        child_choices = [('', '--- Select Child Customer ---')] + [
                            (c['CardCode'], f"{c['CardCode']} - {c['CardName']}") 
                            for c in child_customers
                        ]
                        # Set choices on the actual u_s_card_code field
                        if 'u_s_card_code' in self.fields:
                            self.fields['u_s_card_code'].widget.choices = child_choices
                            print(f"Set {len(child_choices)} choices on u_s_card_code widget")
                    except Exception as e:
                        print(f"Error loading child customers: {e}")
                        import traceback
                        traceback.print_exc()
                else:
                    print("No card_code in instance, setting default message")
                    # Set initial empty choices - will be populated via JavaScript when parent is selected
                    if 'u_s_card_code' in self.fields:
                        self.fields['u_s_card_code'].widget.choices = [('', '--- Select Parent Customer First ---')]
                        print("Set default message on u_s_card_code")
                
                db.close()
            else:
                print("Failed to get HANA connection for customers")
        except Exception as e:
            print(f"Error loading customers: {e}")
            import traceback
            traceback.print_exc()
        
        # Set help text on child customer field
        if 'u_s_card_code' in self.fields:
            self.fields['u_s_card_code'].help_text = "Select parent customer first to load child customers"
            print(f"Child customer field widget: {type(self.fields['u_s_card_code'].widget)}")
            print(f"Child customer field widget attrs: {self.fields['u_s_card_code'].widget.attrs}")
            print(f"Child customer field is required: {self.fields['u_s_card_code'].required}")
            print(f"Child customer field is disabled: {self.fields['u_s_card_code'].disabled}")
        else:
            print("WARNING: u_s_card_code field not found in form fields!")
        
        if 'u_s_card_name' in self.fields:
            print(f"Child customer NAME field exists: {type(self.fields['u_s_card_name'].widget)}")
        
        print(f"\n=== FINAL FORM FIELDS ===")
        print(f"Total fields in form: {len(self.fields)}")
        print(f"Field names: {list(self.fields.keys())}")
        # Add help text for fields - NO styling, NO readonly - let JavaScript handle it
        if 'card_code' in self.fields:
            self.fields['card_code'].help_text = "Select customer from SAP"
        if 'card_name' in self.fields:
            self.fields['card_name'].help_text = "Auto-filled from customer (editable if needed)"
        if 'contact_person_code' in self.fields:
            self.fields['contact_person_code'].help_text = "Auto-filled from customer"
        if 'federal_tax_id' in self.fields:
            self.fields['federal_tax_id'].help_text = "Auto-filled from customer (NTN)"
    
    def clean(self):
        """Override clean to ensure all customer fields are captured"""
        cleaned_data = super().clean()
        import logging
        logger = logging.getLogger(__name__)
        
        # Log all customer-related fields
        card_code = cleaned_data.get('card_code')
        card_name = cleaned_data.get('card_name')
        contact_person = cleaned_data.get('contact_person_code')
        federal_tax = cleaned_data.get('federal_tax_id')
        pay_to = cleaned_data.get('pay_to_code')
        address = cleaned_data.get('address')
        
        logger.info(f"[FORM CLEAN] === Customer Fields ===")
        logger.info(f"[FORM CLEAN] card_code: '{card_code}' (type: {type(card_code)})")
        logger.info(f"[FORM CLEAN] card_name: '{card_name}' (type: {type(card_name)})")
        logger.info(f"[FORM CLEAN] contact_person_code: '{contact_person}'")
        logger.info(f"[FORM CLEAN] federal_tax_id: '{federal_tax}'")
        logger.info(f"[FORM CLEAN] pay_to_code: '{pay_to}'")
        logger.info(f"[FORM CLEAN] address: '{address}'")
        logger.info(f"[FORM CLEAN] All cleaned_data keys: {list(cleaned_data.keys())}")
        
        # CRITICAL: Ensure empty strings are preserved (not converted to None)
        if card_code == '':
            logger.warning("[FORM CLEAN] ⚠️ card_code is empty string!")
        if card_name == '':
            logger.warning("[FORM CLEAN] ⚠️ card_name is empty string!")
        
        return cleaned_data


class SalesOrderLineInlineForm(forms.ModelForm):
    """Custom form for sales order lines with LOV dropdowns"""
    
    class Meta:
        model = SalesOrderLine
        fields = '__all__'
        widgets = {
            'item_code': forms.Select(attrs={'class': 'sap-item-lov'}),
            'warehouse_code': forms.Select(attrs={'class': 'sap-warehouse-lov'}),
            'vat_group': forms.Select(attrs={'class': 'sap-tax-lov'}),
            'project_code': forms.Select(attrs={'class': 'sap-project-lov'}),
            'u_crop': forms.Select(attrs={'class': 'sap-crop-lov'}),
            'u_policy': forms.Select(attrs={'class': 'sap-policy-lov'}),
        }
    
    def __init__(self, *args, **kwargs):
        # Capture request to honor selected database
        self.request = kwargs.pop('request', None)
        selected_db_key = kwargs.pop('selected_db_key', None)
        if not selected_db_key and self.request and hasattr(self.request, 'session'):
            selected_db_key = self.request.session.get('selected_db')

        super().__init__(*args, **kwargs)
        
        try:
            db = get_hana_connection(selected_db_key)
            if db:
                # Populate item dropdown
                items = hana_connect.item_lov(db)
                item_choices = [('', '--- Select Item ---')] + [
                    (item['ItemCode'], f"{item['ItemCode']} - {item['ItemName']}") 
                    for item in items[:500]  # Limit to first 500 items for performance
                ]
                if 'item_code' in self.fields:
                    self.fields['item_code'].widget = forms.Select(choices=item_choices, attrs={'class': 'sap-item-lov', 'style': 'width: 400px;'})
                    print(f"DEBUG: Item LOV loaded with {len(item_choices)} items")
                
                # Populate tax group dropdown
                try:
                    tax_codes = hana_connect.sales_tax_codes(db)
                    
                    # Validate tax codes to prevent corruption
                    valid_tax_choices = []
                    for tax in tax_codes:
                        try:
                            code = str(tax.get('Code', '')).strip()
                            name = str(tax.get('Name', '')).strip()
                            rate = tax.get('Rate', 0)
                            
                            # Skip if code is empty or contains invalid characters
                            if code and '\ufffd' not in code and '\xff' not in code:
                                valid_tax_choices.append((code, f"{code} - {name} ({rate}%)"))
                        except Exception as e:
                            print(f"WARNING: Skipping corrupted tax code: {e}")
                            continue
                    
                    # Use default 'SE' if no valid codes found
                    if not valid_tax_choices:
                        print("WARNING: No valid tax codes loaded from SAP, using defaults")
                        valid_tax_choices = [
                            ('SE', 'SE - Standard Exempted (0%)'),
                            ('AT1', 'AT1 - Taxable (17%)'),
                        ]
                    
                    tax_choices = [('', '--- Select Tax Code ---')] + valid_tax_choices
                    
                    if 'vat_group' in self.fields:
                        self.fields['vat_group'].widget = forms.Select(choices=tax_choices, attrs={'class': 'sap-tax-lov', 'style': 'width: 300px;'})
                        self.fields['vat_group'].initial = 'SE'  # Default to Standard Exempted
                        print(f"DEBUG: Tax LOV loaded with {len(valid_tax_choices)} valid tax codes (set default to 'SE')")
                    else:
                        print("DEBUG: vat_group field not in form fields (expected for inline forms)")
                
                except Exception as e:
                    print(f"ERROR loading tax codes: {e}")
                    # Fallback to safe defaults
                    if 'vat_group' in self.fields:
                        tax_choices = [
                            ('', '--- Select Tax Code ---'),
                            ('SE', 'SE - Standard Exempted'),
                            ('AT1', 'AT1 - Standard Taxable (17%)'),
                        ]
                        self.fields['vat_group'].widget = forms.Select(choices=tax_choices, attrs={'class': 'sap-tax-lov', 'style': 'width: 300px;'})
                        self.fields['vat_group'].initial = 'SE'
                
                # Populate project dropdown
                projects = hana_connect.projects_lov(db)
                project_choices = [('', '--- Select Project ---')] + [
                    (proj['PrjCode'], f"{proj['PrjCode']} - {proj['PrjName']}") 
                    for proj in projects[:200]  # Limit to first 200 projects
                ]
                if 'project_code' in self.fields:
                    self.fields['project_code'].widget = forms.Select(choices=project_choices, attrs={'class': 'sap-project-lov', 'style': 'width: 350px;'})
                    print(f"DEBUG: Project LOV loaded with {len(project_choices)} projects")
                
                # Populate crop dropdown
                try:
                    crops = hana_connect.crop_lov(db)
                    print(f"DEBUG: Loaded {len(crops) if crops else 0} crops from HANA")
                    if crops:
                        print(f"DEBUG: First crop sample: {crops[0] if len(crops) > 0 else 'None'}")
                    crop_choices = [('', '--- Select Crop ---')] + [
                        (crop['Code'], f"{crop['Code']} - {crop['Name']}") 
                        for crop in crops
                    ]
                    print(f"DEBUG: Created {len(crop_choices)} crop choices")
                    if 'u_crop' in self.fields:
                        self.fields['u_crop'].widget = forms.Select(choices=crop_choices, attrs={'class': 'sap-crop-lov', 'style': 'width: 250px;'})
                        print(f"DEBUG: Crop widget assigned successfully to u_crop field")
                    else:
                        print("DEBUG: u_crop field not found in form fields")
                except Exception as e:
                    print(f"ERROR loading crops: {e}")
                    import traceback
                    traceback.print_exc()
                    # Set empty choices on error
                    if 'u_crop' in self.fields:
                        self.fields['u_crop'].widget = forms.Select(choices=[('', '--- No Crops Available ---')], attrs={'class': 'sap-crop-lov', 'style': 'width: 250px;'})
                
                # Load all warehouses (we'll show all available warehouses)
                # Note: In a real scenario, you might want to filter by item, but for now showing all
                try:
                    # Get a sample warehouse list - you can enhance this to load all warehouses
                    # For now, we'll create a basic list
                    warehouse_choices = [
                        ('', '--- Select Warehouse ---'),
                        ('WH01', 'WH01 - Main Warehouse'),
                        ('WH02', 'WH02 - Secondary Warehouse'),
                        ('WH03', 'WH03 - Regional Warehouse'),
                        ('WH04', 'WH04 - Distribution Center'),
                        ('WH05', 'WH05 - Storage Facility'),
                        ('WH06', 'WH06 - Branch Warehouse'),
                    ]
                    self.fields['warehouse_code'].widget = forms.Select(
                        choices=warehouse_choices,
                        attrs={'class': 'sap-warehouse-lov', 'style': 'width: 250px;'}
                    )
                except Exception as e:
                    print(f"Error loading warehouses: {e}")
                
                # Policy dropdown will be populated entirely by JavaScript
                # This allows dynamic loading when customer changes
                # Initialize with placeholder that indicates JS will handle it
                if 'u_policy' in self.fields:
                    self.fields['u_policy'].widget = forms.Select(
                        choices=[('', '--- Select Policy ---')],
                        attrs={'class': 'sap-policy-lov', 'style': 'width: 400px;'}
                    )
                    print(f"DEBUG: Policy dropdown initialized with placeholder (JS will populate)")
                
                db.close()
                print(f"Loaded {len(item_choices)-1} items, {len(tax_choices)-1} tax codes, {len(project_choices)-1} projects, {len(crop_choices)-1} crops")
            else:
                print("Failed to get HANA connection for LOVs")
        except Exception as e:
            print(f"Error loading LOVs: {e}")
            import traceback
            traceback.print_exc()
        
        # Set readonly fields only if they exist in the form
        if 'item_description' in self.fields:
            self.fields['item_description'].widget.attrs['readonly'] = True
        if 'measure_unit' in self.fields:
            self.fields['measure_unit'].widget.attrs['readonly'] = True
        if 'uom_code' in self.fields:
            self.fields['uom_code'].widget.attrs['readonly'] = True


class SalesOrderLineInline(admin.TabularInline):
    model = SalesOrderLine
    form = SalesOrderLineInlineForm
    extra = 1
    # Reorganized fields - hiding line_num, uom_entry, uom_code, vat_group, tax_percentage_per_row
    fields = (
        'u_pl',
        'u_policy',
        'item_code', 
        'item_description', 
        'measure_unit',
        'u_crop',
        'warehouse_code', 
        'quantity', 
        'unit_price',
        'discount_percent',
    )
    readonly_fields = ('item_description', 'measure_unit')
    # Add CSS classes for styling
    classes = ['collapse', 'open']
    
    def get_formset(self, request, obj=None, **kwargs):
        """Customize the formset with proper labels and help text"""
        base_formset = super().get_formset(request, obj, **kwargs)
        selected_db_key = request.session.get('selected_db') if hasattr(request, 'session') else None

        class RequestAwareFormSet(base_formset):
            def _construct_form(self2, i, **kw):
                kw['request'] = request
                kw['selected_db_key'] = selected_db_key
                return super()._construct_form(i, **kw)

        formset = RequestAwareFormSet
        # Update labels and help text only for fields that exist
        if hasattr(formset.form, 'base_fields'):
            if 'u_pl' in formset.form.base_fields:
                formset.form.base_fields['u_pl'].label = 'Policy Link'
            if 'u_policy' in formset.form.base_fields:
                formset.form.base_fields['u_policy'].label = 'Policy'
            if 'item_code' in formset.form.base_fields:
                formset.form.base_fields['item_code'].label = 'Item No'
                formset.form.base_fields['item_code'].help_text = "Select item from catalog"
            if 'u_crop' in formset.form.base_fields:
                formset.form.base_fields['u_crop'].label = 'Crop'
            if 'warehouse_code' in formset.form.base_fields:
                formset.form.base_fields['warehouse_code'].label = 'Warehouse'
                formset.form.base_fields['warehouse_code'].help_text = "Select warehouse (filtered by item)"
            if 'quantity' in formset.form.base_fields:
                formset.form.base_fields['quantity'].label = 'Quantity'
                formset.form.base_fields['quantity'].help_text = "Enter quantity"
            if 'unit_price' in formset.form.base_fields:
                formset.form.base_fields['unit_price'].label = 'Unit Price'
                formset.form.base_fields['unit_price'].help_text = "Unit price from policy"
            if 'discount_percent' in formset.form.base_fields:
                formset.form.base_fields['discount_percent'].label = 'Discount %'
                formset.form.base_fields['discount_percent'].help_text = "Discount percentage"
                formset.form.base_fields['discount_percent'].initial = 0.0
        return formset


@admin.register(SalesOrder, site=admin_site)
class SalesOrderAdmin(admin.ModelAdmin, _CompanySessionResolver):
    form = SalesOrderForm
    list_display = ('id', 'portal_order_id', 'card_code', 'card_name', 'doc_date', 'status', 'is_posted_to_sap', 'sap_doc_num', 'created_at')
    list_filter = ('status', 'is_posted_to_sap', 'doc_date', 'created_at')
    search_fields = ('card_code', 'card_name', 'federal_tax_id', 'u_s_card_code', 'portal_order_id')
    readonly_fields = ('portal_order_id', 'current_database', 'created_at', 'sap_doc_entry', 'sap_doc_num', 'sap_error', 'sap_response_display', 'posted_at', 'is_posted_to_sap', 'add_to_sap_button', 'series', 'doc_type', 'summery_type', 'doc_object_code')
    ordering = ['-id']
    
    fieldsets = (
        ('Database Selection', {
            'fields': ('current_database',),
            'description': 'Current active database. Use the global DB selector at the top-right to switch between companies.'
        }),
        ('Basic Information', {
            'fields': ('portal_order_id', 'staff', 'dealer', 'status', 'created_at')
        }),
        ('Document Dates', {
            'fields': ('doc_date', 'doc_due_date', 'tax_date'),
            'description': 'Posting Date, Delivery Date, and Document Date'
        }),
        ('Customer Information', {
            'fields': ('card_code', 'card_name', 'contact_person_code', 'federal_tax_id', 'pay_to_code', 'address'),
            'description': 'Customer Code, Name, Contact Person, and Billing Address (auto-filled based on Customer Code)'
        }),
        ('Sales Type & Portal User', {
            'fields': ('u_sotyp', 'u_usid'),
            'description': 'Sales Type: 01=Regular, 02=Advance | Portal User ID'
        }),
        ('Child Customer (Optional)', {
            'fields': ('u_s_card_code', 'u_s_card_name'),
            'description': 'Select a child customer if applicable. This dropdown populates automatically when you select a parent customer above.'
        }),
        ('Additional Comments', {
            'fields': ('comments',),
            'classes': ('collapse',)
        }),
        ('SAP Integration', {
            'fields': ('add_to_sap_button', 'is_posted_to_sap', 'sap_doc_entry', 'sap_doc_num', 'posted_at', 'sap_error', 'sap_response_display'),
        }),
    )
    
    inlines = [SalesOrderLineInline]
    
    actions = ['post_to_sap']

    class Media:
        css = {
            'all': (
                'admin/salesorder_loading.css',
                'FieldAdvisoryService/css/sap_response_json.css',
            )
        }
        js = (
            'admin/salesorder_loading.js',
            'FieldAdvisoryService/js/sap_json_scroll.js',
            'FieldAdvisoryService/salesorder_customer.js',
            'FieldAdvisoryService/salesorder_policy.js',
        )
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        """Filter dealers based on selected company database."""
        if db_field.name == 'dealer':
            # Get the selected company from session
            selected_company = self._get_selected_company(request)
            
            if selected_company:
                # Filter dealers by the selected company
                kwargs['queryset'] = Dealer.objects.filter(
                    company=selected_company,
                    is_active=True
                ).select_related('user', 'company')
                print(f"[SalesOrderAdmin] Filtering dealers for company: {selected_company.Company_name} (schema: {selected_company.name})")
            else:
                # If no company selected, show all active dealers
                kwargs['queryset'] = Dealer.objects.filter(
                    is_active=True
                ).select_related('user', 'company')
                print("[SalesOrderAdmin] No company selected, showing all dealers")
        
        return super().formfield_for_foreignkey(db_field, request, **kwargs)
    
    def get_form(self, request, obj=None, **kwargs):
        base_form = super().get_form(request, obj, **kwargs)
        # Wrap form to inject request so dropdown data respects the selected DB
        class RequestAwareSalesOrderForm(base_form):
            def __init__(self2, *args, **kw):
                kw['request'] = request
                # Explicitly pass selected_db_key so form always knows the intended company
                kw['selected_db_key'] = request.GET.get('company_db') or request.session.get('selected_db')
                super().__init__(*args, **kw)

        form = RequestAwareSalesOrderForm

        # Update field labels (only for fields that are editable)
        if 'doc_date' in form.base_fields:
            form.base_fields['doc_date'].label = 'Posting Date'
        if 'doc_due_date' in form.base_fields:
            form.base_fields['doc_due_date'].label = 'Delivery Date'
        if 'tax_date' in form.base_fields:
            form.base_fields['tax_date'].label = 'Document Date'
        if 'card_code' in form.base_fields:
            form.base_fields['card_code'].label = 'Customer Code'
        if 'u_sotyp' in form.base_fields:
            form.base_fields['u_sotyp'].label = 'Sales Type'
        if 'u_usid' in form.base_fields:
            form.base_fields['u_usid'].label = 'Portal User ID'
        if 'u_s_card_code' in form.base_fields:
            form.base_fields['u_s_card_code'].label = 'Child Card Code'
        if 'u_s_card_name' in form.base_fields:
            form.base_fields['u_s_card_name'].label = 'Child Card Name'
        
        # Set initial values for dates to current date (only for new objects)
        if not obj:
            from django.utils import timezone
            today = timezone.now().date()
            if 'doc_date' in form.base_fields:
                form.base_fields['doc_date'].initial = today
            if 'doc_due_date' in form.base_fields:
                form.base_fields['doc_due_date'].initial = today
            if 'tax_date' in form.base_fields:
                form.base_fields['tax_date'].initial = today
            if 'u_sotyp' in form.base_fields:
                form.base_fields['u_sotyp'].initial = '01'
        
        return form
    
    def current_database(self, obj):
        """Display currently selected database/company"""
        from django.utils.html import format_html
        from django.contrib.admin.templatetags.admin_urls import admin_urlname
        from django.urls import reverse
        
        # Get selected database from various sources
        request = getattr(self, '_current_request', None)
        selected_db = None
        
        if request:
            # Try session first
            if hasattr(request, 'session'):
                selected_db = request.session.get('selected_db')
            # Try GET param
            if not selected_db and request.GET.get('company_db'):
                selected_db = request.GET.get('company_db')
        
        # Get display name from Company model dynamically
        display_name = selected_db or 'No database selected'
        
        if selected_db:
            try:
                # Try to get the company display name
                company = Company.objects.filter(Company_name=selected_db, is_active=True).first()
                if not company:
                    company = Company.objects.filter(name=selected_db, is_active=True).first()
                
                if company:
                    display_name = f"{company.Company_name} ({company.name})"
            except Exception as e:
                # If lookup fails, just use the selected_db value
                display_name = selected_db
        
        return format_html(
            '<div style=\"padding: 12px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); '
            'border-radius: 8px; color: white; font-weight: 600; font-size: 14px; '
            'box-shadow: 0 4px 6px rgba(0,0,0,0.1); display: inline-block;\">'
            '<svg xmlns=\"http://www.w3.org/2000/svg\" width=\"16\" height=\"16\" viewBox=\"0 0 24 24\" '
            'fill=\"none\" stroke=\"currentColor\" stroke-width=\"2\" style=\"vertical-align: middle; margin-right: 8px;\">'
            '<ellipse cx=\"12\" cy=\"5\" rx=\"9\" ry=\"3\"></ellipse>'
            '<path d=\"M21 12c0 1.66-4 3-9 3s-9-1.34-9-3\"></path>'
            '<path d=\"M3 5v14c0 1.66 4 3 9 3s9-1.34 9-3V5\"></path>'
            '</svg>'
            '{}'
            '</div>'
            '<p style=\"margin-top: 8px; color: #666; font-size: 12px;\">'
            '💡 To change database, use the <strong>DB selector</strong> at the top of the page.'
            '</p>',
            display_name
        )
    current_database.short_description = 'Active Company Database'
    
    def get_form(self, request, obj=None, **kwargs):
        # Store request for use in current_database method
        self._current_request = request
        
        base_form = super().get_form(request, obj, **kwargs)
        # Wrap form to inject request so dropdown data respects the selected DB
        class RequestAwareSalesOrderForm(base_form):
            def __init__(self2, *args, **kw):
                kw['request'] = request
                # Explicitly pass selected_db_key so form always knows the intended company
                kw['selected_db_key'] = request.GET.get('company_db') or request.session.get('selected_db')
                super().__init__(*args, **kw)

        form = RequestAwareSalesOrderForm

        # Update field labels (only for fields that are editable)
        if 'doc_date' in form.base_fields:
            form.base_fields['doc_date'].label = 'Posting Date'
        if 'doc_due_date' in form.base_fields:
            form.base_fields['doc_due_date'].label = 'Delivery Date'
        if 'tax_date' in form.base_fields:
            form.base_fields['tax_date'].label = 'Document Date'
        if 'card_code' in form.base_fields:
            form.base_fields['card_code'].label = 'Customer Code'
        if 'u_sotyp' in form.base_fields:
            form.base_fields['u_sotyp'].label = 'Sales Type'
        if 'u_usid' in form.base_fields:
            form.base_fields['u_usid'].label = 'Portal User ID'
        if 'u_s_card_code' in form.base_fields:
            form.base_fields['u_s_card_code'].label = 'Child Card Code'
        if 'u_s_card_name' in form.base_fields:
            form.base_fields['u_s_card_name'].label = 'Child Card Name'
        
        # Set initial values for dates to current date (only for new objects)
        if not obj:
            from django.utils import timezone
            today = timezone.now().date()
            if 'doc_date' in form.base_fields:
                form.base_fields['doc_date'].initial = today
            if 'doc_due_date' in form.base_fields:
                form.base_fields['doc_due_date'].initial = today
            if 'tax_date' in form.base_fields:
                form.base_fields['tax_date'].initial = today
            if 'u_sotyp' in form.base_fields:
                form.base_fields['u_sotyp'].initial = '01'
        
        return form
    
    def add_to_sap_button(self, obj):
        """Display a button to post this order to SAP"""
        from django.utils.html import format_html
        from django.urls import reverse
        import json
        
        if obj.pk:
            if obj.is_posted_to_sap:
                return format_html(
                    '<div style="padding: 10px; background: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px; color: #155724;">'
                    '<strong>✓ Posted to SAP</strong><br>'
                    'DocEntry: {}<br>DocNum: {}'
                    '</div>',
                    obj.sap_doc_entry, obj.sap_doc_num
                )
            else:
                # Check if customer is selected
                if not obj.card_code:
                    return format_html(
                        '<div style="padding: 12px; background: #fff3cd; border: 1px solid #ffc107; border-radius: 4px; color: #856404;">'
                        '<strong>⚠️ Customer Not Selected</strong><br>'
                        'Please select a <strong>Customer Code</strong> from the dropdown above and click <strong>SAVE</strong> before posting to SAP.'
                        '</div>'
                    )
                
                # Build SAP payload preview - COMPLETE structure
                try:
                    payload_lines = []
                    for line in obj.document_lines.all().order_by('line_num'):
                        # Sanitize tax code the same way as in actual SAP posting
                        import re
                        vat_group_preview = (line.vat_group or "SE").strip()
                        vat_group_preview = re.sub(r'[^A-Za-z0-9\-_]', '', vat_group_preview) or "SE"
                        
                        payload_lines.append({
                            "LineNum": line.line_num,
                            "ItemCode": line.item_code,
                            "ItemDescription": line.item_description,
                            "Quantity": float(line.quantity),
                            "DiscountPercent": float(line.discount_percent),
                            "WarehouseCode": line.warehouse_code,
                            # Service Layer expects TaxCode on document lines (sanitized)
                            "TaxCode": vat_group_preview,
                            "UnitsOfMeasurment": float(line.units_of_measurment),
                            "TaxPercentagePerRow": float(line.tax_percentage_per_row),
                            "UnitPrice": float(line.unit_price),
                            "UoMEntry": line.uom_entry,
                            "MeasureUnit": line.measure_unit,
                            "UoMCode": line.uom_code,
                            "ProjectCode": line.project_code,
                            "U_SD": float(line.u_sd),
                            "U_AD": float(line.u_ad),
                            "U_EXD": float(line.u_exd),
                            "U_zerop": float(line.u_zerop),
                            "U_pl": line.u_pl,
                            "U_BP": float(line.u_bp) if line.u_bp else None,
                            "U_policy": line.u_policy,
                            "U_focitem": line.u_focitem,
                            "U_crop": line.u_crop
                        })
                    
                    payload_preview = {
                        "Series": obj.series,
                        "DocType": obj.doc_type,
                        "DocDate": obj.doc_date.strftime('%Y-%m-%d') if obj.doc_date else None,
                        "DocDueDate": obj.doc_due_date.strftime('%Y-%m-%d') if obj.doc_due_date else None,
                        "TaxDate": obj.tax_date.strftime('%Y-%m-%d') if obj.tax_date else None,
                        "CardCode": obj.card_code or "",
                        "CardName": obj.card_name or "",
                        "ContactPersonCode": obj.contact_person_code,
                        "FederalTaxID": obj.federal_tax_id,
                        "PayToCode": obj.pay_to_code,
                        "Address": obj.address,
                        "DocCurrency": obj.doc_currency,
                        "DocRate": float(obj.doc_rate),
                        "Comments": obj.comments or "",
                        "SummeryType": obj.summery_type,
                        "DocObjectCode": obj.doc_object_code,
                        "U_sotyp": obj.u_sotyp,
                        "U_USID": obj.u_usid,
                        "U_SWJE": obj.u_swje,
                        "U_SECJE": obj.u_secje,
                        "U_CRJE": obj.u_crje,
                        "U_SCardCode": obj.u_s_card_code or "",
                        "U_SCardName": obj.u_s_card_name or "",
                        "DocumentLines": payload_lines[:5]  # Show first 5 lines
                    }
                    
                    payload_html = format_html(
                        '<details style="margin: 10px 0; padding: 10px; background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 4px;">'
                        '<summary style="cursor: pointer; font-weight: bold; color: #495057;">📋 Complete SAP Payload Preview (Click to expand)</summary>'
                        '<pre style="margin: 10px 0; padding: 10px; background: #ffffff; border: 1px solid #ced4da; border-radius: 3px; '
                        'overflow-x: auto; font-size: 11px; font-family: monospace; max-height: 500px;">{}</pre>'
                        '<p style="color: #6c757d; font-size: 12px; margin: 5px 0;">Showing {} of {} document lines. Check server logs for complete payload.</p>'
                        '</details>',
                        json.dumps(payload_preview, indent=2, ensure_ascii=False),
                        min(5, len(payload_lines)),
                        len(payload_lines)
                    )
                except Exception as e:
                    payload_html = format_html('<p style="color: #856404;">Could not generate payload preview: {}</p>', str(e))
                
                url = reverse('admin:post_order_to_sap', args=[obj.pk])
                return format_html(
                    '{}'
                    '<a class="button" href="{}" style="padding: 10px 15px; background-color: #417690; color: white; '
                    'text-decoration: none; border-radius: 4px; display: inline-block; font-weight: bold;">'
                    'Add to SAP'
                    '</a>',
                    payload_html,
                    url
                )
        return "-"
    add_to_sap_button.short_description = "SAP Action"
    
    def sap_response_display(self, obj):
        """Display SAP response JSON in a scrollable container with visible scrollbar"""
        from django.utils.html import format_html
        from django.utils.safestring import mark_safe
        import json
        
        if not obj.sap_response_json:
            return format_html('<div style="padding: 10px; color: #999; font-style: italic;">No SAP response yet</div>')
        
        # Pretty print the JSON
        try:
            json_obj = json.loads(obj.sap_response_json)
            formatted_json = json.dumps(json_obj, indent=2)
        except:
            formatted_json = obj.sap_response_json
        
        return format_html(
            '<div style="'
            'width: 100%; '
            'height: 400px; '
            'overflow: scroll !important; '
            'overflow-y: scroll !important; '
            'overflow-x: scroll !important; '
            'border: 3px solid #ff6600; '
            'border-radius: 6px; '
            'background: #f8f9fa; '
            'padding: 0; '
            'position: relative; '
            'display: block; '
            'box-sizing: border-box; '
            'scrollbar-width: auto; '
            'scrollbar-color: #ff6600 #e0e0e0;'
            '">'
            '<pre style="'
            'margin: 0; '
            'padding: 12px; '
            'font-family: Consolas, Monaco, monospace; '
            'font-size: 13px; '
            'line-height: 1.6; '
            'color: #2c3e50; '
            'white-space: pre; '
            'word-wrap: normal; '
            'background: transparent; '
            'border: none; '
            'min-height: 100%; '
            'overflow: visible;'
            '">{}</pre>'
            '</div>'
            '<style>'
            'div[style*="border: 3px solid #ff6600"]::-webkit-scrollbar {{ '
            'width: 16px !important; '
            'height: 16px !important; '
            'display: block !important; '
            '}} '
            'div[style*="border: 3px solid #ff6600"]::-webkit-scrollbar-track {{ '
            'background: #e0e0e0 !important; '
            '}} '
            'div[style*="border: 3px solid #ff6600"]::-webkit-scrollbar-thumb {{ '
            'background: #ff6600 !important; '
            'border: 3px solid #e0e0e0 !important; '
            '}} '
            'div[style*="border: 3px solid #ff6600"]::-webkit-scrollbar-thumb:hover {{ '
            'background: #cc5200 !important; '
            '}}'
            '</style>',
            formatted_json
        )
    sap_response_display.short_description = "Sap response json"
    
    def post_to_sap(self, request, queryset):
        """Action to post selected sales orders to SAP"""
        success_count = 0
        error_count = 0
        
        for order in queryset:
            if order.is_posted_to_sap:
                self.message_user(request, f"Order #{order.id} already posted to SAP", messages.WARNING)
                continue
            
            try:
                # Build SAP payload
                payload = {
                    "Series": order.series,
                    "DocType": order.doc_type,
                    "DocDate": order.doc_date.strftime('%Y-%m-%d') if order.doc_date else None,
                    "DocDueDate": order.doc_due_date.strftime('%Y-%m-%d') if order.doc_due_date else None,
                    "TaxDate": order.tax_date.strftime('%Y-%m-%d') if order.tax_date else None,
                    "CardCode": order.card_code or "",
                    "CardName": order.card_name or "",
                    "ContactPersonCode": order.contact_person_code or None,
                    "FederalTaxID": order.federal_tax_id or None,
                    "PayToCode": order.pay_to_code or None,
                    "Address": order.address or "",
                    "DocCurrency": order.doc_currency or "PKR",
                    "DocRate": float(order.doc_rate) if order.doc_rate else 1.0,
                    "Comments": order.comments or "",
                    "SummeryType": order.summery_type or "dNoSummary",
                    "DocObjectCode": order.doc_object_code or "oOrders",
                    "U_sotyp": order.u_sotyp or "01",
                    "U_USID": order.u_usid or None,
                    "U_SWJE": order.u_swje or None,
                    "U_SECJE": order.u_secje or None,
                    "U_CRJE": order.u_crje or None,
                    "U_SCardCode": order.u_s_card_code or "",
                    "U_SCardName": order.u_s_card_name or "",
                    "DocumentLines": []
                }
                
                # Validate required fields
                if not payload["CardCode"]:
                    raise ValueError(f"Order #{order.id}: CardCode is required")
                if not payload["CardName"]:
                    raise ValueError(f"Order #{order.id}: CardName is required")
                
                # Add document lines
                for line in order.document_lines.all().order_by('line_num'):
                    line_data = {
                        "LineNum": line.line_num,
                        "ItemCode": line.item_code,
                        "ItemDescription": line.item_description,
                        "Quantity": float(line.quantity),
                        "DiscountPercent": float(line.discount_percent),
                        "WarehouseCode": line.warehouse_code,
                        # Service Layer expects TaxCode on document lines
                        "TaxCode": line.vat_group,
                        "UnitsOfMeasurment": float(line.units_of_measurment),
                        "TaxPercentagePerRow": float(line.tax_percentage_per_row),
                        "UnitPrice": float(line.unit_price),
                        "UoMEntry": line.uom_entry,
                        "MeasureUnit": line.measure_unit,
                        "UoMCode": line.uom_code,
                        "ProjectCode": line.project_code,
                        "U_SD": float(line.u_sd),
                        "U_AD": float(line.u_ad),
                        "U_EXD": float(line.u_exd),
                        "U_zerop": float(line.u_zerop),
                        "U_pl": line.u_pl,
                        "U_BP": float(line.u_bp) if line.u_bp else None,
                        "U_policy": line.u_policy,
                        "U_focitem": line.u_focitem,
                        "U_crop": line.u_crop
                    }
                    payload["DocumentLines"].append(line_data)
                
                # Post to SAP
                selected_db = request.session.get('selected_db', '4B-BIO')
                sap_client = SAPClient(company_db_key=selected_db)
                response = sap_client.post('Orders', payload)
                
                # Store complete response
                order.sap_response_json = json.dumps(response, indent=2) if response else None
                
                if response and 'DocEntry' in response:
                    order.sap_doc_entry = response.get('DocEntry')
                    order.sap_doc_num = response.get('DocNum')
                    order.is_posted_to_sap = True
                    order.posted_at = timezone.now()
                    order.sap_error = None
                    order.save()
                    success_count += 1
                    self.message_user(request, f"Order #{order.id} posted successfully. SAP DocNum: {order.sap_doc_num}", messages.SUCCESS)
                else:
                    error_count += 1
                    order.sap_error = "No DocEntry in response"
                    order.save()
                    self.message_user(request, f"Order #{order.id} failed: No DocEntry in response", messages.ERROR)
                    
            except Exception as e:
                error_count += 1
                order.sap_error = str(e)
                order.save()
                self.message_user(request, f"Order #{order.id} failed: {str(e)}", messages.ERROR)
        
        if success_count > 0:
            self.message_user(request, f"{success_count} order(s) posted successfully", messages.SUCCESS)
        if error_count > 0:
            self.message_user(request, f"{error_count} order(s) failed", messages.ERROR)
    
    post_to_sap.short_description = "Post selected orders to SAP"
    
    def save_model(self, request, obj, form, change):
        """Override save_model to ensure all customer fields are saved"""
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info(f"[SAVE_MODEL] Saving SalesOrder #{obj.pk if obj.pk else 'NEW'}")
        logger.info(f"[SAVE_MODEL] card_code from form.cleaned_data: {form.cleaned_data.get('card_code')}")
        logger.info(f"[SAVE_MODEL] card_name from form.cleaned_data: {form.cleaned_data.get('card_name')}")
        logger.info(f"[SAVE_MODEL] address from form.cleaned_data: {form.cleaned_data.get('address')}")
        logger.info(f"[SAVE_MODEL] card_code from obj BEFORE save: {obj.card_code}")
        logger.info(f"[SAVE_MODEL] card_name from obj BEFORE save: {obj.card_name}")
        
        # Explicitly ensure all customer fields are set from form data
        # CRITICAL: Preserve values even if they're empty strings (don't convert to None)
        if 'card_code' in form.cleaned_data and form.cleaned_data['card_code'] is not None:
            obj.card_code = form.cleaned_data['card_code']
            logger.info(f"[SAVE_MODEL] Set obj.card_code = '{obj.card_code}'")
        if 'card_name' in form.cleaned_data and form.cleaned_data['card_name'] is not None:
            obj.card_name = form.cleaned_data['card_name']
            logger.info(f"[SAVE_MODEL] Set obj.card_name = '{obj.card_name}'")
        if 'contact_person_code' in form.cleaned_data and form.cleaned_data['contact_person_code'] is not None:
            obj.contact_person_code = form.cleaned_data['contact_person_code']
        if 'federal_tax_id' in form.cleaned_data and form.cleaned_data['federal_tax_id'] is not None:
            obj.federal_tax_id = form.cleaned_data['federal_tax_id']
        if 'pay_to_code' in form.cleaned_data and form.cleaned_data['pay_to_code'] is not None:
            obj.pay_to_code = form.cleaned_data['pay_to_code']
        if 'address' in form.cleaned_data and form.cleaned_data['address'] is not None:
            obj.address = form.cleaned_data['address']
        
        # Also ensure child customer fields are saved
        if 'u_s_card_code' in form.cleaned_data and form.cleaned_data['u_s_card_code'] is not None:
            obj.u_s_card_code = form.cleaned_data['u_s_card_code']
            logger.info(f"[SAVE_MODEL] Set obj.u_s_card_code = '{obj.u_s_card_code}'")
        if 'u_s_card_name' in form.cleaned_data and form.cleaned_data['u_s_card_name'] is not None:
            obj.u_s_card_name = form.cleaned_data['u_s_card_name']
            logger.info(f"[SAVE_MODEL] Set obj.u_s_card_name = '{obj.u_s_card_name}'")
        
        # Call parent save_model
        super().save_model(request, obj, form, change)
        
        # Log after save
        obj.refresh_from_db()
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - card_code in DB: '{obj.card_code}'")
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - card_name in DB: '{obj.card_name}'")
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - address in DB: '{obj.address}'")
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - contact_person_code in DB: {obj.contact_person_code}")
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - federal_tax_id in DB: '{obj.federal_tax_id}'")
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - u_s_card_code in DB: '{obj.u_s_card_code}'")
        logger.info(f"[SAVE_MODEL] ✓ After save and refresh - u_s_card_name in DB: '{obj.u_s_card_name}'")
    
    def get_urls(self):
        """Add custom URL for individual order posting"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('<int:order_id>/post-to-sap/', 
                 self.admin_site.admin_view(self.post_single_order_to_sap),
                 name='post_order_to_sap'),
        ]
        return custom_urls + urls
    
    def post_single_order_to_sap(self, request, order_id):
        """Handle posting a single order to SAP with async JSON response"""
        import logging
        import re
        from django.shortcuts import redirect
        from django.urls import reverse
        from django.http import JsonResponse
        
        # Entry log and safe order fetch (handles 404s gracefully)
        logger = logging.getLogger(__name__)
        try:
            logger.info(f"[POST_TO_SAP] Entry: method={request.method}, ajax={request.headers.get('X-Requested-With') == 'XMLHttpRequest'}, order_id={order_id}")
            # Fetch order and refresh from database to ensure we have latest data
            order = SalesOrder.objects.select_related('staff', 'dealer').get(pk=order_id)
            order.refresh_from_db()
        except SalesOrder.DoesNotExist:
            err = f"Order #{order_id} not found"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': err}, status=404)
            else:
                self.message_user(request, err, messages.ERROR)
                return redirect(reverse('admin:FieldAdvisoryService_salesorder_changelist'))
        
        # Log the order data from database
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[ORDER DATA] Order #{order_id} from database (refreshed):")
        logger.info(f"  CardCode: '{order.card_code}'")
        logger.info(f"  CardCode type: {type(order.card_code)}")
        logger.info(f"  CardCode is None: {order.card_code is None}")
        logger.info(f"  CardCode is empty string: {order.card_code == ''}")
        logger.info(f"  CardName: '{order.card_name}'")
        logger.info(f"  U_SCardCode: '{order.u_s_card_code}'")
        logger.info(f"  U_SCardName: '{order.u_s_card_name}'")
        logger.info(f"  DocumentLines: {order.document_lines.count()}")
        if order.is_posted_to_sap:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': f"Order #{order.id} is already posted to SAP"
                }, status=400)
            else:
                self.message_user(request, f"Order #{order.id} already posted to SAP", messages.WARNING)
                return redirect(reverse('admin:FieldAdvisoryService_salesorder_change', args=[order_id]))
        
        try:
            # Build SAP payload
            payload = {
                "Series": order.series,
                "DocType": order.doc_type,
                "DocDate": order.doc_date.strftime('%Y-%m-%d') if order.doc_date else None,
                "DocDueDate": order.doc_due_date.strftime('%Y-%m-%d') if order.doc_due_date else None,
                "TaxDate": order.tax_date.strftime('%Y-%m-%d') if order.tax_date else None,
                "CardCode": order.card_code or "",
                "CardName": order.card_name or "",
                "ContactPersonCode": order.contact_person_code or None,
                "FederalTaxID": order.federal_tax_id or None,
                "PayToCode": order.pay_to_code or None,
                "Address": order.address or "",
                "DocCurrency": order.doc_currency or "PKR",
                "DocRate": float(order.doc_rate) if order.doc_rate else 1.0,
                "Comments": order.comments or "",
                "SummeryType": order.summery_type or "dNoSummary",
                "DocObjectCode": order.doc_object_code or "oOrders",
                "U_sotyp": order.u_sotyp or "01",
                "U_USID": order.u_usid or None,
                "U_SWJE": order.u_swje or None,
                "U_SECJE": order.u_secje or None,
                "U_CRJE": order.u_crje or None,
                "U_SCardCode": order.u_s_card_code or "",
                "U_SCardName": order.u_s_card_name or "",
                "DocumentLines": []
            }
            
            # Validate required fields
            if not payload["CardCode"]:
                raise ValueError(
                    "Customer Code (CardCode) is required but is empty. "
                    "Please select a customer from the 'Customer Code' dropdown and SAVE the form before clicking 'Add to SAP'."
                )
            if not payload["CardName"]:
                raise ValueError(
                    "Customer Name (CardName) is required but is empty. "
                    "Please select a customer from the 'Customer Code' dropdown and SAVE the form before clicking 'Add to SAP'."
                )
            
            # Log the complete payload before posting
            logger.info(f"[SAP PAYLOAD] Order #{order.id} payload:")
            logger.info(json.dumps(payload, indent=2, ensure_ascii=False))
            
            # Add document lines
            for line in order.document_lines.all().order_by('line_num'):
                # CRITICAL: Clean vat_group to prevent SAP validation errors
                vat_group = (line.vat_group or "SE").strip()
                
                # Aggressively sanitize: keep only alphanumeric, dash, underscore
                # This catches any corruption, non-ASCII, or invalid characters
                vat_group_clean = re.sub(r'[^A-Za-z0-9\-_]', '', vat_group)
                
                if not vat_group_clean or vat_group_clean != vat_group:
                    logger.warning(f"[SAP PAYLOAD] Line {line.line_num}: Invalid/corrupted vat_group {repr(vat_group)} → sanitized to {repr(vat_group_clean)}")
                    if not vat_group_clean:
                        vat_group = "SE"
                    else:
                        vat_group = vat_group_clean
                else:
                    vat_group = vat_group_clean
                
                # Final safety: validate it's a known SAP tax code (fallback list)
                VALID_TAX_CODES = {'SE', 'AT1', 'ST', 'VAT0', 'VAT5', 'VAT17', 'VAT20'}
                if vat_group not in VALID_TAX_CODES:
                    logger.warning(f"[SAP PAYLOAD] Line {line.line_num}: Unknown tax code {repr(vat_group)}, using default 'SE'")
                    vat_group = "SE"
                
                line_data = {
                    "LineNum": line.line_num,
                    "ItemCode": line.item_code or "",
                    "ItemDescription": line.item_description or "",
                    "Quantity": float(line.quantity) if line.quantity else 0.0,
                    "DiscountPercent": float(line.discount_percent) if line.discount_percent else 0.0,
                    "WarehouseCode": line.warehouse_code or "",
                    # Service Layer expects TaxCode on document lines
                    "TaxCode": vat_group,
                    "UnitsOfMeasurment": float(line.units_of_measurment) if line.units_of_measurment else 1.0,
                    "TaxPercentagePerRow": float(line.tax_percentage_per_row) if line.tax_percentage_per_row else 0.0,
                    "UnitPrice": float(line.unit_price) if line.unit_price else 0.0,
                    "UoMEntry": line.uom_entry or None,
                    "MeasureUnit": line.measure_unit or "",
                    "UoMCode": line.uom_code or "",
                    "ProjectCode": line.project_code or "",
                    "U_SD": float(line.u_sd) if line.u_sd else 0.0,
                    "U_AD": float(line.u_ad) if line.u_ad else 0.0,
                    "U_EXD": float(line.u_exd) if line.u_exd else 0.0,
                    "U_zerop": float(line.u_zerop) if line.u_zerop else 0.0,
                    "U_pl": line.u_pl or None,
                    "U_BP": float(line.u_bp) if line.u_bp else None,
                    "U_policy": line.u_policy or "",
                    "U_focitem": line.u_focitem or "No",
                    "U_crop": line.u_crop or None
                }
                payload["DocumentLines"].append(line_data)
            
            # Log the complete payload with document lines
            logger.info(f"[SAP PAYLOAD] Complete payload with {len(payload['DocumentLines'])} lines:")
            logger.info(json.dumps(payload, indent=2, ensure_ascii=False))
            
            # Post to SAP (this is the blocking call that takes 10-20 seconds)
            selected_db = request.session.get('selected_db', '4B-BIO')
            logger.info(f"[SAP PAYLOAD] Using company DB: {selected_db}")
            sap_client = SAPClient(company_db_key=selected_db)
            response = sap_client.post('Orders', payload)
            
            # Store complete response
            order.sap_response_json = json.dumps(response, indent=2) if response else None
            
            if response and 'DocEntry' in response:
                order.sap_doc_entry = response.get('DocEntry')
                order.sap_doc_num = response.get('DocNum')
                order.is_posted_to_sap = True
                order.posted_at = timezone.now()
                order.sap_error = None
                order.save()
                
                success_message = f"Order #{order.id} posted successfully to SAP"
                
                # Return JSON response for AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': success_message,
                        'doc_entry': order.sap_doc_entry,
                        'doc_num': order.sap_doc_num
                    })
                else:
                    self.message_user(request, 
                        f"✓ {success_message}! DocEntry: {order.sap_doc_entry}, DocNum: {order.sap_doc_num}", 
                        messages.SUCCESS)
            else:
                error_msg = "No DocEntry in SAP response"
                order.sap_error = error_msg
                order.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    }, status=400)
                else:
                    self.message_user(request, f"Order #{order.id} failed: {error_msg}", messages.ERROR)
                
        except Exception as e:
            error_msg = str(e)
            order.sap_error = error_msg
            order.sap_response_json = json.dumps({"error": error_msg}, indent=2)
            order.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=500)
            else:
                self.message_user(request, f"Order #{order.id} failed: {error_msg}", messages.ERROR)
        
        # For non-AJAX requests, redirect back
        return redirect(reverse('admin:FieldAdvisoryService_salesorder_change', args=[order_id]))


admin.site.register(SalesOrderAttachment)





@admin.register(Company, site=admin_site)
class CompanyAdmin(admin.ModelAdmin):
    list_display = ('Company_name', 'name', 'email', 'contact_number', 'is_active')
    search_fields = ('Company_name', 'name', 'email')
    list_filter = ('is_active',)
    
    fieldsets = (
        ('Company Information', {
            'fields': ('Company_name', 'description', 'is_active')
        }),
        ('HANA Database Schema', {
            'fields': ('name',),
            'description': '<strong>Note:</strong> The "Name" field represents the HANA database schema name (e.g., 4B-BIO_APP, 4B-ORANG_APP). This value is used in HANA Connect for database connections.'
        }),
        ('Contact Information', {
            'fields': ('email', 'contact_number', 'address')
        }),
        ('Location', {
            'fields': ('latitude', 'longitude'),
            'classes': ('collapse',)
        }),
        ('Additional Information', {
            'fields': ('remarks', 'created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('created_at', 'updated_at')

@admin.register(Region, site=admin_site)
class RegionAdmin(admin.ModelAdmin, _CompanySessionResolver):
    list_display = ('name', 'company')
    list_filter = ('company',)
    search_fields = ('name',)
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        comp = self._get_selected_company(request)
        return qs.filter(company=comp) if comp else qs

@admin.register(Zone, site=admin_site)
class ZoneAdmin(admin.ModelAdmin, _CompanySessionResolver):
    list_display = ('name', 'region', 'company')
    list_filter = ('region', 'company')
    search_fields = ('name', 'region__name')
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        comp = self._get_selected_company(request)
        return qs.filter(company=comp) if comp else qs

@admin.register(Territory, site=admin_site)
class TerritoryAdmin(admin.ModelAdmin, _CompanySessionResolver):
    def region(self, obj):
        return obj.zone.region if obj.zone else None
    region.short_description = 'Region'
    list_display = ('name', 'zone', 'region', 'company')
    list_filter = ('zone', 'zone__region', 'company')
    search_fields = ('name', 'zone__name')
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        comp = self._get_selected_company(request)
        return qs.filter(company=comp) if comp else qs

@admin.register(Dealer, site=admin_site)
class DealerAdmin(admin.ModelAdmin):
    list_display = ('name', 'business_name', 'user_link', 'user_email', 'contact_number', 'company', 'filer_status', 'is_active', 'created_at')
    list_select_related = ('user', 'company')
    list_filter = ('is_active', 'company', 'filer_status', 'card_type', 'vat_liable', 'created_at')
    search_fields = ('name', 'business_name', 'contact_number', 'mobile_phone', 'email', 'user__username', 'user__email', 'cnic_number', 'card_code')
    raw_id_fields = ('user', 'company', 'region', 'zone', 'territory')
    readonly_fields = ('created_at', 'updated_at', 'card_code')
    
    fieldsets = (
        ('User Account', {
            'fields': ('user', 'card_code'),
            'description': 'Link or create a user account for dealer login'
        }),
        ('Dealer Information', {
            'fields': ('name', 'business_name', 'cnic_number')
        }),
        ('Contact Information', {
            'fields': ('email', 'contact_number', 'mobile_phone')
        }),
        ('Address Information', {
            'fields': ('address', 'city', 'state', 'country', 'latitude', 'longitude')
        }),
        ('Geographic Assignment', {
            'fields': ('company', 'region', 'zone', 'territory')
        }),
        ('Tax & Legal Information', {
            'fields': ('federal_tax_id', 'additional_id', 'unified_federal_tax_id', 'filer_status'),
            'classes': ('collapse',)
        }),
        ('License Information', {
            'fields': ('govt_license_number', 'license_expiry', 'u_leg'),
            'classes': ('collapse',)
        }),
        ('SAP Configuration', {
            'fields': ('sap_series', 'card_type', 'group_code', 'debitor_account', 'vat_group', 'vat_liable', 'whatsapp_messages'),
            'classes': ('collapse',)
        }),
        ('Financial', {
            'fields': ('minimum_investment',),
            'classes': ('collapse',)
        }),
        ('CNIC Images', {
            'fields': ('cnic_front_image', 'cnic_back_image'),
            'classes': ('collapse',)
        }),
        ('Additional Information', {
            'fields': ('remarks',),
            'classes': ('collapse',)
        }),
        ('Status & Audit', {
            'fields': ('is_active', 'created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def user_link(self, obj):
        """Display user with link to user admin"""
        if obj.user:
            from django.urls import reverse
            from django.utils.html import format_html
            url = reverse('admin:accounts_user_change', args=[obj.user.pk])
            return format_html('<a href="{}">{}</a>', url, obj.user.username)
        return format_html('<span style="color: #999;">No user assigned</span>')
    user_link.short_description = 'User Account'

    def user_email(self, obj):
        return obj.user.email if obj.user else '-'
    user_email.short_description = 'Email'
    
    def save_model(self, request, obj, form, change):
        """Set created_by, sync is_dealer flag, and derive name from user"""
        if not change:  # Creating new object
            obj.created_by = request.user
        
        # Sync is_dealer flag with user
        if obj.user and not obj.user.is_dealer:
            obj.user.is_dealer = True
            obj.user.save(update_fields=["is_dealer"])

        # Derive name from linked user's name if not provided or to keep in sync
        if obj.user:
            first = getattr(obj.user, 'first_name', '') or ''
            last = getattr(obj.user, 'last_name', '') or ''
            full = (first + ' ' + last).strip()
            if not full:
                full = getattr(obj.user, 'username', None) or getattr(obj.user, 'email', '')
            obj.name = full
        
        super().save_model(request, obj, form, change)
    
@admin.register(DealerRequest, site=admin_site)
class DealerRequestAdmin(admin.ModelAdmin):
    change_list_template = 'admin/fieldadvisoryservice/dealerrequest/change_list.html'
    change_form_template = 'admin/fieldadvisoryservice/dealerrequest/change_form.html'
    list_display = (
        'business_name', 'owner_name', 'status', 'is_posted_to_sap', 'sap_card_code',
        'requested_by', 'reviewed_by', 'filer_status', 'created_at'
    )
    list_filter = ('status', 'is_posted_to_sap', 'filer_status', 'card_type', 'company', 'region', 'zone', 'territory')
    search_fields = ('business_name', 'owner_name', 'requested_by__username', 'cnic_number', 'sap_card_code', 'email')
    actions = ['approve_create_bp']
    readonly_fields = ('created_at', 'updated_at', 'reviewed_at', 'is_posted_to_sap', 'sap_card_code', 'sap_doc_entry', 'sap_error', 'posted_at', 'sap_response_json')
    
    fieldsets = (
        ('Request Information', {
            'fields': ('requested_by', 'status', 'reason')
        }),
        ('Business Partner Basic Info', {
            'fields': ('business_name', 'owner_name', 'contact_number', 'mobile_phone', 'email')
        }),
        ('Address Information', {
            'fields': ('address', 'city', 'state', 'country')
        }),
        ('Tax & Legal Information', {
            'fields': ('cnic_number', 'federal_tax_id', 'additional_id', 'unified_federal_tax_id', 'filer_status')
        }),
        ('License Information', {
            'fields': ('govt_license_number', 'license_expiry', 'u_leg')
        }),
        ('Documents', {
            'fields': ('cnic_front', 'cnic_back')
        }),
        ('Territory & Organization', {
            'fields': ('company', 'region', 'zone', 'territory')
        }),
        ('SAP Configuration', {
            'fields': ('sap_series', 'card_type', 'group_code', 'debitor_account', 'vat_group', 'vat_liable', 'whatsapp_messages'),
            'classes': ('collapse',)
        }),
        ('Financial', {
            'fields': ('minimum_investment',)
        }),
        ('Review Information', {
            'fields': ('reviewed_at', 'reviewed_by'),
            'classes': ('collapse',)
        }),
        ('SAP Integration Status', {
            'fields': ('is_posted_to_sap', 'sap_card_code', 'sap_doc_entry', 'sap_error', 'posted_at', 'sap_response_json'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        return form

    def save_model(self, request, obj, form, change):
        if change and not obj.reviewed_by:
            obj.reviewed_by = request.user
        # Capture previous status before saving
        prev_status = None
        if change and obj.pk:
            try:
                prev_status = DealerRequest.objects.filter(pk=obj.pk).values_list('status', flat=True).first()
            except Exception:
                prev_status = None
        super().save_model(request, obj, form, change)
        
        # Call SAP only when transitioning to 'approved' or 'posted_to_sap'
        if obj.status in ['approved', 'posted_to_sap'] and prev_status not in ['approved', 'posted_to_sap'] and not obj.is_posted_to_sap:
            selected_db = request.session.get('selected_db', '4B-BIO')
            sap = SAPClient(company_db_key=selected_db)
            territory_id = None
            if obj.territory and obj.territory.name:
                territory_id = sap.get_territory_id_by_name(obj.territory.name)
            
            addr_name = 'Bill To'
            addr_type = 'bo_BillTo'
            
            # Build payload using model fields directly
            payload = {
                'Series': obj.sap_series or 70,
                'CardName': obj.business_name or '',
                'CardType': obj.card_type or 'cCustomer',
                'GroupCode': obj.group_code or 100,
                'Address': (obj.address or '')[:100],
                'Phone1': obj.contact_number or '',
                'MobilePhone': obj.mobile_phone or obj.contact_number or '',
                'ContactPerson': obj.owner_name or '',
                'FederalTaxID': obj.federal_tax_id or obj.cnic_number or None,
                'AdditionalID': obj.additional_id or None,
                'OwnerIDNumber': obj.cnic_number or None,
                'UnifiedFederalTaxID': obj.unified_federal_tax_id or obj.cnic_number or None,
                'Territory': territory_id,
                'DebitorAccount': obj.debitor_account or 'A020301001',
                'U_leg': obj.u_leg or '17-5349',
                'U_gov': obj.license_expiry.isoformat() if obj.license_expiry else None,
                'U_fil': obj.filer_status or None,
                'U_lic': obj.govt_license_number or None,
                'U_region': obj.region.name if obj.region else None,
                'U_zone': obj.zone.name if obj.zone else None,
                'U_WhatsappMessages': obj.whatsapp_messages or 'YES',
                'VatGroup': obj.vat_group or 'AT1',
                'VatLiable': obj.vat_liable or 'vLiable',
                'BPAddresses': [
                    {
                        'AddressName': addr_name,
                        'AddressName2': None,
                        'AddressName3': None,
                        'City': obj.city or None,
                        'Country': obj.country or 'PK',
                        'State': obj.state or None,
                        'Street': (obj.address or '')[:50],
                        'AddressType': addr_type,
                    }
                ],
                'ContactEmployees': [
                    {
                        'Name': obj.owner_name or '',
                        'Position': None,
                        'MobilePhone': obj.mobile_phone or obj.contact_number or None,
                        'E_Mail': obj.email or None,
                    }
                ],
            }
            
            try:
                # Log payload
                try:
                    print("SAP Business Partner payload:", json.dumps(payload, ensure_ascii=False), flush=True)
                except Exception:
                    pass
                
                # Create BP in SAP
                result = sap.create_business_partner(payload)
                summary = None
                card_code = None
                doc_entry = None
                
                if isinstance(result, dict):
                    card_code = result.get('CardCode') or result.get('code')
                    doc_entry = result.get('DocEntry')
                    
                    # Try to extract CardCode from Location header if not in body
                    if not card_code:
                        hdrs = result.get('headers') if isinstance(result.get('headers'), dict) else None
                        if hdrs:
                            loc = hdrs.get('Location') or hdrs.get('location')
                            if isinstance(loc, str):
                                try:
                                    start = loc.find("BusinessPartners('")
                                    if start != -1:
                                        start += len("BusinessPartners('")
                                        end = loc.find("')", start)
                                        if end != -1:
                                            card_code = loc[start:end]
                                except Exception:
                                    pass
                    
                    if card_code:
                        summary = f"CardCode={card_code}"
                
                # Get full BP details if CardCode available
                try:
                    details = None
                    if card_code:
                        try:
                            details = sap.get_bp_details(card_code)
                        except Exception:
                            try:
                                details = sap.get_business_partner(card_code)
                            except Exception:
                                details = None
                    msg_json = json.dumps(details if details is not None else result, ensure_ascii=False, indent=2)
                except Exception:
                    try:
                        msg_json = json.dumps(result, ensure_ascii=False, indent=2)
                    except Exception:
                        msg_json = str(result)
                
                # Log response
                try:
                    logging.getLogger('sap').info(msg_json)
                except Exception:
                    pass
                
                try:
                    print("SAP Business Partner response:", msg_json, flush=True)
                except Exception:
                    pass
                
                # Show success message
                messages.success(request, f"SAP_RESPONSE_JSON:{msg_json}")
                if summary:
                    messages.success(request, f"SAP Business Partner created ({summary}).")
                else:
                    messages.success(request, "SAP Business Partner created.")
                
                # Update dealer request with SAP response
                try:
                    obj.sap_response_json = msg_json
                    obj.posted_at = timezone.now()
                    obj.is_posted_to_sap = True
                    obj.status = 'posted_to_sap'
                    if card_code:
                        obj.sap_card_code = card_code
                    if doc_entry:
                        obj.sap_doc_entry = doc_entry
                    obj.sap_error = None  # Clear previous errors
                    obj.save(update_fields=['sap_response_json', 'posted_at', 'is_posted_to_sap', 'status', 'sap_card_code', 'sap_doc_entry', 'sap_error'])
                except Exception as save_err:
                    print(f"Error saving SAP response: {save_err}", flush=True)
                    
            except Exception as e:
                # Log error
                error_msg = str(e)
                try:
                    logging.getLogger('sap').error(error_msg)
                except Exception:
                    pass
                
                try:
                    print("SAP Business Partner error:", error_msg, flush=True)
                except Exception:
                    pass
                
                # Show error message and save to model
                messages.error(request, f"SAP Business Partner creation failed: {error_msg}")
                try:
                    obj.sap_error = error_msg[:500]  # Store first 500 chars
                    obj.save(update_fields=['sap_error'])
                except Exception:
                    pass
                messages.error(request, f"SAP_RESPONSE_JSON:{str(e)}")
                messages.error(request, f"SAP Business Partner creation failed: {str(e)}")
                try:
                    obj.sap_response_json = str(e)
                    obj.sap_response_at = timezone.now()
                    obj.save(update_fields=['sap_response_json','sap_response_at'])
                except Exception:
                    pass

    def approve_create_bp(self, request, queryset):
        for obj in queryset:
            prev_status = obj.status
            if obj.reviewed_by is None:
                obj.reviewed_by = request.user
            obj.status = 'approved'
            obj.save()
            if prev_status == 'approved':
                continue
            selected_db = request.session.get('selected_db', '4B-BIO')
            sap = SAPClient(company_db_key=selected_db)
            territory_id = None
            if obj.territory and obj.territory.name:
                territory_id = sap.get_territory_id_by_name(obj.territory.name)
            payload = {
                'Series': 70,
                'CardName': obj.business_name,
                'CardType': 'cCustomer',
                'GroupCode': 100,
                'Address': obj.address or '',
                'Phone1': obj.contact_number,
                'MobilePhone': obj.contact_number,
                'ContactPerson': obj.owner_name,
                'FederalTaxID': obj.cnic_number,
                'AdditionalID': None,
                'OwnerIDNumber': obj.cnic_number,
                'UnifiedFederalTaxID': obj.cnic_number,
                'DebitorAccount': 'A020301001',
                'U_leg': '17-5349',
                'U_gov': obj.license_expiry.isoformat() if obj.license_expiry else None,
                'U_fil': obj.filer_status,
                'U_lic': obj.govt_license_number,
                'U_region': obj.region.name if obj.region else None,
                'U_zone': obj.zone.name if obj.zone else None,
                'U_WhatsappMessages': 'YES',
                'VatGroup': 'AT1',
                'VatLiable': 'vLiable',
                'BPAddresses': [
                    {
                        'AddressName': 'Bill To',
                        'AddressName2': None,
                        'AddressName3': None,
                        'City': None,
                        'Country': 'PK',
                        'State': None,
                        'Street': (obj.address or '')[:50],
                        'AddressType': 'bo_BillTo',
                    }
                ],
                'ContactEmployees': [
                    {
                        'Name': obj.owner_name,
                        'Position': None,
                        'MobilePhone': obj.contact_number or None,
                        'E_Mail': None,
                    }
                ],
            }
            if territory_id is not None:
                payload['Territory'] = territory_id
            try:
                result = sap.create_business_partner(payload)
                # Extract CardCode from response
                summary = None
                card_code = None
                if isinstance(result, dict):
                    card_code = result.get('CardCode') or result.get('code')
                    if not card_code:
                        hdrs = result.get('headers') if isinstance(result.get('headers'), dict) else None
                        if hdrs:
                            loc = hdrs.get('Location') or hdrs.get('location')
                            if isinstance(loc, str):
                                try:
                                    start = loc.find("BusinessPartners('")
                                    if start != -1:
                                        start += len("BusinessPartners('")
                                        end = loc.find("')", start)
                                        if end != -1:
                                            card_code = loc[start:end]
                                except Exception:
                                    pass
                    if card_code:
                        summary = f"CardCode={card_code}"
                # Prefer full details payload for persisted response
                try:
                    details = None
                    if card_code:
                        try:
                            details = sap.get_bp_details(card_code)
                        except Exception:
                            # Fallback to full GET entity
                            try:
                                details = sap.get_business_partner(card_code)
                            except Exception:
                                details = None
                    msg_json = json.dumps(details if details is not None else result, ensure_ascii=False, indent=2)
                except Exception:
                    try:
                        msg_json = json.dumps(result, ensure_ascii=False, indent=2)
                    except Exception:
                        msg_json = str(result)
                messages.success(request, f"SAP_RESPONSE_JSON:{msg_json}")
                if summary:
                    messages.success(request, f"SAP Business Partner created for request {obj.pk} ({summary}).")
                else:
                    messages.success(request, f"SAP Business Partner created for request {obj.pk}.")
                try:
                    obj.sap_response_json = msg_json
                    obj.sap_response_at = timezone.now()
                    if card_code:
                        obj.sap_card_code = card_code
                    obj.save(update_fields=['sap_response_json','sap_response_at','sap_card_code'])
                except Exception:
                    pass
            except Exception as e:
                messages.error(request, f"SAP_RESPONSE_JSON:{str(e)}")
                messages.error(request, f"SAP Business Partner creation failed for request {obj.pk}: {str(e)}")
                try:
                    obj.sap_response_json = str(e)
                    obj.sap_response_at = timezone.now()
                    obj.save(update_fields=['sap_response_json','sap_response_at'])
                except Exception:
                    pass

    def get_urls(self):
        urls = super().get_urls()
        from django.urls import path
        custom = [
            path('<int:pk>/approve-add-to-sap/', self.admin_site.admin_view(self.approve_add_view), name='dealerrequest_approve_add_to_sap'),
        ]
        return custom + urls

    def approve_add_view(self, request, pk):
        from django.shortcuts import redirect, get_object_or_404
        obj = get_object_or_404(DealerRequest, pk=pk)
        prev_status = obj.status
        if obj.reviewed_by is None:
            obj.reviewed_by = request.user
        obj.status = 'approved'
        obj.save()
        if prev_status != 'approved':
            self.approve_create_bp(request, queryset=[obj])
        else:
            messages.info(request, f"SAP Business Partner already processed for request {obj.pk} (status approved).")
            # Also fetch and display current BP details if CardCode is known
            try:
                if obj.sap_card_code:
                    selected_db = request.session.get('selected_db', '4B-BIO')
                    sap = SAPClient(company_db_key=selected_db)
                    try:
                        details = sap.get_bp_details(obj.sap_card_code)
                    except Exception:
                        try:
                            details = sap.get_business_partner(obj.sap_card_code)
                        except Exception:
                            details = None
                    if details is not None:
                        msg_json = json.dumps(details, ensure_ascii=False, indent=2)
                        messages.success(request, f"SAP_RESPONSE_JSON:{msg_json}")
                        try:
                            obj.sap_response_json = msg_json
                            obj.sap_response_at = timezone.now()
                            obj.save(update_fields=['sap_response_json','sap_response_at'])
                        except Exception:
                            pass
            except Exception:
                pass
        return redirect(f'../../{pk}/change/')