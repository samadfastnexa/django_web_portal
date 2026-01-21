from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from .models import Meeting
from .serializers import MeetingSerializer
from rest_framework import viewsets, filters, status, serializers
from rest_framework.decorators import action
from .models import FieldDay
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .serializers import FieldDaySerializer
from FieldAdvisoryService.serializers import CompanySerializer, RegionSerializer, ZoneSerializer, TerritorySerializer,Company,Region,Zone,Territory

class MeetingViewSet(viewsets.ModelViewSet):
    # queryset = Meeting.objects.all()  # Use .filter(is_active=True) after adding the field
    queryset = Meeting.objects.select_related(
        'region_fk', 'zone_fk', 'territory_fk', 'company_fk'
    ).all().order_by('-id')
    serializer_class = MeetingSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    # Filters, search, ordering
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["fsm_name", "region_fk", "zone_fk", "territory_fk", "company_fk", "presence_of_zm", "presence_of_rsm", "location", "user_id"]
    search_fields = [
        'fsm_name',
        'region_fk__name',
        'zone_fk__name',
        'territory_fk__name',
        'location',
        'key_topics_discussed',
    ]
    ordering_fields = ["date", "fsm_name", "region_fk__name", "zone_fk__name", "territory_fk__name", "total_attendees", "id"]
    ordering = ["-id"]
    
    def get_queryset(self):
        """
        Filter meetings based on user's geographic hierarchy AND reporting hierarchy.
        Users see:
        1. Meetings in their assigned geographic areas (companies/regions/zones/territories)
        2. Meetings created by their subordinates (reporting hierarchy)
        """
        from django.db.models import Q
        
        queryset = super().get_queryset()
        user = self.request.user
        
        # During schema generation, return empty queryset for AnonymousUser
        if not user.is_authenticated:
            return queryset.none()
        
        # Superusers see everything
        if user.is_superuser or user.is_staff:
            return queryset
        
        # Get user's sales profile
        sales_profile = getattr(user, 'sales_profile', None)
        if not sales_profile:
            # Non-sales staff see only their own meetings
            return queryset.filter(user_id=user)
        
        # Build filter for geographic hierarchy
        geo_filter = Q()
        
        # Add filters for assigned geographic areas
        company_ids = list(sales_profile.companies.values_list('id', flat=True))
        region_ids = list(sales_profile.regions.values_list('id', flat=True))
        zone_ids = list(sales_profile.zones.values_list('id', flat=True))
        territory_ids = list(sales_profile.territories.values_list('id', flat=True))
        
        if company_ids:
            geo_filter |= Q(company_fk_id__in=company_ids)
        if region_ids:
            geo_filter |= Q(region_fk_id__in=region_ids)
        if zone_ids:
            geo_filter |= Q(zone_fk_id__in=zone_ids)
        if territory_ids:
            geo_filter |= Q(territory_fk_id__in=territory_ids)
        
        # Build filter for reporting hierarchy (subordinates)
        reporting_filter = Q()
        
        # Get all subordinates' user IDs
        subordinate_users = sales_profile.get_subordinate_users()
        subordinate_user_ids = list(subordinate_users.values_list('id', flat=True))
        
        # Get all subordinates' territory IDs
        subordinate_territory_ids = sales_profile.get_subordinate_territory_ids()
        
        if subordinate_user_ids:
            reporting_filter |= Q(user_id__in=subordinate_user_ids)
        if subordinate_territory_ids:
            reporting_filter |= Q(territory_fk_id__in=subordinate_territory_ids)
        
        # Include own meetings
        own_filter = Q(user_id=user)
        
        # Combine all filters with OR
        final_filter = own_filter | geo_filter | reporting_filter
        
        return queryset.filter(final_filter).distinct()
    
     # ---------------- Global Extra Data ----------------
    # Removed verbose nested data to reduce response size
    # The API now returns only IDs and names for companies, regions, zones, and territories
    
    # Define common parameters
    common_parameters = [
        openapi.Parameter('fsm_name', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True,
                         description='Field Sales Manager name'),
        openapi.Parameter('company_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                         description='Company ID (Foreign Key)'),
        openapi.Parameter('territory_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                         description='Territory ID (Foreign Key)'),
        openapi.Parameter('zone_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                         description='Zone ID (Foreign Key)'),
        openapi.Parameter('region_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                         description='Region ID (Foreign Key)'),
        openapi.Parameter('date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=True,
                         description='Meeting date (YYYY-MM-DD)'),
        openapi.Parameter('location', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True,
                         description='Meeting location'),
        openapi.Parameter('total_attendees', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=True,
                         description='Total number of attendees'),
        openapi.Parameter('key_topics_discussed', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False,
                         description='Key topics discussed in the meeting'),
        openapi.Parameter('presence_of_zm', openapi.IN_FORM, type=openapi.TYPE_BOOLEAN, required=False,
                         description='Presence of Zone Manager (ZM)'),
        openapi.Parameter('presence_of_rsm', openapi.IN_FORM, type=openapi.TYPE_BOOLEAN, required=False,
                         description='Presence of Regional Sales Manager (RSM)'),
        openapi.Parameter('feedback_from_attendees', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False,
                         description='Feedback from attendees'),
        openapi.Parameter('suggestions_for_future', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False,
                         description='Suggestions for future meetings'),
        openapi.Parameter('attachments', openapi.IN_FORM, type=openapi.TYPE_FILE, required=False,
                         description='File attachments (can upload multiple)'),
        
        # Multiple fields for attendees
        openapi.Parameter('attendee_name', openapi.IN_FORM, type=openapi.TYPE_ARRAY, 
                         items=openapi.Items(type=openapi.TYPE_STRING), required=False,
                         description='List of farmer names. Example: ["John Doe", "Jane Smith"]'),
        openapi.Parameter('attendee_contact', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                         items=openapi.Items(type=openapi.TYPE_STRING), required=False,
                         description='List of contact numbers. Example: ["123-456-7890", "098-765-4321"]'),
        openapi.Parameter('attendee_acreage', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                         items=openapi.Items(type=openapi.TYPE_NUMBER), required=False,
                         description='List of acreage values. Example: [5.5, 3.2]'),
        openapi.Parameter('attendee_crop', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                         items=openapi.Items(type=openapi.TYPE_STRING), required=False,
                         description='List of crops. Example: ["Wheat", "Corn"]'),
    ]

    # ---------------- List / Retrieve ----------------
    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Retrieve a list of all farmer advisory meetings with comprehensive filtering, search, and ordering capabilities. Includes related company, region, zone, and territory data.",
        responses={
            200: openapi.Response(
                description='List of farmer meetings with dropdown data',
                examples={
                    'application/json': {
                        'results': [
                            {
                                'id': 1,
                                'fsm_name': 'Ahmed Ali',
                                'company_id': 1,
                         'company_name': 'ABC Agriculture',
                         'region_id': 1,
                         'region_name': 'Punjab',
                         'zone_id': 1,
                         'zone_name': 'Lahore',
                         'territory_id': 1,
                         'territory_name': 'Model Town',
                                'date': '2024-01-15',
                                'location': 'Community Center, Model Town',
                                'total_attendees': 25,
                                'key_topics_discussed': 'Crop rotation, pest management',
                                'presence_of_zm': True,
                                'presence_of_rsm': False,
                                'feedback_from_attendees': 'Very informative session',
                                'suggestions_for_future': 'More practical demonstrations',
                                'attendees': [
                                    {
                                        'name': 'Farmer John',
                                        'contact': '+92-300-1234567',
                                        'acreage': 5.5,
                                        'crop': 'Wheat'
                                    }
                                ]
                            }
                        ],
                        'companies': [{'id': 1, 'name': 'ABC Agriculture'}],
                        'regions': [{'id': 1, 'name': 'Punjab'}],
                        'zones': [{'id': 1, 'name': 'Lahore'}],
                        'territories': [{'id': 1, 'name': 'Model Town'}]
                    }
                }
            )
        },
        manual_parameters=[
            openapi.Parameter('fsm_name', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False,
                            description='Filter by FSM name'),
            openapi.Parameter('company_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                            description='Filter by company ID'),
            openapi.Parameter('region_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                            description='Filter by region ID'),
            openapi.Parameter('zone_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                            description='Filter by zone ID'),
            openapi.Parameter('territory_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                            description='Filter by territory ID'),
            openapi.Parameter('location', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False,
                            description='Filter by location'),
            openapi.Parameter('presence_of_zm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False,
                             description='Filter by presence of Zone Manager (ZM)'),
            openapi.Parameter('presence_of_rsm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False,
                             description='Filter by presence of Regional Sales Manager (RSM)'),
            openapi.Parameter('user_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                             description='Filter by User ID (Created By) - Use to see specific user records.'),
        ]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Retrieve detailed information of a specific farmer advisory meeting including all attendees and attachments.",
        responses={
            200: openapi.Response(
                description='Meeting details with attendees and attachments',
                examples={
                    'application/json': {
                        'id': 1,
                        'fsm_name': 'Ahmed Ali',
                        'company_id': 1,
                        'company_name': 'ABC Agriculture',
                        'region_id': 1,
                        'region_name': 'Punjab',
                        'zone_id': 1,
                        'zone_name': 'Lahore',
                        'territory_id': 1,
                        'territory_name': 'Model Town',
                        'date': '2024-01-15',
                        'location': 'Community Center, Model Town',
                        'total_attendees': 25,
                        'key_topics_discussed': 'Crop rotation, pest management, irrigation techniques',
                        'presence_of_zm': True,
                        'presence_of_rsm': True,
                        'feedback_from_attendees': 'Very informative session, farmers appreciated practical tips',
                        'suggestions_for_future': 'More practical demonstrations, field visits',
                        'attendees': [
                            {
                                'name': 'Farmer John',
                                'contact': '+92-300-1234567',
                                'acreage': 5.5,
                                'crop': 'Wheat'
                            },
                            {
                                'name': 'Farmer Ali',
                                'contact': '+92-301-7654321',
                                'acreage': 3.2,
                                'crop': 'Rice'
                            }
                        ],
                        'attachments': [
                            {
                                'id': 1,
                                'file': '/media/meetings/presentation.pdf',
                                'uploaded_at': '2024-01-15T10:30:00Z'
                            }
                        ]
                    }
                }
            ),
            404: 'Meeting not found'
        }
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    # ---------------- Export to Excel ----------------
    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Export filtered farmer advisory meetings to Excel format with all attendee details. Use 'id' parameter to download a specific record. Respects all applied filters and returns an .xlsx file.",
        manual_parameters=[
            openapi.Parameter('id', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Download specific meeting by ID (e.g., FM123ABC)'),
            openapi.Parameter('fsm_name', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by FSM name'),
            openapi.Parameter('company_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by company ID'),
            openapi.Parameter('region_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by region ID'),
            openapi.Parameter('zone_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by zone ID'),
            openapi.Parameter('territory_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by territory ID'),
            openapi.Parameter('location', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by location'),
            openapi.Parameter('presence_of_zm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by presence of Zone Manager'),
            openapi.Parameter('presence_of_rsm', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by presence of Regional Sales Manager'),
            openapi.Parameter('user_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by User ID (Created By)'),
            openapi.Parameter('search', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Search in FSM name, region, zone, territory, location, topics'),
            openapi.Parameter('ordering', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Order by field (e.g., date, -date, fsm_name, -id)'),
        ],
        responses={
            200: openapi.Response(
                description='Excel file download',
                schema=openapi.Schema(
                    type=openapi.TYPE_FILE
                )
            )
        }
    )
    @action(detail=False, methods=['get'])
    def export_to_excel(self, request):
        """Export filtered meetings to Excel"""
        from datetime import datetime
        # Get filtered queryset based on current query params
        queryset = self.filter_queryset(self.get_queryset())
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Farmer Meetings"
        
        # Add export date/time at the top
        export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"Farmer Advisory Meeting Export - {export_time}"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:R1')
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        data_alignment = Alignment(vertical="top", wrap_text=True)
        
        # Headers - Meeting Info
        main_headers = [
            'ID', 'FSM Name', 'Date', 'Company', 'Region', 'Zone', 'Territory',
            'Location', 'Total Attendees', 'ZM Present', 'RSM Present',
            'Key Topics', 'Feedback', 'Suggestions'
        ]
        
        # Attendee headers
        attendee_headers = ['Attendee Name', 'Contact Number', 'Acreage', 'Crop']
        
        # Write main headers in row 3 (columns A-N)
        for col_num, header in enumerate(main_headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write attendee headers (columns O-R)
        for col_num, header in enumerate(attendee_headers, len(main_headers) + 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        row_num = 4
        for meeting in queryset:
            attendees = meeting.attendees.all()
            
            # Meeting main info (only once per meeting)
            meeting_data = [
                meeting.id, meeting.fsm_name,
                meeting.date.strftime('%Y-%m-%d %H:%M') if meeting.date else '',
                meeting.company_fk.Company_name if meeting.company_fk else '',
                meeting.region_fk.name if meeting.region_fk else '',
                meeting.zone_fk.name if meeting.zone_fk else '',
                meeting.territory_fk.name if meeting.territory_fk else '',
                meeting.location, meeting.total_attendees,
                'Yes' if meeting.presence_of_zm else 'No',
                'Yes' if meeting.presence_of_rsm else 'No',
                meeting.key_topics_discussed,
                meeting.feedback_from_attendees or '',
                meeting.suggestions_for_future or ''
            ]
            
            is_odd_row = (row_num - 4) % 2 == 0
            row_fill = odd_row_fill if is_odd_row else even_row_fill
            
            if attendees.exists():
                first_row = True
                for attendee in attendees:
                    # Write meeting info only in the first row
                    if first_row:
                        for col_num, value in enumerate(meeting_data, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            cell.fill = row_fill
                            cell.border = thin_border
                            cell.alignment = data_alignment
                        first_row = False
                    else:
                        # Leave meeting columns empty for subsequent attendees
                        for col_num in range(1, len(main_headers) + 1):
                            cell = ws.cell(row=row_num, column=col_num, value='')
                            cell.fill = row_fill
                            cell.border = thin_border
                    
                    # Write attendee info
                    for col_num, value in enumerate([
                        attendee.farmer_name, attendee.contact_number,
                        attendee.acreage, attendee.crop
                    ], len(main_headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.fill = row_fill
                        cell.border = thin_border
                        cell.alignment = data_alignment
                    row_num += 1
            else:
                # No attendees - just write meeting info
                for col_num, value in enumerate(meeting_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = data_alignment
                row_num += 1
        
        # Adjust column widths
        from openpyxl.cell.cell import MergedCell
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                if col_letter is None:
                    col_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if col_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = 'A4'
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=farmer_meetings.xlsx'
        wb.save(response)
        return response

    # ---------------- Create ----------------
    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Create a new farmer advisory meeting with multiple attendees and file attachments. Supports bulk attendee data entry.",
        manual_parameters=common_parameters,
        responses={
            201: openapi.Response(
                description='Meeting created successfully',
                examples={
                    'application/json': {
                        'id': 1,
                        'fsm_name': 'Ahmed Ali',
                        'company_id': 1,
                         'company_name': 'ABC Agriculture',
                         'region_id': 1,
                         'region_name': 'Punjab',
                         'zone_id': 1,
                         'zone_name': 'Lahore',
                         'territory_id': 1,
                         'territory_name': 'Model Town',
                        'date': '2024-01-15',
                        'location': 'Community Center',
                        'total_attendees': 25,
                        'presence_of_zm': True,
                        'presence_of_rsm': False,
                        'created_at': '2024-01-15T10:30:00Z'
                    }
                }
            ),
            400: 'Bad Request - Invalid data or missing required fields'
        }
    )
    def create(self, request, *args, **kwargs):
        # DEBUG: Log raw request data
        print("=== FieldDayViewSet CREATE DEBUG ===")
        print(f"Raw request.data: {request.data}")
        print(f"attendee_farmer_id: {request.data.get('attendee_farmer_id', 'NOT_PROVIDED')}")
        print(f"attendee_crop: {request.data.get('attendee_crop', 'NOT_PROVIDED')}")
        print(f"attendee_farmer_id type: {type(request.data.get('attendee_farmer_id'))}")
        print(f"attendee_crop type: {type(request.data.get('attendee_crop'))}")
        if 'attendee_crop' in request.data:
            crop_data = request.data.get('attendee_crop')
            if isinstance(crop_data, list):
                for i, crop in enumerate(crop_data):
                    print(f"  attendee_crop[{i}]: '{crop}' (type: {type(crop)})")
        print("=== END DEBUG ===")
        
        # Automatically assign company based on logged-in user's sales profile
        if hasattr(request.user, 'sales_profile') and request.user.sales_profile:
            # Get the first company from user's sales profile (many-to-many relationship)
            user_companies = request.user.sales_profile.companies.all()
            if user_companies.exists() and 'company_id' not in request.data:
                # Create a mutable copy of request.data
                data = request.data.copy()
                data['company_id'] = user_companies.first().id
                request._full_data = data
        
        return super().create(request, *args, **kwargs)

    # ---------------- Update ----------------
    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Update an existing Farmer Meeting (attendees + files supported).",
        manual_parameters=common_parameters,
        responses={200: MeetingSerializer}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    # ---------------- Partial Update ----------------
    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Partial update of a Farmer Meeting.",
        manual_parameters=common_parameters,
        responses={200: MeetingSerializer}
    )
    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    # ---------------- Delete ----------------
    @swagger_auto_schema(
        tags=["12. Farmer Advisory Meeting"],
        operation_description="Delete a Farmer Meeting.",
        responses={204: "Meeting deleted"}
    )
    def destroy(self, request, *args, **kwargs):
        meeting = self.get_object()
        meeting.delete()
        return Response(
            {"detail": "Meeting deleted."}, 
            status=status.HTTP_204_NO_CONTENT
        )
        
# ------------------------------------------------
# Field Day ViewSet (similar structure to MeetingViewSet)
# ------------------------------------------------
class FieldDayViewSet(viewsets.ModelViewSet):
    queryset = FieldDay.objects.select_related(
        'region_fk', 'zone_fk', 'territory_fk', 'company_fk'
    ).filter(is_active=True)
    serializer_class = FieldDaySerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    # Filters, search, ordering
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["region_fk", "zone_fk", "territory_fk", "company_fk", "location", "total_participants", "user"]
    search_fields = ["title", "region_fk__name", "zone_fk__name", "territory_fk__name", "company_fk__Company_name", "location", "feedback"]
    ordering_fields = ["date", "title", "region_fk__name", "zone_fk__name", "territory_fk__name"]
    ordering = ["-id"]
    
    def get_queryset(self):
        """
        Filter field days based on user's geographic hierarchy AND reporting hierarchy.
        Same logic as MeetingViewSet - users see their data + subordinates' data.
        """
        from django.db.models import Q
        
        queryset = super().get_queryset()
        user = self.request.user
        
        # During schema generation, return empty queryset for AnonymousUser
        if not user.is_authenticated:
            return queryset.none()
        
        # Superusers see everything
        if user.is_superuser or user.is_staff:
            return queryset
        
        # Get user's sales profile
        sales_profile = getattr(user, 'sales_profile', None)
        if not sales_profile:
            # Non-sales staff see only their own field days
            return queryset.filter(user=user)
        
        # Build filter for geographic hierarchy
        geo_filter = Q()
        
        # Add filters for assigned geographic areas
        company_ids = list(sales_profile.companies.values_list('id', flat=True))
        region_ids = list(sales_profile.regions.values_list('id', flat=True))
        zone_ids = list(sales_profile.zones.values_list('id', flat=True))
        territory_ids = list(sales_profile.territories.values_list('id', flat=True))
        
        if company_ids:
            geo_filter |= Q(company_fk_id__in=company_ids)
        if region_ids:
            geo_filter |= Q(region_fk_id__in=region_ids)
        if zone_ids:
            geo_filter |= Q(zone_fk_id__in=zone_ids)
        if territory_ids:
            geo_filter |= Q(territory_fk_id__in=territory_ids)
        
        # Build filter for reporting hierarchy (subordinates)
        reporting_filter = Q()
        
        # Get all subordinates' user IDs
        subordinate_users = sales_profile.get_subordinate_users()
        subordinate_user_ids = list(subordinate_users.values_list('id', flat=True))
        
        # Get all subordinates' territory IDs
        subordinate_territory_ids = sales_profile.get_subordinate_territory_ids()
        
        if subordinate_user_ids:
            reporting_filter |= Q(user_id__in=subordinate_user_ids)
        if subordinate_territory_ids:
            reporting_filter |= Q(territory_fk_id__in=subordinate_territory_ids)
        
        # Include own field days
        own_filter = Q(user=user)
        
        # Combine all filters with OR
        final_filter = own_filter | geo_filter | reporting_filter
        
        return queryset.filter(final_filter).distinct()
    
     # ---------------- Global Extra Data ----------------
    # Removed finalize_response method to eliminate verbose nested data for companies, regions, zones, and territories
    # The API now returns only IDs and names for these entities to reduce response size
    # ---------------- Common Parameters ----------------
    common_parameters = [
        openapi.Parameter('fsm_name', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True,
                          description='Name of FSM'),
        openapi.Parameter('company_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                          description='Company ID (Foreign Key) - Automatically assigned from logged-in user if not provided'),
        openapi.Parameter('region_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                          description='Region ID (Foreign Key)'),
        openapi.Parameter('zone_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                          description='Zone ID (Foreign Key)'),
        openapi.Parameter('territory_id', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                          description='Territory ID (Foreign Key)'),
        openapi.Parameter('date', openapi.IN_FORM, type=openapi.TYPE_STRING, format='date', required=True,
                          description='Field Day date (YYYY-MM-DD)'),
        openapi.Parameter('location', openapi.IN_FORM, type=openapi.TYPE_STRING, required=True,
                          description='Field Day location'),
        openapi.Parameter('total_participants', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                          description='Total number of participants in the field day'),
        openapi.Parameter('demonstrations_conducted', openapi.IN_FORM, type=openapi.TYPE_INTEGER, required=False,
                          description='Number of demonstrations conducted during the Field Day'),
        openapi.Parameter('feedback', openapi.IN_FORM, type=openapi.TYPE_STRING, required=False,
                          description='Feedback about the Field Day'),
        openapi.Parameter('attachments', openapi.IN_FORM, type=openapi.TYPE_FILE, required=False,
                          description='File attachments (can upload multiple photos/documents)'),

        # Multiple attendees - can link existing farmers or add manual entries
        openapi.Parameter('attendee_farmer_id', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                          items=openapi.Items(type=openapi.TYPE_INTEGER), required=False,
                          description='List of farmer IDs to link existing farmers. When provided, attendee_name, attendee_contact, and attendee_acreage are auto-filled from farmer records. Example: [1, 2, 3]'),
        openapi.Parameter('attendee_name', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                          items=openapi.Items(type=openapi.TYPE_STRING), required=False,
                          description='List of farmer names. Auto-filled from linked farmer records when attendee_farmer_id is provided. For manual entry (when not linking existing farmers), provide names directly. Example: ["Ali", "Ahmed"]'),
        openapi.Parameter('attendee_contact', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                          items=openapi.Items(type=openapi.TYPE_STRING), required=False,
                          description='List of contact numbers. Auto-filled from linked farmer records when attendee_farmer_id is provided. Can be overridden for additional contact info or manual entry.'),
        openapi.Parameter('attendee_acreage', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                          items=openapi.Items(type=openapi.TYPE_NUMBER), required=False,
                          description='List of acreage values for each attendee. Auto-filled from farmer total_land_area when attendee_farmer_id is provided and acreage is 0.0. Can be overridden with specific values.'),
        openapi.Parameter('attendee_crop', openapi.IN_FORM, type=openapi.TYPE_ARRAY,
                          items=openapi.Items(type=openapi.TYPE_STRING), required=False,
                          description='List of crops for each farmer attendee. Each entry corresponds to one farmer attendee (comma-separated for multiple crops per farmer). Example: ["rice,wheat", "cotton", "potato,sugarcane"] - farmer 1 has rice and wheat, farmer 2 has cotton, farmer 3 has potato and sugarcane'),
    ]

    # ---------------- List ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Retrieve a list of all field day events with comprehensive filtering, search, and ordering capabilities. Field days are educational events for farmers.",
        responses={
            200: openapi.Response(
                description='List of field day events',
                examples={
                    'application/json': [
                        {
                            'id': 1,
                            'fsm_name': 'Ahmed Hassan',
                            'company_id': 1,
                            'company_name': 'ABC Agriculture',
                            'region_id': 1,
                            'region_name': 'Punjab',
                            'zone_id': 1,
                            'zone_name': 'Lahore',
                            'territory_id': 1,
                            'territory_name': 'Model Town',
                            'date': '2024-02-15T10:00:00Z',
                            'location': 'Agricultural Research Center',
                            'total_participants': 15,
                            'demonstrations_conducted': 3,
                            'feedback': 'Successful event with high farmer participation and positive feedback',
                            'attendees': [
                                {
                                    'id': 1,
                                    'farmer': 15,
                                    'farmer_id': 'F-2024-001',
                                    'farmer_full_name': 'Hassan Ali Khan',
                                    'farmer_primary_phone': '+92-300-9876543',
                                    'farmer_district': 'Lahore',
                                    'farmer_village': 'Model Town',
                                    'farmer_name': 'Hassan Ali Khan',
                                    'contact_number': '+92-300-9876543',
                                    'acreage': 8.0,
                                    'crop': 'Cotton',
                                    'crops': [
                                        {
                                            'id': 1,
                                            'crop_name': 'Cotton',
                                            'acreage': 8.0
                                        }
                                    ]
                                }
                            ],
                            'attachments': [
                                {
                                    'id': 1,
                                    'file': '/media/field_day_uploads/field_photo_1.jpg',
                                    'uploaded_at': '2024-02-15T14:30:00Z'
                                }
                            ]
                        }
                    ]
                }
            )
        },
        manual_parameters=[
            openapi.Parameter('user', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                             description='Filter by User ID (Created By) - Use to see specific user records.'),
            openapi.Parameter('fsm_name', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False,
                              description='Filter by Name of FSM'),
            openapi.Parameter('company_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                              description='Filter by company ID'),
            openapi.Parameter('region_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                              description='Filter by region ID'),
            openapi.Parameter('zone_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                              description='Filter by zone ID'),
            openapi.Parameter('territory_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                              description='Filter by territory ID'),
            openapi.Parameter('location', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False,
                              description='Filter by location'),
            openapi.Parameter('total_participants', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False,
                              description='Filter by total number of participants'),
        ]
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    # ---------------- Retrieve ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Retrieve detailed information of a specific field day event including all attendees and event outcomes.",
        responses={
            200: openapi.Response(
                description='Field day details with attendees',
                examples={
                    'application/json': {
                        'id': 1,
                        'fsm_name': 'Ahmed Hassan',
                        'company_id': 1,
                        'company_name': 'ABC Agriculture',
                        'region_id': 1,
                        'region_name': 'Punjab',
                        'zone_id': 1,
                        'zone_name': 'Lahore',
                        'territory_id': 1,
                        'territory_name': 'Model Town',
                        'date': '2024-02-15T10:00:00Z',
                        'location': 'Agricultural Research Center, Lahore',
                        'total_participants': 25,
                        'demonstrations_conducted': 5,
                        'feedback': 'Successful event with high farmer participation. 95% satisfaction rate. Farmers appreciated the practical demonstrations.',
                        'attendees': [
                            {
                                'id': 1,
                                'farmer': 15,
                                'farmer_id': 'F-2024-001',
                                'farmer_full_name': 'Hassan Ali Khan',
                                'farmer_primary_phone': '+92-300-9876543',
                                'farmer_district': 'Lahore',
                                'farmer_village': 'Model Town',
                                'farmer_name': 'Hassan Ali Khan',
                                'contact_number': '+92-300-9876543',
                                'acreage': 8.0,
                                'crop': 'Cotton',
                                'crops': [
                                    {
                                        'id': 1,
                                        'crop_name': 'Cotton',
                                        'acreage': 8.0
                                    }
                                ]
                            },
                            {
                                'id': 2,
                                'farmer': 22,
                                'farmer_id': 'F-2024-002',
                                'farmer_full_name': 'Fatima Bibi',
                                'farmer_primary_phone': '+92-301-1122334',
                                'farmer_district': 'Lahore',
                                'farmer_village': 'Johar Town',
                                'farmer_name': 'Fatima Bibi',
                                'contact_number': '+92-301-1122334',
                                'acreage': 4.5,
                                'crop': 'Sugarcane,Rice',
                                'crops': [
                                    {
                                        'id': 2,
                                        'crop_name': 'Sugarcane',
                                        'acreage': 2.5
                                    },
                                    {
                                        'id': 3,
                                        'crop_name': 'Rice',
                                        'acreage': 2.0
                                    }
                                ]
                            }
                        ],
                        'attachments': [
                            {
                                'id': 1,
                                'file': '/media/field_day_uploads/field_photo_1.jpg',
                                'uploaded_at': '2024-02-15T14:30:00Z'
                            },
                            {
                                'id': 2,
                                'file': '/media/field_day_uploads/field_photo_2.jpg',
                                'uploaded_at': '2024-02-15T15:00:00Z'
                            }
                        ],
                        'created_at': '2024-02-10T09:00:00Z',
                        'updated_at': '2024-02-15T17:00:00Z'
                    }
                }
            ),
            404: 'Field day not found'
        }
    )
    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    # ---------------- Export to Excel ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Export filtered field days to Excel format with all attendee details. Use 'id' parameter to download a specific record. Respects all applied filters and returns an .xlsx file.",
        manual_parameters=[
            openapi.Parameter('id', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Download specific field day by ID (e.g., FD123ABC)'),
            openapi.Parameter('title', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by title/FSM name'),
            openapi.Parameter('company_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by company ID'),
            openapi.Parameter('region_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by region ID'),
            openapi.Parameter('zone_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by zone ID'),
            openapi.Parameter('territory_fk', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by territory ID'),
            openapi.Parameter('location', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Filter by location'),
            openapi.Parameter('total_participants', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by total participants'),
            openapi.Parameter('is_active', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, required=False, description='Filter by active status'),
            openapi.Parameter('user', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, required=False, description='Filter by User ID (Created By)'),
            openapi.Parameter('search', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Search in title, company, region, zone, territory, location'),
            openapi.Parameter('ordering', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False, description='Order by field (e.g., date, -date, title, -id)'),
        ],
        responses={
            200: openapi.Response(
                description='Excel file download',
                schema=openapi.Schema(
                    type=openapi.TYPE_FILE
                )
            )
        }
    )
    @action(detail=False, methods=['get'])
    def export_to_excel(self, request):
        """Export filtered field days to Excel"""
        from datetime import datetime
        # Get filtered queryset based on current query params
        queryset = self.filter_queryset(self.get_queryset())
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Field Days"
        
        # Add export date/time at the top
        export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"Field Day Export - {export_time}"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        ws.merge_cells('A1:P1')
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        data_alignment = Alignment(vertical="top", wrap_text=True)
        
        # Headers - Field Day Info
        main_headers = [
            'ID', 'Title', 'Date', 'Company', 'Region', 'Zone', 'Territory',
            'Total Participants', 'Demonstrations Conducted', 'User', 'Active', 'Feedback'
        ]
        
        # Attendee headers
        attendee_headers = ['Attendee Name', 'Contact Number', 'Acreage', 'Crop']
        
        # Write main headers in row 3 (columns A-L)
        for col_num, header in enumerate(main_headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write attendee headers (columns M-P)
        for col_num, header in enumerate(attendee_headers, len(main_headers) + 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        row_num = 4
        for field_day in queryset:
            attendees = field_day.attendees.all()
            
            # Field day main info (only once per field day)
            field_day_data = [
                field_day.id, field_day.title,
                field_day.date.strftime('%Y-%m-%d %H:%M') if field_day.date else '',
                field_day.company_fk.Company_name if field_day.company_fk else '',
                field_day.region_fk.name if field_day.region_fk else '',
                field_day.zone_fk.name if field_day.zone_fk else '',
                field_day.territory_fk.name if field_day.territory_fk else '',
                field_day.total_participants, field_day.demonstrations_conducted,
                field_day.user.username if field_day.user else '',
                'Yes' if field_day.is_active else 'No',
                field_day.feedback or ''
            ]
            
            is_odd_row = (row_num - 4) % 2 == 0
            row_fill = odd_row_fill if is_odd_row else even_row_fill
            
            if attendees.exists():
                first_row = True
                for attendee in attendees:
                    # Write field day info only in the first row
                    if first_row:
                        for col_num, value in enumerate(field_day_data, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            cell.fill = row_fill
                            cell.border = thin_border
                            cell.alignment = data_alignment
                        first_row = False
                    else:
                        # Leave field day columns empty for subsequent attendees
                        for col_num in range(1, len(main_headers) + 1):
                            cell = ws.cell(row=row_num, column=col_num, value='')
                            cell.fill = row_fill
                            cell.border = thin_border
                    
                    # Write attendee info
                    for col_num, value in enumerate([
                        attendee.farmer_name, attendee.contact_number,
                        attendee.acreage, attendee.crop
                    ], len(main_headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.fill = row_fill
                        cell.border = thin_border
                        cell.alignment = data_alignment
                    row_num += 1
            else:
                # No attendees - just write field day info
                for col_num, value in enumerate(field_day_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = data_alignment
                row_num += 1
        
        # Adjust column widths
        from openpyxl.cell.cell import MergedCell
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                if col_letter is None:
                    col_letter = cell.column_letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if col_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = 'A4'
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=field_days.xlsx'
        wb.save(response)
        return response

    # ---------------- Create ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Create a new Field Day with attendees.",
        manual_parameters=common_parameters,
        responses={201: FieldDaySerializer}
    )
    def create(self, request, *args, **kwargs):
        # Automatically assign company based on logged-in user's sales profile
        if hasattr(request.user, 'sales_profile') and request.user.sales_profile:
            # Get the first company from user's sales profile (many-to-many relationship)
            user_companies = request.user.sales_profile.companies.all()
            if user_companies.exists() and 'company_id' not in request.data:
                # Create a mutable copy of request.data
                data = request.data.copy()
                data['company_id'] = user_companies.first().id
                request._full_data = data
        
        return super().create(request, *args, **kwargs)

    # ---------------- Update ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Update a Field Day with attendees.",
        manual_parameters=common_parameters,
        responses={200: FieldDaySerializer}
    )
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    # ---------------- Partial Update ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Partially update a Field Day.",
        manual_parameters=common_parameters,
        responses={200: FieldDaySerializer}
    )
    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    # ---------------- Delete ----------------
    @swagger_auto_schema(
        tags=["13. Field Day"],
        operation_description="Soft delete a Field Day by setting is_active to False.",
        responses={204: "Field Day deleted"}
    )
    def destroy(self, request, *args, **kwargs):
        field_day = self.get_object()
        field_day.is_active = False
        field_day.save()
        return Response({"detail": "Field Day deleted."}, status=status.HTTP_204_NO_CONTENT)

# NOTE: This is a template for your products catalog serializer.
# Apply this pattern to your actual serializer (not shown in the provided code).

class ProductCatalogSerializer(serializers.ModelSerializer):
    # ...existing code...

    def to_representation(self, instance):
        data = super().to_representation(instance)
        # Ensure these fields are never null
        for field in [
            "product_image_url",
            "product_description_urdu_url",
            "Product_Image_Name",
            "Product_Image_Ext",
            "Product_Urdu_Name",
            "Product_Urdu_Ext"
        ]:
            if data.get(field) is None:
                data[field] = ""
        return data

    # ...existing code...
