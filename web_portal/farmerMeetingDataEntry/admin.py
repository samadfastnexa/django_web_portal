from django.contrib import admin
from web_portal.admin import admin_site
from .models import Meeting, FarmerAttendance, MeetingAttachment, FieldDay, FieldDayAttendance, FieldDayAttachment, FieldDayAttendanceCrop
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db import models
from django.contrib.admin import widgets as admin_widgets
from datetime import datetime

class FarmerAttendanceInline(admin.TabularInline):
    model = FarmerAttendance
    extra = 1
    fields = ('farmer', 'farmer_name', 'contact_number', 'acreage', 'crop')
    readonly_fields = ()
    autocomplete_fields = ['farmer']
    
    class Media:
        js = ('admin/js/farmer_autocomplete.js',)


class MeetingAttachmentInline(admin.TabularInline):
    model = MeetingAttachment
    extra = 1

def export_farmer_meeting_to_excel(modeladmin, request, queryset):
    """Export selected Farmer Meetings to Excel with attendance details"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Farmer Meetings"
    
    # Add export date/time at the top
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Farmer Advisory Meeting Export - {export_time}"
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells('A1:R1')
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    # Headers - Meeting Info
    main_headers = [
        'ID', 'FSM Name', 'Date', 'Company', 'Region', 'Zone', 'Territory',
        'Location', 'Total Attendees', 'ZM Present', 'RSM Present',
        'Key Topics', 'Feedback', 'Suggestions'
    ]
    
    # Attendee headers
    attendee_headers = ['Attendee Name', 'Contact Number', 'Acreage', 'Crop']
    
    # Write main headers in row 3 (columns A-N)
    for col_num, header in enumerate(main_headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write attendee headers (columns O-R)
    for col_num, header in enumerate(attendee_headers, len(main_headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    row_num = 4
    for meeting in queryset:
        attendees = meeting.attendees.all()
        
        # Meeting main info (only once per meeting)
        meeting_data = [
            meeting.id, meeting.fsm_name,
            meeting.date.strftime('%Y-%m-%d %H:%M') if meeting.date else '',
            meeting.company_fk.Company_name if meeting.company_fk else '',
            meeting.region_fk.name if meeting.region_fk else '',
            meeting.zone_fk.name if meeting.zone_fk else '',
            meeting.territory_fk.name if meeting.territory_fk else '',
            meeting.location, meeting.total_attendees,
            'Yes' if meeting.presence_of_zm else 'No',
            'Yes' if meeting.presence_of_rsm else 'No',
            meeting.key_topics_discussed,
            meeting.feedback_from_attendees or '',
            meeting.suggestions_for_future or ''
        ]
        
        is_odd_row = (row_num - 4) % 2 == 0
        row_fill = odd_row_fill if is_odd_row else even_row_fill
        
        if attendees.exists():
            first_row = True
            for attendee in attendees:
                # Write meeting info only in the first row
                if first_row:
                    for col_num, value in enumerate(meeting_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.fill = row_fill
                        cell.border = thin_border
                        cell.alignment = data_alignment
                    first_row = False
                else:
                    # Leave meeting columns empty for subsequent attendees
                    for col_num in range(1, len(main_headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num, value='')
                        cell.fill = row_fill
                        cell.border = thin_border
                
                # Write attendee info
                for col_num, value in enumerate([
                    attendee.farmer_name, attendee.contact_number,
                    attendee.acreage, attendee.crop
                ], len(main_headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = data_alignment
                row_num += 1
        else:
            # No attendees - just write meeting info
            for col_num, value in enumerate(meeting_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = data_alignment
            row_num += 1
    
    # Adjust column widths
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if col_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[3].height = 30
    
    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A4'
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=farmer_meetings.xlsx'
    wb.save(response)
    return response

export_farmer_meeting_to_excel.short_description = "Export selected to Excel"

@admin.register(Meeting, site=admin_site)
class MeetingAdmin(admin.ModelAdmin):
    inlines = [FarmerAttendanceInline, MeetingAttachmentInline]

    list_display = [
        'id',
        'fsm_name',
        'formatted_date',
        'region_fk',
        'zone_fk',
        'territory_fk',
        'total_attendees',
    ]

    search_fields = [
        'fsm_name',
        'region_fk__name',
        'zone_fk__name',
        'territory_fk__name',
        'location',
    ]

    list_filter = [
        'region_fk',
        'zone_fk',
        'territory_fk',
        ('date', admin.DateFieldListFilter),
    ]
    ordering = ['-id']
    actions = [export_farmer_meeting_to_excel]
    
    # Configure form to show datetime input with separate date and time fields
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'date':
            kwargs['widget'] = admin_widgets.AdminSplitDateTime()
        return super().formfield_for_dbfield(db_field, request, **kwargs)
    
    def formatted_date(self, obj):
        """Display date with time"""
        if obj.date:
            return obj.date.strftime('%Y-%m-%d %H:%M')
        return '-'
    formatted_date.short_description = 'Date & Time'
    formatted_date.admin_order_field = 'date'


class FieldDayAttendanceInline(admin.TabularInline):
    model = FieldDayAttendance
    extra = 1
    fields = ('farmer', 'farmer_name', 'contact_number', 'acreage', 'crop')
    readonly_fields = ()
    autocomplete_fields = ['farmer']
    
    class Media:
        js = ('admin/js/farmer_autocomplete.js',)
    
    def get_readonly_fields(self, request, obj=None):
        """Make acreage readonly when farmer is linked and has total_land_area"""
        readonly = list(self.readonly_fields)
        return readonly

class FieldDayAttachmentInline(admin.TabularInline):
    model = FieldDayAttachment
    extra = 1

def export_field_day_to_excel(modeladmin, request, queryset):
    """Export selected Field Days to Excel with attendance details"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Days"
    
    # Add export date/time at the top
    export_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Field Day Export - {export_time}"
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells('A1:D1')
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    info_label_font = Font(bold=True, size=10, color="1F4E78")
    
    odd_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    current_row = 3
    
    for field_day in queryset:
        # Field Day Information Section
        info_fields = [
            ('ID', field_day.id),
            ('Title', field_day.title),
            ('Date', field_day.date.strftime('%Y-%m-%d %H:%M') if field_day.date else ''),
            ('Company', field_day.company_fk.Company_name if field_day.company_fk else ''),
            ('Region', field_day.region_fk.name if field_day.region_fk else ''),
            ('Zone', field_day.zone_fk.name if field_day.zone_fk else ''),
            ('Territory', field_day.territory_fk.name if field_day.territory_fk else ''),
            ('Total Participants', field_day.total_participants),
            ('Demonstrations Conducted', field_day.demonstrations_conducted),
            ('User', field_day.user.username if field_day.user else ''),
            ('Active', 'Yes' if field_day.is_active else 'No'),
            ('Feedback', field_day.feedback or '')
        ]
        
        # Write field day information as key-value pairs
        for label, value in info_fields:
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.fill = info_label_fill
            label_cell.font = info_label_font
            label_cell.border = thin_border
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            value_cell = ws.cell(row=current_row, column=2, value=value)
            value_cell.border = thin_border
            value_cell.alignment = data_alignment
            ws.merge_cells(f'B{current_row}:D{current_row}')
            
            current_row += 1
        
        current_row += 1  # Empty row
        
        # Attendee table headers
        attendee_headers = ['Attendee Name', 'Contact Number', 'Acreage', 'Crop']
        for col_num, header in enumerate(attendee_headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        current_row += 1
        
        # Write attendees
        attendees = field_day.attendees.all()
        if attendees.exists():
            for idx, attendee in enumerate(attendees):
                is_odd_row = idx % 2 == 0
                row_fill = odd_row_fill if is_odd_row else even_row_fill
                
                attendee_data = [
                    attendee.farmer_name,
                    attendee.contact_number,
                    attendee.acreage,
                    attendee.crop
                ]
                
                for col_num, value in enumerate(attendee_data, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=value)
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = data_alignment
                
                current_row += 1
        else:
            # No attendees message
            no_attendee_cell = ws.cell(row=current_row, column=1, value="No attendees")
            no_attendee_cell.font = Font(italic=True, color="999999")
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
        
        current_row += 2  # Space before next field day
    
    # Adjust column widths
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        if col_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=field_days.xlsx'
    wb.save(response)
    return response

export_field_day_to_excel.short_description = "Export selected to Excel"

@admin.register(FieldDay, site=admin_site)
class FieldDayAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'title', 'company_fk', 'territory_fk', 'zone_fk', 'region_fk', 
        'formatted_date', 'total_participants', 'demonstrations_conducted', 'user', 'is_active'
    )
    list_filter = (
        'company_fk', 'region_fk', 'zone_fk', 'territory_fk', 
        ('date', admin.DateFieldListFilter), 'total_participants', 'demonstrations_conducted', 'is_active'
    )
    search_fields = (
        'id', 'title', 'company_fk__Company_name', 'territory_fk__name', 
        'zone_fk__name', 'region_fk__name', 'user__email', 'feedback'
    )
    readonly_fields = ('id',)
    ordering = ['-id']
    inlines = [FieldDayAttendanceInline, FieldDayAttachmentInline]
    actions = [export_field_day_to_excel]
    
    # Configure form to show datetime input with separate date and time fields
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        if db_field.name == 'date':
            kwargs['widget'] = admin_widgets.AdminSplitDateTime()
        return super().formfield_for_dbfield(db_field, request, **kwargs)
    
    def formatted_date(self, obj):
        """Display date with time"""
        if obj.date:
            return obj.date.strftime('%Y-%m-%d %H:%M')
        return '-'
    formatted_date.short_description = 'Date & Time'
    formatted_date.admin_order_field = 'date'

class FieldDayAttendanceCropInline(admin.TabularInline):
    model = FieldDayAttendanceCrop
    extra = 1
    fields = ['crop_name', 'acreage']


@admin.register(FieldDayAttendance, site=admin_site)
class FieldDayAttendanceAdmin(admin.ModelAdmin):
    list_display = ('field_day', 'farmer_info', 'farmer_name', 'contact_number', 'acreage', 'crop', 'get_crops_display')
    search_fields = ('farmer__farmer_id', 'farmer__first_name', 'farmer__last_name', 'farmer_name', 'contact_number', 'crop')
    list_filter = ('crop', 'farmer__district', 'farmer__village', 'field_day__date')
    autocomplete_fields = ['farmer']
    readonly_fields = ('farmer_name', 'contact_number')  # Auto-filled from farmer
    inlines = [FieldDayAttendanceCropInline]
    fieldsets = (
        ('Farmer Linking', {
            'fields': ('farmer',),
            'description': 'Link to an existing farmer to auto-fill attendee information'
        }),
        ('Attendee Information', {
            'fields': ('farmer_name', 'contact_number', 'acreage', 'crop'),
            'description': 'Attendee details (auto-filled from linked farmer if available)'
        }),
    )
    
    def farmer_info(self, obj):
        """Display farmer ID and name if linked"""
        if obj.farmer:
            return f"{obj.farmer.farmer_id} - {obj.farmer.full_name}"
        return "Manual Entry"
    farmer_info.short_description = "Farmer Link"
    farmer_info.admin_order_field = 'farmer__farmer_id'
    
    def get_crops_display(self, obj):
        """Display all crops for this attendance"""
        crops = obj.crops.all()
        if crops:
            return ", ".join([f"{crop.crop_name} ({crop.acreage})" for crop in crops])
        return "-"
    get_crops_display.short_description = "Crops (New)"


@admin.register(FieldDayAttendanceCrop, site=admin_site)
class FieldDayAttendanceCropAdmin(admin.ModelAdmin):
    list_display = ['attendance', 'crop_name', 'acreage']
    list_filter = ['crop_name']
    search_fields = ['crop_name', 'attendance__farmer_name']
