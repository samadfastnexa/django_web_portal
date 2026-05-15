"""
Separate Journal Entry Ledger export API using SAP Business One Service Layer.
This module intentionally does not reuse SQL/HANA ledger query paths.
"""

from datetime import datetime
from io import BytesIO
import re
import requests
import logging
import urllib3
import socket

from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

logger = logging.getLogger(__name__)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from sap_integration.sap_client import SAPClient


def _parse_sap_date(value):
    """Parse SAP Service Layer date values into datetime.date when possible."""
    if not value:
        return None

    text = str(value)

    # SAP JSON date format: /Date(1735689600000)/
    m = re.match(r"/Date\((\d+)(?:[+-]\d+)?\)/", text)
    if m:
        try:
            millis = int(m.group(1))
            return datetime.utcfromtimestamp(millis / 1000).date()
        except Exception:
            return None

    # ISO style: 2026-01-01 or 2026-01-01T00:00:00(.sss)(Z)
    cleaned = text.rstrip('Z')
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except Exception:
            continue

    # Fallback: first 10 chars as yyyy-mm-dd
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _to_float(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _fetch_journal_entries_via_requests_fallback(client, max_entries, page_size=20):
    """Fallback Service Layer fetch using requests (tries HTTPS first, then HTTP).
    
    Note: Fetches JournalEntries without $expand to reduce timeout, then fetches lines separately
    if needed. This is a trade-off for slower initial query vs. avoiding timeouts on large datasets.
    """
    entries = []
    skip = 0
    last_error = None

    schemes = ['http'] if getattr(client, 'use_http', False) else ['https', 'http']

    for scheme in schemes:
        session = requests.Session()
        base_url = f"{scheme}://{client.host}:{client.port}{client.base_path}"

        try:
            login_payload = {
                'UserName': client.username,
                'Password': client.password,
                'CompanyDB': client.company_db,
            }

            login_resp = session.post(
                f"{base_url}/Login",
                json=login_payload,
                timeout=60,
                verify=False if scheme == 'https' else True,
            )
            if login_resp.status_code >= 400:
                raise Exception(f"Service Layer login failed ({scheme.upper()}): {login_resp.status_code} {login_resp.text}")

            while len(entries) < max_entries:
                # Fetch JournalEntries WITHOUT $expand to avoid timeout on large result sets.
                # Lines will be fetched on-demand if needed.
                params = {
                    '$orderby': 'JdtNum asc',
                    '$top': page_size,
                    '$skip': skip,
                }
                resp = session.get(
                    f"{base_url}/JournalEntries",
                    params=params,
                    timeout=60,
                    verify=False if scheme == 'https' else True,
                )
                if resp.status_code >= 400:
                    raise Exception(f"JournalEntries fetch failed ({scheme.upper()}): {resp.status_code} {resp.text}")

                payload = resp.json() if resp.content else {}
                page_rows = payload.get('value', []) if isinstance(payload, dict) else []
                if not page_rows:
                    break

                # Fetch lines for each entry individually (slower but avoids expansion timeout)
                for entry in page_rows:
                    jdt_num = entry.get('JdtNum')
                    if jdt_num:
                        try:
                            lines_resp = session.get(
                                f"{base_url}/JournalEntries({jdt_num})/JournalEntryLines",
                                timeout=30,
                                verify=False if scheme == 'https' else True,
                            )
                            if lines_resp.status_code == 200:
                                lines_payload = lines_resp.json() if lines_resp.content else {}
                                entry['JournalEntryLines'] = lines_payload.get('value', [])
                        except Exception:
                            # If lines fetch fails, continue without them
                            entry['JournalEntryLines'] = []

                entries.extend(page_rows)
                skip += page_size
                if len(page_rows) < page_size:
                    break

            try:
                session.post(
                    f"{base_url}/Logout",
                    timeout=10,
                    verify=False if scheme == 'https' else True,
                )
            except Exception:
                pass

            return entries[:max_entries]

        except Exception as e:
            last_error = e
            entries = []
            skip = 0
            continue

    raise Exception(str(last_error) if last_error else 'Service Layer fallback failed')


def _fetch_entries_from_sap(client, max_entries=50, skip_lines=False, page_size=10):
    """Fetch JournalEntries from SAP with optional line expansion via per-entry call."""
    entries = []
    skip = 0

    try:
        while len(entries) < max_entries:
            query = f"$orderby=JdtNum asc&$top={page_size}&$skip={skip}"
            res = client.get('JournalEntries', query)
            page_rows = res.get('value', []) if isinstance(res, dict) else (res or [])

            if not page_rows:
                break

            if not skip_lines:
                for entry in page_rows:
                    jdt_num = entry.get('JdtNum')
                    if not jdt_num:
                        entry['JournalEntryLines'] = []
                        continue
                    try:
                        lines_res = client.get(f'JournalEntries({jdt_num})/JournalEntryLines', '')
                        entry['JournalEntryLines'] = lines_res.get('value', []) if isinstance(lines_res, dict) else (lines_res or [])
                    except Exception:
                        entry['JournalEntryLines'] = []
            else:
                for entry in page_rows:
                    entry['JournalEntryLines'] = []

            entries.extend(page_rows)
            skip += page_size
            if len(page_rows) < page_size:
                break

    except Exception as e:
        logger.error(f"[JE API] Error during entry fetch: {e}", exc_info=True)
        if 'SAP Server SSL Internal Error' in str(e) or 'TLSV1_ALERT_INTERNAL_ERROR' in str(e):
            logger.info("[JE API] Attempting fallback via requests...")
            entries = _fetch_journal_entries_via_requests_fallback(client, max_entries=max_entries, page_size=page_size)
        else:
            raise

    if len(entries) > max_entries:
        entries = entries[:max_entries]
    return entries


def _flatten_entries(entries, parsed_from, parsed_to, bp_code, account_code, project_code):
    """Flatten JE header + lines into row dictionaries with running balance."""
    rows = []
    running_balance = 0.0

    for je in entries:
        posting_date = _parse_sap_date(je.get('ReferenceDate') or je.get('RefDate'))
        if parsed_from and posting_date and posting_date < parsed_from:
            continue
        if parsed_to and posting_date and posting_date > parsed_to:
            continue

        due_date = _parse_sap_date(je.get('DueDate'))
        doc_date = _parse_sap_date(je.get('TaxDate') or je.get('DocumentDate'))

        lines = je.get('JournalEntryLines') or []
        if not isinstance(lines, list):
            lines = []

        for line in lines:
            line_account = str(line.get('AccountCode') or line.get('Account') or '').strip()
            line_bp = str(line.get('ShortName') or '').strip()
            line_project = str(line.get('ProjectCode') or line.get('Project') or '').strip()

            if account_code and line_account != account_code:
                continue
            if bp_code and line_bp != bp_code:
                continue
            if project_code and line_project != project_code:
                continue

            debit = _to_float(line.get('Debit'))
            credit = _to_float(line.get('Credit'))
            running_balance += (debit - credit)

            rows.append({
                'je_no': je.get('JdtNum') or je.get('TransId') or '',
                'posting_date': posting_date.strftime('%Y-%m-%d') if posting_date else '',
                'due_date': due_date.strftime('%Y-%m-%d') if due_date else '',
                'doc_date': doc_date.strftime('%Y-%m-%d') if doc_date else '',
                'memo': str(je.get('Memo') or ''),
                'line_no': line.get('Line_ID') or line.get('LineNum') or '',
                'account': line_account,
                'bp_code': line_bp,
                'line_memo': str(line.get('LineMemo') or ''),
                'project': line_project,
                'debit': debit if debit > 0 else 0,
                'credit': credit if credit > 0 else 0,
                'balance': running_balance,
            })

    return rows


def _tcp_probe(host, port, timeout_sec=5):
    """Return quick TCP connectivity probe for SAP Service Layer host/port."""
    try:
        with socket.create_connection((host, int(port)), timeout=timeout_sec):
            return {'ok': True, 'detail': 'tcp_connect_ok'}
    except Exception as e:
        return {'ok': False, 'detail': str(e)}


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='SAP Service Layer Health Check (Login only)',
    operation_description='Performs SAP login/session creation only. Does not query JournalEntries.',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Company DB key (e.g., 4B-AGRI)', required=False),
    ],
    responses={200: 'SAP login success', 503: 'SAP login failed'}
)
@api_view(['GET'])
def sap_journal_entry_health_api(request):
    try:
        company = (request.GET.get('company') or '').strip() or None
        client = SAPClient(company_db_key=company)
        tcp = _tcp_probe(client.host, client.port, timeout_sec=5)
        session_id = client.get_session_id()
        return Response({
            'success': True,
            'message': 'SAP Service Layer login successful',
            'company_db': client.company_db,
            'host': client.host,
            'port': client.port,
            'use_http': client.use_http,
            'tcp_probe': tcp,
            'session': bool(session_id),
        }, status=200)
    except Exception as e:
        error_text = str(e)
        lowered = error_text.lower()
        status_code = 503 if ('timed out' in lowered or 'timeout' in lowered or 'ssl' in lowered or 'tls' in lowered) else 500

        host = None
        port = None
        use_http = None
        company_db = None
        tcp = None
        try:
            host = client.host
            port = client.port
            use_http = client.use_http
            company_db = client.company_db
            tcp = _tcp_probe(client.host, client.port, timeout_sec=5)
        except Exception:
            pass

        hint = 'Check SAP Service Layer status and credentials.'
        if 'timed out' in lowered or 'timeout' in lowered:
            hint = 'Login timed out. Verify SAP Service Layer is running and reachable on host/port, and check server load.'
        elif 'ssl' in lowered or 'tls' in lowered:
            hint = 'TLS/SSL handshake failed. Verify Service Layer certificate/protocol settings or use HTTP mode if allowed.'

        logger.error(f"[JE API][HEALTH] Login failed: {error_text}")
        return Response({
            'success': False,
            'message': 'SAP Service Layer login failed',
            'error': error_text,
            'hint': hint,
            'company_db': company_db,
            'host': host,
            'port': port,
            'use_http': use_http,
            'tcp_probe': tcp,
        }, status=status_code)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Download Journal Entry Ledger PDF from SAP API',
    operation_description='Downloads Journal Entry ledger directly from SAP Business One Service Layer as PDF.',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Company DB key', required=False),
        openapi.Parameter('from_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start date (YYYY-MM-DD)', required=False),
        openapi.Parameter('to_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End date (YYYY-MM-DD)', required=False),
        openapi.Parameter('bp_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional BP code filter', required=False),
        openapi.Parameter('account_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional account code filter', required=False),
        openapi.Parameter('project_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional project code filter', required=False),
        openapi.Parameter('max_entries', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Max JournalEntries to fetch (default 50, max 500)', required=False),
    ],
    responses={200: openapi.Response(description='PDF file download', schema=openapi.Schema(type=openapi.TYPE_FILE)), 500: 'Error generating PDF'}
)
@api_view(['GET'])
def export_journal_entry_ledger_sap_pdf_api(request):
    try:
        company = (request.GET.get('company') or '').strip() or None
        from_date = (request.GET.get('from_date') or '').strip()
        to_date = (request.GET.get('to_date') or '').strip()
        bp_code = (request.GET.get('bp_code') or '').strip()
        account_code = (request.GET.get('account_code') or '').strip()
        project_code = (request.GET.get('project_code') or '').strip()

        try:
            max_entries = int(request.GET.get('max_entries', 50) or 50)
        except Exception:
            max_entries = 50
        max_entries = max(1, min(max_entries, 500))

        parsed_from = datetime.strptime(from_date, '%Y-%m-%d').date() if from_date else None
        parsed_to = datetime.strptime(to_date, '%Y-%m-%d').date() if to_date else None

        client = SAPClient(company_db_key=company)
        entries = _fetch_entries_from_sap(client=client, max_entries=max_entries, skip_lines=False, page_size=10)
        rows = _flatten_entries(entries, parsed_from, parsed_to, bp_code, account_code, project_code)

        response = HttpResponse(content_type='application/pdf')
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        company_part = company or 'default'
        response['Content-Disposition'] = f'attachment; filename="sap_journal_entry_ledger_{company_part}_{stamp}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph('SAP Journal Entry Ledger', styles['Title']))
        elements.append(Paragraph(f'Company: {company_part}', styles['Normal']))
        elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Spacer(1, 12))

        table_data = [[
            'JE No', 'Posting', 'Due', 'Doc', 'Memo', 'Ln', 'Account', 'BP', 'Line Memo', 'Project', 'Debit', 'Credit', 'Balance'
        ]]

        for r in rows:
            table_data.append([
                str(r['je_no']),
                r['posting_date'],
                r['due_date'],
                r['doc_date'],
                r['memo'][:30],
                str(r['line_no']),
                r['account'],
                r['bp_code'],
                r['line_memo'][:30],
                r['project'],
                f"{r['debit']:.2f}",
                f"{r['credit']:.2f}",
                f"{r['balance']:.2f}",
            ])

        col_widths = [45, 55, 55, 55, 90, 30, 65, 65, 95, 55, 60, 60, 65]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('ALIGN', (10, 1), (12, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    except Exception as e:
        error_text = str(e)
        lowered = error_text.lower()
        if 'timed out' in lowered or 'timeout' in lowered:
            return Response({
                'success': False,
                'error': (
                    'SAP Service Layer timeout at fourb.vdc.services:5588. '
                    'Please retry later or contact SAP administrator to verify Service Layer health.'
                )
            }, status=503)
        return Response({'success': False, 'error': error_text}, status=500)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Download Journal Entry Ledger from SAP API',
    operation_description='Downloads Journal Entry ledger (TransType=Journal Entry) directly from SAP Business One Service Layer as Excel. Note: SAP Service Layer can be slow; use smaller date ranges and max_entries for faster response.',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Company DB key (e.g., 4B-BIO, 4B-AGRI, 4B-ORANG)', required=False),
        openapi.Parameter('from_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start date (YYYY-MM-DD)', required=False),
        openapi.Parameter('to_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End date (YYYY-MM-DD)', required=False),
        openapi.Parameter('bp_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional BP code filter (line-level)', required=False),
        openapi.Parameter('account_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional account code filter (line-level)', required=False),
        openapi.Parameter('project_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional project code filter (line-level)', required=False),
        openapi.Parameter('max_entries', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Max JournalEntries to fetch from SAP (default 50, max 500). Use smaller values for better performance.', required=False),
        openapi.Parameter('skip_lines', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, description='If true, only fetch entry headers without lines (faster, default false)', required=False),
    ],
    responses={
        200: openapi.Response(description='Excel file download', schema=openapi.Schema(type=openapi.TYPE_FILE)),
        500: 'Error generating Journal Entry ledger export'
    }
)
@api_view(['GET'])
def export_journal_entry_ledger_sap_api(request):
    """
    GET /api/general-ledger/sap-journal-entry/export-excel/

    Separate downloadable Journal Entry ledger endpoint using SAP Service Layer.
    Optimized for SAP B1 Service Layer performance constraints.
    """
    if not OPENPYXL_AVAILABLE:
        return Response({
            'success': False,
            'error': 'Excel export not available. Please install openpyxl package.'
        }, status=500)

    try:
        company = (request.GET.get('company') or '').strip() or None
        from_date = (request.GET.get('from_date') or '').strip()
        to_date = (request.GET.get('to_date') or '').strip()
        bp_code = (request.GET.get('bp_code') or '').strip()
        account_code = (request.GET.get('account_code') or '').strip()
        project_code = (request.GET.get('project_code') or '').strip()
        skip_lines = (request.GET.get('skip_lines') or '').lower() in ('true', '1', 'yes', 'on')

        try:
            max_entries = int(request.GET.get('max_entries', 50) or 50)
        except Exception:
            max_entries = 50
        max_entries = max(1, min(max_entries, 500))

        logger.info(f"[JE API] Request: company={company}, from_date={from_date}, to_date={to_date}, bp_code={bp_code}, max_entries={max_entries}, skip_lines={skip_lines}")

        parsed_from = None
        parsed_to = None
        if from_date:
            try:
                parsed_from = datetime.strptime(from_date, '%Y-%m-%d').date()
            except Exception as e:
                logger.error(f"[JE API] Failed to parse from_date '{from_date}': {e}")
                raise ValueError(f"Invalid from_date format: {from_date}")
        if to_date:
            try:
                parsed_to = datetime.strptime(to_date, '%Y-%m-%d').date()
            except Exception as e:
                logger.error(f"[JE API] Failed to parse to_date '{to_date}': {e}")
                raise ValueError(f"Invalid to_date format: {to_date}")

        client = SAPClient(company_db_key=company)
        logger.info(f"[JE API] SAP Client initialized for company_db={client.company_db}, use_http={client.use_http}")

        entries = _fetch_entries_from_sap(client=client, max_entries=max_entries, skip_lines=skip_lines, page_size=10)

        logger.info(f"[JE API] Building Excel with {len(entries)} entries (before line filtering)")

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Journal Entry Ledger'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        headers = [
            'JE No', 'Posting Date', 'Due Date', 'Document Date',
            'Memo', 'Line No', 'Account', 'BP Code', 'Line Memo',
            'Project', 'Debit', 'Credit', 'Balance'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        running_balance = 0.0
        row_num = 2
        rows_written = 0
        flat_rows = _flatten_entries(entries, parsed_from, parsed_to, bp_code, account_code, project_code)

        for r in flat_rows:
            running_balance = _to_float(r.get('balance'))
            ws.cell(row=row_num, column=1, value=r.get('je_no', ''))
            ws.cell(row=row_num, column=2, value=r.get('posting_date', ''))
            ws.cell(row=row_num, column=3, value=r.get('due_date', ''))
            ws.cell(row=row_num, column=4, value=r.get('doc_date', ''))
            ws.cell(row=row_num, column=5, value=r.get('memo', ''))
            ws.cell(row=row_num, column=6, value=r.get('line_no', ''))
            ws.cell(row=row_num, column=7, value=r.get('account', ''))
            ws.cell(row=row_num, column=8, value=r.get('bp_code', ''))
            ws.cell(row=row_num, column=9, value=r.get('line_memo', ''))
            ws.cell(row=row_num, column=10, value=r.get('project', ''))
            ws.cell(row=row_num, column=11, value=r.get('debit', 0))
            ws.cell(row=row_num, column=12, value=r.get('credit', 0))
            ws.cell(row=row_num, column=13, value=running_balance)
            row_num += 1
            rows_written += 1

        logger.info(f"[JE API] Excel built: {rows_written} rows written")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = '' if cell.value is None else str(cell.value)
                    if len(val) > max_length:
                        max_length = len(val)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        company_part = company or 'default'

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="sap_journal_entry_ledger_{company_part}_{stamp}.xlsx"'
        )
        logger.info(f"[JE API] Response generated successfully: {len(output.getvalue())} bytes")
        return response

    except Exception as e:
        error_text = str(e)
        logger.error(f"[JE API] Endpoint error: {error_text}", exc_info=True)

        lowered = error_text.lower()
        if 'timed out' in lowered or 'timeout' in lowered:
            return Response({
                'success': False,
                'error': (
                    'SAP Service Layer timeout at fourb.vdc.services:5588. '
                    'Please retry later or contact SAP administrator to verify Service Layer health.'
                )
            }, status=503)

        if 'ssl' in lowered or 'tls' in lowered:
            return Response({
                'success': False,
                'error': (
                    'SAP Service Layer TLS/SSL connection error at fourb.vdc.services:5588. '
                    'Please verify Service Layer certificate/protocol settings.'
                )
            }, status=503)

        return Response({'success': False, 'error': error_text}, status=500)
