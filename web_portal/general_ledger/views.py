"""
General Ledger API Views and Admin Interface
"""
from rest_framework.response import Response
from rest_framework.decorators import api_view
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
import json
import logging
import csv
import re
from datetime import datetime
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.cell.cell import WriteOnlyCell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage, KeepTogether
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen.canvas import Canvas as _RLCanvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def _make_last_page_only_canvas(footer_fn):
    """
    Build a reportlab Canvas subclass that calls footer_fn(canvas) only on the
    final page. footer_fn receives just the canvas; it must not depend on the
    SimpleDocTemplate doc argument.
    """
    class _LastPageOnlyCanvas(_RLCanvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._saved_page_states)
            for idx, state in enumerate(self._saved_page_states):
                self.__dict__.update(state)
                if idx == total - 1:
                    try:
                        footer_fn(self)
                    except Exception:
                        pass
                _RLCanvas.showPage(self)
            _RLCanvas.save(self)

    return _LastPageOnlyCanvas

try:
    import arabic_reshaper
    from bidi.algorithm import get_display as bidi_display
    ARABIC_AVAILABLE = True
    # Urdu-tuned reshaper: enables Urdu-specific ligatures and presentation
    # forms. The default config is Arabic-centric and misses some Urdu
    # letterforms (e.g. ٹ ڈ ڑ ۂ).
    try:
        _urdu_reshaper = arabic_reshaper.ArabicReshaper(configuration={
            'language': 'Urdu',
            'support_ligatures': True,
            'delete_harakat': False,
        })
    except Exception:
        _urdu_reshaper = None
except ImportError:
    ARABIC_AVAILABLE = False
    _urdu_reshaper = None

# Register a Unicode/Urdu-capable font once at module load.
# Discovery order:
#   1. Project-bundled font in general_ledger/fonts/  (deployment-agnostic)
#   2. Hardcoded common system paths (Windows / Linux)
#   3. Recursive scan of /usr/share/fonts/ for NotoNastaliqUrdu*
#   4. fc-match query (Linux only)
_URDU_FONT = 'Helvetica'
if REPORTLAB_AVAILABLE:
    import os as _font_os
    import sys as _font_sys
    import glob as _font_glob
    import subprocess as _font_subproc

    # 1. Project-bundled font directory.
    #    Naskh-style fonts (NotoNaskhArabic, Amiri) ship precomposed glyphs for
    #    Arabic Presentation Forms (U+FE70-FEFF) — they render correctly with
    #    ReportLab. Nastaliq fonts (NotoNastaliqUrdu, JameelNooriNastaleeq)
    #    rely on OpenType GSUB shaping which ReportLab does NOT perform, so
    #    they render as blank/missing glyphs. Naskh is checked FIRST.
    _bundled_font_dir = _font_os.path.join(_font_os.path.dirname(__file__), 'fonts')
    _bundled_candidates = [
        ('NotoNaskhArabic',     _font_os.path.join(_bundled_font_dir, 'NotoNaskhArabic-Regular.ttf')),
        ('Amiri',               _font_os.path.join(_bundled_font_dir, 'Amiri-Regular.ttf')),
        ('JameelNooriNastaleeq',_font_os.path.join(_bundled_font_dir, 'Jameel Noori Nastaleeq.ttf')),
        ('NotoNastaliqUrdu',    _font_os.path.join(_bundled_font_dir, 'NotoNastaliqUrdu-Regular.ttf')),
    ]

    # 2. Common system paths — Naskh-style fonts first.
    if _font_sys.platform.startswith('win'):
        _system_candidates = [
            ('Tahoma',          r'C:\Windows\Fonts\tahoma.ttf'),
            ('Arial',           r'C:\Windows\Fonts\arial.ttf'),
            ('Calibri',         r'C:\Windows\Fonts\calibri.ttf'),
            ('TimesNewRoman',   r'C:\Windows\Fonts\times.ttf'),
            ('NotoNastaliqUrdu', r'C:\Windows\Fonts\NotoNastaliqUrdu-Regular.ttf'),
            ('JameelNooriNastaleeq', r'C:\Windows\Fonts\Jameel Noori Nastaleeq.ttf'),
        ]
    else:
        # Note: on modern Ubuntu the apt package `fonts-noto-nastaliq-urdu`
        # installs to /usr/share/fonts/opentype/noto/ (NOT truetype/), so we
        # check both locations.
        _system_candidates = [
            ('NotoNaskhArabic',  '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf'),
            ('NotoNaskhArabic',  '/usr/share/fonts/opentype/noto/NotoNaskhArabic-Regular.ttf'),
            ('Amiri',            '/usr/share/fonts/truetype/amiri/amiri-regular.ttf'),
            ('NotoSans',         '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf'),
            ('DejaVuSans',       '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
            ('LiberationSans',   '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'),
            ('FreeSans',         '/usr/share/fonts/truetype/freefont/FreeSans.ttf'),
            ('NotoNastaliqUrdu', '/usr/share/fonts/opentype/noto/NotoNastaliqUrdu-Regular.ttf'),
            ('NotoNastaliqUrdu', '/usr/share/fonts/truetype/noto/NotoNastaliqUrdu-Regular.ttf'),
        ]

    _URDU_FONT_CANDIDATES = _bundled_candidates + _system_candidates

    # 3. Recursive scan for Arabic-capable fonts under common roots — Naskh
    #    style preferred (it actually renders with ReportLab).
    if not _font_sys.platform.startswith('win'):
        try:
            for _name, _pat in (
                ('NotoNaskhArabic',  '/usr/share/fonts/**/NotoNaskhArabic*.ttf'),
                ('NotoNaskhArabic',  '/usr/share/fonts/**/NotoNaskhArabic*.otf'),
                ('Amiri',            '/usr/share/fonts/**/[Aa]miri*.ttf'),
                ('NotoNastaliqUrdu', '/usr/share/fonts/**/NotoNastaliqUrdu*.ttf'),
                ('NotoNastaliqUrdu', '/usr/share/fonts/**/NotoNastaliqUrdu*.otf'),
            ):
                for _match in _font_glob.glob(_pat, recursive=True):
                    _URDU_FONT_CANDIDATES.append((_name, _match))
        except Exception:
            pass

    # 4. fc-match fallback — ask fontconfig where Arabic fonts live (Naskh first)
    if not _font_sys.platform.startswith('win'):
        for _name, _query in (
            ('NotoNaskhArabic',  'Noto Naskh Arabic'),
            ('Amiri',            'Amiri'),
            ('NotoNastaliqUrdu', 'Noto Nastaliq Urdu'),
        ):
            try:
                _fc_out = _font_subproc.run(
                    ['fc-match', '-f', '%{file}', _query],
                    capture_output=True, text=True, timeout=5,
                )
                _fc_path = (_fc_out.stdout or '').strip()
                if _fc_path and _font_os.path.exists(_fc_path):
                    _URDU_FONT_CANDIDATES.append((_name, _fc_path))
            except Exception:
                pass

    try:
        from reportlab.pdfbase import pdfmetrics as _pm
        from reportlab.pdfbase.ttfonts import TTFont as _TTFont
        _registered = False
        _seen_paths = set()
        for _fname, _fpath in _URDU_FONT_CANDIDATES:
            if not _fpath or _fpath in _seen_paths:
                continue
            _seen_paths.add(_fpath)
            if _font_os.path.exists(_fpath):
                try:
                    if _fname not in _pm.getRegisteredFontNames():
                        _pm.registerFont(_TTFont(_fname, _fpath))
                    _URDU_FONT = _fname
                    print(f"[General Ledger] Successfully registered Urdu font: {_fname} from {_fpath}")
                    _registered = True
                    break
                except Exception as e:
                    print(f"[General Ledger] Failed to register font {_fname} at {_fpath}: {e}")
                    continue
        if not _registered:
            print(f"[General Ledger] No suitable Urdu font found. Falling back to {_URDU_FONT}")
            print(f"[General Ledger] Searched {len(_seen_paths)} paths. Drop a TTF into {_bundled_font_dir} for guaranteed support.")
    except Exception as e:
        print(f"[General Ledger] Error during font registration: {e}")

from . import hana_queries
from .utils import (
    get_hana_connection,
    get_company_options,
    calculate_running_balance,
    group_by_account,
    calculate_totals
)

logger = logging.getLogger("general_ledger")

# Custom Flowable for Urdu text that draws directly on canvas
# This is necessary because Paragraph class breaks reshaped RTL text
if REPORTLAB_AVAILABLE:
    from reportlab.platypus.flowables import Flowable
    
    class UrduTextBox(Flowable):
        """Custom flowable that draws Urdu text directly on canvas"""
        def __init__(self, text, font_name, font_size, width, text_color, bg_color, padding=8):
            Flowable.__init__(self)
            self.text = text
            self.font_name = font_name
            self.font_size = font_size
            self.width = width
            self.text_color = text_color
            self.bg_color = bg_color
            self.padding = padding
            # Calculate height with extra space for proper text rendering
            self.height = (font_size * 1.6) + (padding * 2)
        
        def draw(self):
            """Draw the Urdu text on the canvas"""
            canvas = self.canv
            
            print(f"[UrduTextBox] Drawing text: {self.text[:50]}...")
            print(f"[UrduTextBox] Font: {self.font_name}, Size: {self.font_size}")
            print(f"[UrduTextBox] Box dimensions: {self.width} x {self.height}")
            
            try:
                # Save graphics state
                canvas.saveState()
                
                # Draw background box
                canvas.setFillColor(self.bg_color)
                canvas.setStrokeColor(colors.HexColor('#6366f1'))
                canvas.setLineWidth(0.75)
                canvas.rect(0, 0, self.width, self.height, fill=1, stroke=1)
                print(f"[UrduTextBox] Background drawn")
                
                # Set text rendering mode to fill
                canvas.setFillColor(self.text_color)
                
                # Try to register font in canvas context if not already done
                try:
                    from reportlab.pdfbase import pdfmetrics
                    if self.font_name not in pdfmetrics.getRegisteredFontNames():
                        print(f"[UrduTextBox] WARNING: Font {self.font_name} not in registered fonts!")
                except Exception as e:
                    print(f"[UrduTextBox] Could not check font registration: {e}")
                
                canvas.setFont(self.font_name, self.font_size)
                
                # Position text right-aligned with proper baseline
                text_x = self.width - self.padding
                # Use simpler baseline calculation - bottom padding + small offset
                text_y = self.padding + 2
                
                print(f"[UrduTextBox] Drawing text at position ({text_x}, {text_y})")
                
                # Draw the text
                canvas.drawRightString(text_x, text_y, self.text)
                
                # Restore graphics state
                canvas.restoreState()
                
                print(f"[UrduTextBox] Text drawn successfully")
            except Exception as e:
                print(f"[UrduTextBox] ERROR in draw(): {e}")
                import traceback
                traceback.print_exc()



# ============================================================================
# REST API Views for Mobile/React App
# ============================================================================

@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Get General Ledger Report',
    operation_description='Fetch general ledger transactions. **NO REQUIRED FIELDS**. All parameters are optional filters for account range, date range, business partner, and project. Use user parameter to filter by user\'s assigned customers.',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
        openapi.Parameter('account_from', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Start of account range', required=False),
        openapi.Parameter('account_to', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: End of account range', required=False),
        openapi.Parameter('from_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Start date (YYYY-MM-DD)', required=False),
        openapi.Parameter('to_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: End date (YYYY-MM-DD)', required=False),
        openapi.Parameter('bp_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Business Partner CardCode. If user parameter is provided, validates that bp_code belongs to that user.', required=False),
        openapi.Parameter('user', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: User ID or username. Filters transactions to only user\'s assigned customers. Can be combined with bp_code for validation.', required=False),
        openapi.Parameter('project_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Project Code', required=False),
        openapi.Parameter('trans_type', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Transaction Type (e.g., 13, 30)', required=False),
        openapi.Parameter('page', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Optional: Page number', default=1, required=False),
        openapi.Parameter('page_size', openapi.IN_QUERY, type=openapi.TYPE_INTEGER, description='Optional: Records per page', default=50, required=False),
        openapi.Parameter('group_by_account', openapi.IN_QUERY, type=openapi.TYPE_BOOLEAN, description='Optional: Group results by account with opening/closing balances', default=False, required=False),
    ],
    responses={
        200: openapi.Response(
            description='Success',
            examples={
                'application/json': {
                    'success': True,
                    'data': {
                        'accounts': [
                            {
                                'Account': '11010100',
                                'AccountName': 'Cash in Hand',
                                'OpeningBalance': 1000.00,
                                'transactions': [],
                                'TotalDebit': 5000.00,
                                'TotalCredit': 3000.00,
                                'ClosingBalance': 3000.00
                            }
                        ],
                        'grand_total': {
                            'TotalDebit': 50000.00,
                            'TotalCredit': 50000.00,
                            'Difference': 0.00
                        }
                    },
                    'pagination': {
                        'page': 1,
                        'page_size': 50,
                        'total_records': 234,
                        'total_pages': 5
                    }
                }
            }
        ),
        500: 'Error fetching ledger data'
    }
)
@api_view(['GET'])
def general_ledger_api(request):
    """
    GET /api/general-ledger/
    
    Fetch general ledger transactions with optional filters.
    Supports grouping by account with opening/closing balances.
    Supports user-based filtering to show only user's assigned customers.
    """
    try:
        # Extract query parameters
        company = (request.GET.get('company') or '').strip()
        account_from = request.GET.get('account_from', '').strip()
        account_to = request.GET.get('account_to', '').strip()
        from_date = request.GET.get('from_date', '').strip()
        to_date = request.GET.get('to_date', '').strip()
        bp_code = request.GET.get('bp_code', '').strip()
        user_param = request.GET.get('user', '').strip()
        project_code = request.GET.get('project_code', '').strip()
        trans_type = request.GET.get('trans_type', '').strip()
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        group_by_account_flag = request.GET.get('group_by_account', 'false').lower() == 'true'
        
        # Handle user-based filtering
        if user_param:
            try:
                from accounts.models import User
                from FieldAdvisoryService.models import Dealer
                
                # Get user object
                user_obj = None
                try:
                    user_id = int(user_param)
                    user_obj = User.objects.get(id=user_id)
                except (ValueError, User.DoesNotExist):
                    user_obj = User.objects.filter(username=user_param).first()
                
                if not user_obj:
                    return Response({'success': False, 'error': f'User "{user_param}" not found'}, status=404)
                
                # Get dealer card codes for this user
                dealers = Dealer.objects.filter(user=user_obj).values_list('card_code', flat=True).exclude(card_code__isnull=True).exclude(card_code='')
                dealer_card_codes = list(dealers)
                
                if not dealer_card_codes:
                    return Response({'success': True, 'data': {'transactions': [], 'grand_total': {'TotalDebit': 0, 'TotalCredit': 0, 'Balance': 0}}, 'pagination': {'page': page, 'page_size': page_size, 'total_records': 0, 'total_pages': 0}})
                
                # If bp_code is provided, validate it belongs to user
                if bp_code:
                    if bp_code not in dealer_card_codes:
                        return Response({'success': False, 'error': f'Business Partner "{bp_code}" does not belong to user "{user_param}"'}, status=403)
                    # Use only the specified bp_code
                    bp_code = bp_code
                else:
                    # Use all dealer card codes as a list
                    bp_code = dealer_card_codes
                    
            except Exception as e:
                return Response({'success': False, 'error': f'Error processing user filter: {str(e)}'}, status=500)
        
        # Connect to HANA
        conn = get_hana_connection(company)
        
        # Get total count for pagination
        total_count = hana_queries.general_ledger_count(
            conn,
            account_from=account_from or None,
            account_to=account_to or None,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code or None,
            project_code=project_code or None,
            trans_type=trans_type or None
        )
        
        # Calculate offset
        offset = (page - 1) * page_size
        
        # Fetch transactions
        transactions = hana_queries.general_ledger_report(
            conn,
            account_from=account_from or None,
            account_to=account_to or None,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code or None,
            project_code=project_code or None,
            trans_type=trans_type or None,
            limit=page_size,
            offset=offset
        )
        
        conn.close()
        
        # Prepare response
        if group_by_account_flag:
            # Group by account and calculate opening/closing balances
            grouped = group_by_account(transactions)
            accounts_list = []
            grand_total = {'TotalDebit': 0, 'TotalCredit': 0}
            
            for account_code, account_data in grouped.items():
                # Calculate opening balance if date range provided
                opening_balance = 0
                if from_date:
                    conn = get_hana_connection(company)
                    opening = hana_queries.account_opening_balance(
                        conn,
                        account_code,
                        from_date,
                        bp_code=bp_code or None
                    )
                    conn.close()
                    opening_balance = float(opening.get('Balance', 0))
                
                # Calculate running balance
                account_data['transactions'] = calculate_running_balance(
                    account_data['transactions']
                )
                
                # Calculate totals
                totals = calculate_totals(account_data['transactions'])
                
                accounts_list.append({
                    'Account': account_data['Account'],
                    'AccountName': account_data['AccountName'],
                    'OpeningBalance': opening_balance,
                    'transactions': account_data['transactions'],
                    'TotalDebit': totals['TotalDebit'],
                    'TotalCredit': totals['TotalCredit'],
                    'ClosingBalance': opening_balance + totals['Balance']
                })
                
                grand_total['TotalDebit'] += totals['TotalDebit']
                grand_total['TotalCredit'] += totals['TotalCredit']
            
            grand_total['Difference'] = grand_total['TotalDebit'] - grand_total['TotalCredit']
            
            response_data = {
                'accounts': accounts_list,
                'grand_total': grand_total
            }
        else:
            # Return flat list with running balance
            transactions = calculate_running_balance(transactions)
            grand_total = calculate_totals(transactions)
            
            response_data = {
                'transactions': transactions,
                'grand_total': grand_total
            }
        
        return Response({
            'success': True,
            'data': response_data,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_records': total_count,
                'total_pages': (total_count + page_size - 1) // page_size
            }
        }, status=200)
        
    except Exception as e:
        logger.exception("Error in general_ledger_api")
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Get Chart of Accounts',
    operation_description='Fetch chart of accounts for dropdown/filter',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
        openapi.Parameter('account_type', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Account type filter (A=Assets, L=Liabilities, E=Equity, I=Income, O=Expense)'),
    ],
    responses={
        200: 'Success',
        500: 'Error fetching chart of accounts'
    }
)
@api_view(['GET'])
def chart_of_accounts_api(request):
    """
    GET /api/chart-of-accounts/
    
    Fetch chart of accounts for dropdowns and filters.
    """
    try:
        company = (request.GET.get('company') or '').strip()
        account_type = request.GET.get('account_type', '').strip()
        
        conn = get_hana_connection(company)
        accounts = hana_queries.chart_of_accounts_list(
            conn,
            account_type=account_type or None
        )
        conn.close()
        
        return Response({
            'success': True,
            'data': accounts
        }, status=200)
        
    except Exception as e:
        logger.exception("Error in chart_of_accounts_api")
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Get Transaction Types',
    operation_description='Get list of SAP B1 transaction types for filter dropdown',
    responses={200: 'Success', 500: 'Error'}
)
@api_view(['GET'])
def transaction_types_api(request):
    """
    GET /api/transaction-types/
    
    Get list of transaction types for dropdown filter.
    """
    try:
        conn = get_hana_connection()
        trans_types = hana_queries.transaction_types_lov(conn)
        conn.close()
        
        return Response({
            'success': True,
            'data': trans_types
        }, status=200)
        
    except Exception as e:
        logger.exception("Error in transaction_types_api")
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Get Business Partners LOV',
    operation_description='Get business partners for dropdown filter',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
        openapi.Parameter('bp_type', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='BP type (C=Customer, S=Supplier)'),
    ],
    responses={200: 'Success', 500: 'Error'}
)
@api_view(['GET'])
def business_partners_api(request):
    """
    GET /api/business-partners/
    
    Get business partners for dropdown filter.
    """
    try:
        company = (request.GET.get('company') or '').strip()
        bp_type = request.GET.get('bp_type', '').strip()
        
        conn = get_hana_connection(company)
        partners = hana_queries.business_partner_lov(
            conn,
            bp_type=bp_type or None
        )
        conn.close()
        
        return Response({
            'success': True,
            'data': partners
        }, status=200)
        
    except Exception as e:
        logger.exception("Error in business_partners_api")
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Get Projects LOV',
    operation_description='Get projects for dropdown filter',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
    ],
    responses={200: 'Success', 500: 'Error'}
)
@api_view(['GET'])
def projects_api(request):
    """
    GET /api/projects/
    
    Get projects for dropdown filter.
    """
    try:
        company = (request.GET.get('company') or '').strip()
        
        conn = get_hana_connection(company)
        projects = hana_queries.projects_lov(conn)
        conn.close()
        
        return Response({
            'success': True,
            'data': projects
        }, status=200)
        
    except Exception as e:
        logger.exception("Error in projects_api")
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Export General Ledger to Excel',
    operation_description='Download general ledger report as Excel file with filters applied. Returns formatted XLSX file with headers and styling. Use user parameter to filter by user\'s assigned customers.',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
        openapi.Parameter('account_from', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start of account range'),
        openapi.Parameter('account_to', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End of account range'),
        openapi.Parameter('from_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start date (YYYY-MM-DD)'),
        openapi.Parameter('to_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End date (YYYY-MM-DD)'),
        openapi.Parameter('bp_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Business Partner CardCode. If user parameter is provided, validates that bp_code belongs to that user.'),
        openapi.Parameter('user', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: User ID or username. Filters transactions to only user\'s assigned customers. Can be combined with bp_code for validation.'),
        openapi.Parameter('project_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Project Code'),
        openapi.Parameter('trans_type', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Transaction Type (e.g., 13, 30)'),
    ],
    responses={
        200: openapi.Response(
            description='Excel file download',
            schema=openapi.Schema(type=openapi.TYPE_FILE)
        ),
        500: 'Error generating Excel file'
    }
)
@api_view(['GET'])
def export_ledger_excel_api(request):
    """
    GET /api/general-ledger/export-excel/
    
    Export general ledger to Excel format for mobile app.
    Returns a downloadable XLSX file with formatted headers and data.
    """
    if not OPENPYXL_AVAILABLE:
        return Response({
            'success': False,
            'error': 'Excel export not available. Please install openpyxl package.'
        }, status=500)
    
    def sanitize_for_excel(value):
        """Remove control characters that Excel can't handle."""
        if value is None:
            return ''
        if isinstance(value, (int, float)):
            return value
        # Convert to string and remove control characters (ASCII 0-31 except tab, newline, carriage return)
        text = str(value)
        # Remove control characters except \t (9), \n (10), \r (13)
        text = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', text)
        return text
    
    try:
        # Get filter parameters
        company = (request.GET.get('company') or '').strip()
        account_from = (request.GET.get('account_from') or '').strip()
        account_to = (request.GET.get('account_to') or '').strip()
        from_date = (request.GET.get('from_date') or '').strip()
        to_date = (request.GET.get('to_date') or '').strip()
        bp_code = (request.GET.get('bp_code') or '').strip()
        user_param = request.GET.get('user', '').strip()
        project_code = (request.GET.get('project_code') or '').strip()
        trans_type = (request.GET.get('trans_type') or '').strip()
        
        # Handle user-based filtering
        if user_param:
            try:
                from accounts.models import User
                from FieldAdvisoryService.models import Dealer
                
                # Get user object
                user_obj = None
                try:
                    user_id = int(user_param)
                    user_obj = User.objects.get(id=user_id)
                except (ValueError, User.DoesNotExist):
                    user_obj = User.objects.filter(username=user_param).first()
                
                if not user_obj:
                    return Response({'success': False, 'error': f'User "{user_param}" not found'}, status=404)
                
                # Get dealer card codes for this user
                dealers = Dealer.objects.filter(user=user_obj).values_list('card_code', flat=True).exclude(card_code__isnull=True).exclude(card_code='')
                dealer_card_codes = list(dealers)
                
                if not dealer_card_codes:
                    # Return empty list for bp_code - will result in empty export file
                    bp_code = []
                elif bp_code:
                    # If bp_code is provided, validate it belongs to user
                    if bp_code not in dealer_card_codes:
                        return Response({'success': False, 'error': f'Business Partner "{bp_code}" does not belong to user "{user_param}"'}, status=403)
                    # Use only the specified bp_code
                    bp_code = bp_code
                else:
                    # Use all dealer card codes as a list
                    bp_code = dealer_card_codes
                    
            except Exception as e:
                return Response({'success': False, 'error': f'Error processing user filter: {str(e)}'}, status=500)
        
        # Require date range to prevent full-table HANA scans
        if not from_date or not to_date:
            return Response({
                'success': False,
                'error': 'from_date and to_date are required for export.'
            }, status=400)

        # Fetch data with hard row cap to prevent OOM
        ROW_LIMIT = 50000
        conn = get_hana_connection(company)
        transactions = hana_queries.general_ledger_report(
            conn,
            account_from=account_from or None,
            account_to=account_to or None,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code if bp_code else None,
            project_code=project_code or None,
            trans_type=trans_type or None,
            limit=ROW_LIMIT,
        )
        conn.close()
        truncated = len(transactions) >= ROW_LIMIT

        # Build Excel using write-only (streaming) mode — keeps only one row in RAM at a time
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="General Ledger")

        # Header row with styling via WriteOnlyCell
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        headers = [
            'VNo.', 'VDate', 'Product Name', 'Qty', 'Price', 'Policy',
            'Discount', 'Sale Value', 'Type', 'Debit', 'Credit', 'Balance'
        ]
        header_cells = []
        for header in headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            header_cells.append(cell)
        ws.append(header_cells)

        # Data rows — written one at a time, never all in RAM together
        running_balance = 0
        for txn in transactions:
            debit = float(txn.get('Debit', 0))
            credit = float(txn.get('Credit', 0))
            running_balance += (debit - credit)

            posting_date = txn.get('PostingDate', '')
            try:
                date_obj = datetime.strptime(str(posting_date)[:10], '%Y-%m-%d')
                posting_date = date_obj.strftime('%-m/%-d/%Y').lstrip('0').replace('/-', '/').replace('/0', '/')
            except Exception:
                posting_date = str(posting_date)[:10]

            ws.append([
                sanitize_for_excel(txn.get('TransId', '')),
                posting_date,
                sanitize_for_excel(txn.get('Description', '')),
                txn.get('Qty', 0),
                txn.get('UnitPrice', 0),
                sanitize_for_excel(txn.get('ExtractedProject') or txn.get('ProjectCode', '')),
                txn.get('Discount', 0),
                txn.get('Amount', 0),
                sanitize_for_excel(txn.get('TransTypeName') or txn.get('TransType', '')),
                debit if debit > 0 else 0,
                credit if credit > 0 else 0,
                running_balance,
            ])

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="general_ledger_{company}_{timestamp}.xlsx"'
        if truncated:
            response['X-Truncated'] = 'true'
            response['X-Row-Limit'] = str(ROW_LIMIT)

        return response

    except Exception as e:
        logger.exception("Error exporting Excel")
        return Response({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# Admin Web Interface Views
# ============================================================================

@staff_member_required
def ledger_settings_redirect(request):
    """
    Ensure the singleton LedgerSettings row exists, then redirect to its change page.
    """
    from django.shortcuts import redirect
    from .models import LedgerSettings
    obj = LedgerSettings.get()   # creates row with defaults if not yet saved
    return redirect(f'/admin/general_ledger/ledgersettings/{obj.pk}/change/')


@staff_member_required
def general_ledger_admin(request):
    """
    Admin web interface for General Ledger Report.
    Displays ledger with company selector, filters, grouping, and export.
    """
    error = None
    warning = None
    result_rows = []
    accounts_grouped = []
    grand_total = None
    
    # Get company database options
    db_options = get_company_options()
    selected_db_key = (request.GET.get('company') or '').strip()
    
    # Get filter parameters
    account_from = (request.GET.get('account_from') or '').strip()
    account_to = (request.GET.get('account_to') or '').strip()
    from_date = (request.GET.get('from_date') or '').strip()
    to_date = (request.GET.get('to_date') or '').strip()
    bp_code = (request.GET.get('bp_code') or '').strip()
    project_code = (request.GET.get('project_code') or '').strip()
    trans_type = (request.GET.get('trans_type') or '').strip()
    
    # Pagination parameters
    page_num = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 50))
    
    # Load dropdown data
    chart_of_accounts = []
    business_partners = []
    projects = []
    trans_types = []
    
    try:
        conn = get_hana_connection(selected_db_key)
        chart_of_accounts = hana_queries.chart_of_accounts_list(conn)
        business_partners = hana_queries.business_partner_lov(conn, limit=500)
        projects = hana_queries.projects_lov(conn)
        trans_types = hana_queries.transaction_types_lov(conn)
        conn.close()
    except Exception as e:
        logger.warning(f"SAP connection unavailable for dropdown data: {str(e)}")
        # Continue with empty dropdowns - user can still use manual input
        warning = "⚠️ SAP server is currently unavailable. Dropdown options are disabled, but you can still enter filter values manually."
    
    # Fetch ledger data if filters applied
    if account_from or account_to or from_date or to_date or bp_code or project_code or trans_type:
        try:
            conn = get_hana_connection(selected_db_key)
            
            # Get transactions with a cap to prevent OOM on admin view
            result_rows = hana_queries.general_ledger_report(
                conn,
                account_from=account_from or None,
                account_to=account_to or None,
                from_date=from_date or None,
                to_date=to_date or None,
                bp_code=bp_code or None,
                project_code=project_code or None,
                trans_type=trans_type or None,
                limit=5000,
            )
            
            # Group by account
            grouped = group_by_account(result_rows)
            
            # Calculate opening/closing balances for each account
            for account_code, account_data in grouped.items():
                opening_balance = 0
                if from_date:
                    opening = hana_queries.account_opening_balance(
                        conn,
                        account_code,
                        from_date,
                        bp_code=bp_code or None
                    )
                    opening_balance = float(opening.get('Balance', 0))
                
                # Calculate running balance
                account_data['transactions'] = calculate_running_balance(
                    account_data['transactions']
                )
                
                # Calculate totals
                totals = calculate_totals(account_data['transactions'])
                
                accounts_grouped.append({
                    'Account': account_data['Account'],
                    'AccountName': account_data['AccountName'],
                    'OpeningBalance': opening_balance,
                    'transactions': account_data['transactions'],
                    'TotalDebit': totals['TotalDebit'],
                    'TotalCredit': totals['TotalCredit'],
                    'ClosingBalance': opening_balance + totals['Balance']
                })
            
            conn.close()
            
            # Calculate grand total
            if accounts_grouped:
                grand_total = {
                    'TotalDebit': sum(a['TotalDebit'] for a in accounts_grouped),
                    'TotalCredit': sum(a['TotalCredit'] for a in accounts_grouped),
                }
                grand_total['Difference'] = grand_total['TotalDebit'] - grand_total['TotalCredit']
            
        except Exception as e:
            logger.exception("Error fetching ledger data")
            error = str(e)
    
    # Pagination on accounts level
    paginator = Paginator(accounts_grouped, page_size)
    page_obj = paginator.get_page(page_num)
    
    context = {
        'db_options': db_options,
        'selected_db_key': selected_db_key,
        'chart_of_accounts': chart_of_accounts,
        'business_partners': business_partners,
        'projects': projects,
        'trans_types': trans_types,
        'accounts_grouped': list(page_obj.object_list),
        'grand_total': grand_total,
        'error': error,
        'warning': warning,
        'pagination': {
            'page': page_obj.number,
            'num_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_prev': page_obj.has_previous(),
            'next_page': (page_obj.next_page_number() if page_obj.has_next() else None),
            'prev_page': (page_obj.previous_page_number() if page_obj.has_previous() else None),
            'count': paginator.count,
            'page_size': page_size,
        },
        'filters': {
            'account_from': account_from,
            'account_to': account_to,
            'from_date': from_date,
            'to_date': to_date,
            'bp_code': bp_code,
            'project_code': project_code,
            'trans_type': trans_type,
        }
    }
    
    return render(request, 'general_ledger/general_ledger.html', context)


@staff_member_required
def export_ledger_csv(request):
    """
    Export general ledger to CSV format.
    """
    try:
        # Get filter parameters
        company = (request.GET.get('company') or '').strip()
        account_from = (request.GET.get('account_from') or '').strip()
        account_to = (request.GET.get('account_to') or '').strip()
        from_date = (request.GET.get('from_date') or '').strip()
        to_date = (request.GET.get('to_date') or '').strip()
        bp_code = (request.GET.get('bp_code') or '').strip()
        project_code = (request.GET.get('project_code') or '').strip()
        trans_type = (request.GET.get('trans_type') or '').strip()

        if not from_date or not to_date:
            return HttpResponse('from_date and to_date are required for export.', status=400)

        conn = get_hana_connection(company)
        transactions = hana_queries.general_ledger_report(
            conn,
            account_from=account_from or None,
            account_to=account_to or None,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code or None,
            project_code=project_code or None,
            trans_type=trans_type or None,
            limit=50000,
        )
        conn.close()

        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="general_ledger_{company}_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Posting Date', 'Trans ID', 'Account', 'Account Name', 'BP Code', 'BP Name',
            'Trans Type', 'Reference', 'Description', 'Debit', 'Credit', 'Project'
        ])
        
        # Write data rows
        for txn in transactions:
            writer.writerow([
                txn.get('PostingDate', ''),
                txn.get('TransId', ''),
                txn.get('Account', ''),
                txn.get('AccountName', ''),
                txn.get('BPCode', ''),
                txn.get('BPName', ''),
                txn.get('TransTypeName') or txn.get('TransType', ''),
                txn.get('Reference1', ''),
                txn.get('Description', ''),
                txn.get('Debit', 0),
                txn.get('Credit', 0),
                txn.get('ExtractedProject') or txn.get('ProjectCode', ''),
            ])
        
        return response
        
    except Exception as e:
        logger.exception("Error exporting CSV")
        return HttpResponse(f"Error exporting CSV: {str(e)}", status=500)

@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Export General Ledger to PDF',
    operation_description='Download general ledger report as PDF file with filters applied. Returns formatted PDF file with headers and styling. Use user parameter to filter by user\'s assigned customers.',
    manual_parameters=[
        openapi.Parameter('company', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
        openapi.Parameter('account_from', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start of account range'),
        openapi.Parameter('account_to', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End of account range'),
        openapi.Parameter('from_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start date (YYYY-MM-DD)'),
        openapi.Parameter('to_date', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End date (YYYY-MM-DD)'),
        openapi.Parameter('bp_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Business Partner CardCode. If user parameter is provided, validates that bp_code belongs to that user.'),
        openapi.Parameter('user', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: User ID or username. Filters transactions to only user\'s assigned customers. Can be combined with bp_code for validation.'),
        openapi.Parameter('project_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Project Code'),
        openapi.Parameter('trans_type', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Transaction Type (e.g., 13, 30)'),
    ],
    responses={
        200: openapi.Response(
            description='PDF file download',
            schema=openapi.Schema(type=openapi.TYPE_FILE)
        ),
        500: 'Error generating PDF file'
    }
)
@api_view(['GET'])
def export_ledger_pdf_api(request):
    """
    GET /api/general-ledger/export-pdf/

    Export general ledger to PDF format for mobile app.
    Returns a downloadable PDF file with formatted headers and data.
    """
    if not REPORTLAB_AVAILABLE:
        return Response({
            'success': False,
            'error': 'PDF export not available. Please install reportlab package.'
        }, status=500)

    try:
        # ── Query parameters ──────────────────────────────────────────────────
        company      = (request.GET.get('company')      or '').strip()
        account_from = (request.GET.get('account_from') or '').strip()
        account_to   = (request.GET.get('account_to')   or '').strip()
        from_date    = (request.GET.get('from_date')    or '').strip()
        to_date      = (request.GET.get('to_date')      or '').strip()
        bp_code      = (request.GET.get('bp_code')      or '').strip()
        user_param   = (request.GET.get('user')         or '').strip()
        project_code = (request.GET.get('project_code') or '').strip()
        trans_type   = (request.GET.get('trans_type')   or '').strip()

        # ── User-based filtering (mobile/dealer) ──────────────────────────────
        if user_param:
            try:
                from accounts.models import User
                from FieldAdvisoryService.models import Dealer

                user_obj = None
                try:
                    user_obj = User.objects.get(id=int(user_param))
                except (ValueError, User.DoesNotExist):
                    user_obj = User.objects.filter(username=user_param).first()

                if not user_obj:
                    return Response({'success': False, 'error': f'User "{user_param}" not found'}, status=404)

                dealer_card_codes = list(
                    Dealer.objects.filter(user=user_obj)
                    .exclude(card_code__isnull=True).exclude(card_code='')
                    .values_list('card_code', flat=True)
                )

                if not dealer_card_codes:
                    bp_code = []
                elif bp_code:
                    if bp_code not in dealer_card_codes:
                        return Response({
                            'success': False,
                            'error': f'Business Partner "{bp_code}" does not belong to user "{user_param}"'
                        }, status=403)
                else:
                    bp_code = dealer_card_codes

            except Exception as e:
                return Response({'success': False, 'error': f'Error processing user filter: {str(e)}'}, status=500)

        # ── Fetch & group data ────────────────────────────────────────────────
        if not from_date or not to_date:
            return Response({
                'success': False,
                'error': 'from_date and to_date are required for export.'
            }, status=400)

        conn = get_hana_connection(company)
        transactions = hana_queries.general_ledger_report(
            conn,
            account_from=account_from or None,
            account_to=account_to or None,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code if bp_code else None,
            project_code=project_code or None,
            trans_type=trans_type or None,
            limit=50000,
        )

        grouped_raw = group_by_account(transactions)
        accounts_grouped = []
        for account_code, account_data in grouped_raw.items():
            opening_balance = 0.0
            if from_date:
                opening = hana_queries.account_opening_balance(
                    conn, account_code, from_date, bp_code=bp_code or None
                )
                opening_balance = float(opening.get('Balance', 0))
            account_data['transactions'] = calculate_running_balance(account_data['transactions'])
            totals = calculate_totals(account_data['transactions'])
            accounts_grouped.append({
                'Account':        account_data['Account'],
                'AccountName':    account_data['AccountName'],
                'OpeningBalance': opening_balance,
                'transactions':   account_data['transactions'],
                'TotalDebit':     totals['TotalDebit'],
                'TotalCredit':    totals['TotalCredit'],
                'ClosingBalance': opening_balance + totals['Balance'],
            })
        conn.close()

        grand_total = None
        if accounts_grouped:
            grand_total = {
                'TotalDebit':  sum(a['TotalDebit']  for a in accounts_grouped),
                'TotalCredit': sum(a['TotalCredit'] for a in accounts_grouped),
            }
            grand_total['Difference'] = grand_total['TotalDebit'] - grand_total['TotalCredit']

        # ── PDF setup ─────────────────────────────────────────────────────────
        response = HttpResponse(content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = (
            f'attachment; filename="general_ledger_{company}_{timestamp}.pdf"'
        )

        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            topMargin=0.4 * inch, bottomMargin=0.4 * inch,
            leftMargin=0.3 * inch, rightMargin=0.3 * inch,
        )
        elements = []
        styles = getSampleStyleSheet()

        # cell_style defined after _cfg is loaded (below)

        # ── Date range strings ────────────────────────────────────────────────
        from_disp = from_date or '01/01/2000'
        to_disp   = to_date   or datetime.now().strftime('%m/%d/%Y')
        try:
            if from_date and len(from_date) == 10:
                d = datetime.strptime(from_date, '%Y-%m-%d')
                from_disp = f"{d.month}/{d.day}/{d.year}"
        except Exception:
            pass
        try:
            if to_date and len(to_date) == 10:
                d = datetime.strptime(to_date, '%Y-%m-%d')
                to_disp = f"{d.month}/{d.day}/{d.year}"
        except Exception:
            pass
        now = datetime.now()
        print_date_str = f"{now.month}/{now.day}/{now.year}   {now.strftime('%I:%M:%S%p')}"

        # ── Resolve company full name ─────────────────────────────────────────
        _company_full_name = company
        try:
            from FieldAdvisoryService.models import Company as _CompanyModel
            _co = _CompanyModel.objects.filter(name=company).first() \
                  or _CompanyModel.objects.filter(Company_name=company).first()
            if _co:
                _company_full_name = getattr(_co, 'Company_name', None) or company
        except Exception:
            pass

        # ── Load admin-managed settings ────────────────────────────────────
        from .models import LedgerSettings as _LedgerSettings
        _cfg = _LedgerSettings.get()

        _fs_grp  = int(_cfg.font_size_group_name   or 11)
        _fs_ent  = int(_cfg.font_size_company_name  or 13)
        _fs_ttl  = int(_cfg.font_size_report_title  or 10)
        _fs_dt   = int(_cfg.font_size_dates         or 9)
        _fs_td   = int(_cfg.font_size_table_data    or 8)

        cell_style = ParagraphStyle(
            'PDFCell', parent=styles['Normal'],
            fontSize=_fs_td, fontName='Helvetica', leading=_fs_td + 2, alignment=TA_LEFT,
        )

        # ── Header styles ─────────────────────────────────────────────────────
        _h_group  = ParagraphStyle('HDRGroup',  parent=styles['Normal'],
            fontSize=_fs_grp, fontName='Helvetica-Bold', leading=_fs_grp + 3)
        _h_entity = ParagraphStyle('HDREntity', parent=styles['Normal'],
            fontSize=_fs_ent, fontName='Helvetica-Bold', leading=_fs_ent + 4)
        _h_title  = ParagraphStyle('HDRTitle',  parent=styles['Normal'],
            fontSize=_fs_ttl, fontName='Helvetica',      leading=_fs_ttl + 3)
        _h_dates  = ParagraphStyle('HDRDates',  parent=styles['Normal'],
            fontSize=_fs_dt,  fontName='Helvetica',      leading=_fs_dt + 3)
        _h_print  = ParagraphStyle('HDRPrint',  parent=styles['Normal'],
            fontSize=8,  fontName='Helvetica',      leading=11, alignment=TA_RIGHT)

        # ── Logo (from admin-uploaded image or legacy path) ────────────────────
        import os as _os
        if _cfg.smart_stamp_image:
            _logo_path = _cfg.smart_stamp_image.path
        else:
            _logo_path = _os.path.join(
                _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                'media', 'Ledger_template', 'smart_stamp.png'
            )

        # ── 2-column header table ─────────────────────────────────────────────
        _page_w = landscape(letter)[0] - 0.6 * inch
        _right_col_w = _page_w * 0.30

        # SAP logo: use uploaded image if available, else draw programmatically
        _sap_logo2 = None
        if _cfg.sap_logo_image:
            try:
                _sap_logo2 = RLImage(_cfg.sap_logo_image.path, width=0.65 * inch, height=0.33 * inch)
            except Exception:
                _sap_logo2 = None
        if _sap_logo2 is None:
            try:
                from reportlab.graphics.shapes import Drawing as _RLDrawing, Rect as _RLRect, String as _RLString
                _sap_w2, _sap_h2 = 0.55 * inch, 0.28 * inch
                _sap_d2 = _RLDrawing(_sap_w2, _sap_h2)
                _sap_d2.add(_RLRect(0, 0, _sap_w2, _sap_h2,
                                    fillColor=colors.HexColor('#0070f3'),
                                    strokeColor=colors.HexColor('#0070f3')))
                _sap_d2.add(_RLString(_sap_w2 / 2, (_sap_h2 / 2) - 4, 'SAP',
                                      fontSize=12, fontName='Helvetica-Bold',
                                      fillColor=colors.white, textAnchor='middle'))
                _sap_logo2 = _sap_d2
            except Exception:
                _sap_logo2 = None

        _pdate_para2 = Paragraph(f'Print Date:   {print_date_str}', _h_print)
        _sap_cell2 = _sap_logo2 if _sap_logo2 is not None else Paragraph('SAP', _h_print)
        _right_top2 = Table(
            [[_pdate_para2, _sap_cell2]],
            colWidths=[_right_col_w - 0.65 * inch, 0.65 * inch],
        )
        _right_top2.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN',         (0, 0), (0, 0),   'RIGHT'),
            ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ]))

        _left_content = [
            Paragraph(_cfg.group_name or 'Four Brothers Group', _h_group),
            Paragraph((_cfg.company_name or _company_full_name or '').upper(), _h_entity),
            Paragraph(_cfg.report_title or 'General Ledger', _h_title),
            Spacer(1, 2),
            Paragraph(f'From:  {from_disp}     To:  {to_disp}', _h_dates),
        ]

        _right_content = [_right_top2]
        if _os.path.exists(_logo_path):
            _stamp = RLImage(_logo_path, width=1.1 * inch, height=1.1 * inch)
            _stamp.hAlign = 'RIGHT'
            _stamp.preserveAspectRatio = True
            _right_content.append(Spacer(1, 4))
            _right_content.append(_stamp)

        _hdr_table = Table(
            [[_left_content, _right_content]],
            colWidths=[_page_w * 0.70, _right_col_w],
        )
        _hdr_table.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ]))
        elements.append(_hdr_table)

        elements.append(Spacer(1, 4))
        _divider = Table([['']], colWidths=[_page_w])
        _divider.setStyle(TableStyle([
            ('LINEBELOW',     (0, 0), (-1, -1), 0.75, colors.HexColor('#1e293b')),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(_divider)
        elements.append(Spacer(1, 4))

        # ── Build territory lookup from Dealer model ──────────────────────────
        _bp_territory_map = {}
        try:
            from FieldAdvisoryService.models import Dealer as _Dealer
            for _d in _Dealer.objects.filter(card_code__isnull=False).select_related('region', 'zone', 'territory'):
                _parts = [p for p in [
                    getattr(_d.region,    'name', None),
                    getattr(_d.zone,      'name', None),
                    getattr(_d.territory, 'name', None),
                ] if p]
                _bp_territory_map[_d.card_code] = ' / '.join(_parts)
        except Exception:
            pass

        # ── Group flat transactions by BP ─────────────────────────────────────
        _bp_groups = {}
        _all_txns = [t for acct in accounts_grouped for t in acct['transactions']]
        for _txn in _all_txns:
            _bpc = str(_txn.get('BPCode', '') or '')
            if _bpc not in _bp_groups:
                _bp_groups[_bpc] = {
                    'BPCode': _bpc,
                    'BPName': str(_txn.get('BPName', '') or ''),
                    'Territory': _bp_territory_map.get(_bpc, ''),
                    'transactions': [],
                }
            _bp_groups[_bpc]['transactions'].append(_txn)

        # Recalculate running balance per BP
        for _bpg in _bp_groups.values():
            _bal = 0.0
            for _t in _bpg['transactions']:
                _bal += float(_t.get('Debit', 0) or 0) - float(_t.get('Credit', 0) or 0)
                _t['RunningBalance'] = _bal

        # ── Column widths (landscape letter ≈ 10.4" usable) ──────────────────
        col_widths = [
            0.65 * inch,   # VNo.
            0.80 * inch,   # VDate
            0.68 * inch,   # Type
            1.70 * inch,   # Policy Name (Account Name)
            2.50 * inch,   # Narration (Description)
            0.90 * inch,   # PR NO. (Reference)
            0.90 * inch,   # Debit
            0.90 * inch,   # Credit
            0.90 * inch,   # Balance
        ]
        NCOLS = len(col_widths)

        C_HEADER_BG  = colors.HexColor(_cfg.table_header_bg_color   or '#1e293b')
        C_HEADER_TXT = colors.HexColor(_cfg.table_header_text_color  or '#ffffff')
        C_TERR_BG    = colors.HexColor(_cfg.territory_row_bg_color   or '#f1f5f9')
        C_CLOSE_BG   = colors.HexColor(_cfg.closing_balance_bg_color or '#f1f5f9')
        C_GRAND_BG2  = colors.HexColor(_cfg.grand_total_bg_color     or '#e2e8f0')
        C_GRID       = colors.HexColor(_cfg.grid_color               or '#d1d5db')

        _fs_th = int(_cfg.font_size_table_header or 8)

        table_data   = []
        table_styles = [
            ('FONTSIZE',      (0, 0), (-1, -1), _fs_td),
            ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 3),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            # Horizontal lines only - no vertical column borders
            ('LINEBELOW',     (0, 0), (-1, -1), 0.25, C_GRID),
            ('LINEBEFORE',    (0, 0), (0, -1),  0,    colors.white),
            ('LINEAFTER',     (-1, 0), (-1, -1), 0,   colors.white),
            ('ALIGN',         (6, 0), (8, -1), 'RIGHT'),
        ]

        headers = ['VNo.', 'VDate', 'Type', 'Policy Name', 'Narration', 'PR NO.', 'Debit', 'Credit', 'Balance']
        table_data.append(headers)
        ri = 0
        table_styles += [
            ('BACKGROUND',    (0, ri), (-1, ri), C_HEADER_BG),
            ('TEXTCOLOR',     (0, ri), (-1, ri), C_HEADER_TXT),
            ('FONTNAME',      (0, ri), (-1, ri), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, ri), (-1, ri), _fs_th),
            ('ALIGN',         (0, ri), (-1, ri), 'CENTER'),
            ('BOTTOMPADDING', (0, ri), (-1, ri), 5),
            ('TOPPADDING',    (0, ri), (-1, ri), 5),
            # Top and bottom border around header row
            ('LINEABOVE',     (0, ri), (-1, ri), 0.75, C_GRID),
            ('LINEBELOW',     (0, ri), (-1, ri), 0.75, C_GRID),
        ]
        ri += 1

        # Deduplicate transactions per BP by TransId
        for _bpg in _bp_groups.values():
            _seen_ids = set()
            _deduped  = []
            for _t in _bpg['transactions']:
                _tid = str(_t.get('TransId', '') or '')
                if _tid and _tid in _seen_ids:
                    continue
                if _tid:
                    _seen_ids.add(_tid)
                _deduped.append(_t)
            _bpg['transactions'] = _deduped

        # Recalculate running balance after deduplication
        for _bpg in _bp_groups.values():
            _bal = 0.0
            for _t in _bpg['transactions']:
                _bal += float(_t.get('Debit', 0) or 0) - float(_t.get('Credit', 0) or 0)
                _t['RunningBalance'] = _bal

        # Group BP groups under their territory for display ordering
        _by_territory = {}
        for _bpg in _bp_groups.values():
            _by_territory.setdefault(_bpg['Territory'] or '', []).append(_bpg)

        grand_debit = grand_credit = 0.0

        for _terr_label, _bp_list in _by_territory.items():
            if _terr_label:
                table_data.append([_terr_label] + [''] * (NCOLS - 1))
                table_styles += [
                    ('SPAN',       (0, ri), (NCOLS - 1, ri)),
                    ('BACKGROUND', (0, ri), (-1, ri), C_TERR_BG),
                    ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('FONTSIZE',   (0, ri), (-1, ri), 8),
                    ('LINEABOVE',  (0, ri), (-1, ri), 0.5, C_GRID),
                ]
                ri += 1

            for _bpg in _bp_list:
                _bp_label = (
                    f"{_bpg['BPName']} ({_bpg['BPCode']})"
                    if _bpg['BPCode'] else (_bpg['BPName'] or 'No BP')
                )
                table_data.append([_bp_label] + [''] * (NCOLS - 1))
                table_styles += [
                    ('SPAN',         (0, ri), (NCOLS - 1, ri)),
                    ('FONTNAME',     (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('FONTSIZE',     (0, ri), (-1, ri), 7),
                    ('LINEABOVE',    (0, ri), (-1, ri), 0.5, C_GRID),
                    ('LINEBELOW',    (0, ri), (-1, ri), 0.5, C_GRID),
                    ('BOTTOMPADDING',(0, ri), (-1, ri), 4),
                ]
                ri += 1

                bp_debit = bp_credit = 0.0
                for _txn in _bpg['transactions']:
                    _debit   = float(_txn.get('Debit',  0) or 0)
                    _credit  = float(_txn.get('Credit', 0) or 0)
                    _running = float(_txn.get('RunningBalance', 0) or 0)
                    bp_debit  += _debit
                    bp_credit += _credit

                    _pdate = str(_txn.get('PostingDate', ''))[:10]
                    try:
                        _pd = datetime.strptime(_pdate, '%Y-%m-%d')
                        _pdate = f"{_pd.month}/{_pd.day}/{_pd.year}"
                    except Exception:
                        pass

                    _pol_name  = str(_txn.get('AccountName', '') or '')
                    _narration = str(_txn.get('Description', '') or '')
                    # Suppress narration when it is identical to policy name
                    if _narration.strip() == _pol_name.strip():
                        _narration = ''

                    table_data.append([
                        '',
                        '',
                        str(_txn.get('TransTypeName') or _txn.get('TransType', '') or ''),
                        Paragraph(_pol_name,  cell_style),
                        Paragraph(_narration, cell_style),
                        str(_txn.get('Reference1', '') or ''),
                        '{:,.2f}'.format(_debit)   if _debit   > 0 else '',
                        '{:,.2f}'.format(_credit)  if _credit  > 0 else '',
                        '{:,.2f}'.format(_running) if _running != 0 else '0',
                    ])
                    ri += 1

                # Closing balance per BP
                _bp_close = bp_debit - bp_credit
                table_data.append(
                    ['Closing Balance'] + [''] * (NCOLS - 4)
                    + [
                        '{:,.2f}'.format(bp_debit),
                        '{:,.2f}'.format(bp_credit),
                        '{:,.2f}'.format(_bp_close) if _bp_close != 0 else '0',
                    ]
                )
                table_styles += [
                    ('SPAN',       (0, ri), (NCOLS - 4, ri)),
                    ('BACKGROUND', (0, ri), (-1, ri), C_CLOSE_BG),
                    ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('LINEABOVE',  (0, ri), (-1, ri), 0.5, C_GRID),
                ]
                ri += 1

                grand_debit  += bp_debit
                grand_credit += bp_credit

        # Grand total row
        _grand_close = grand_debit - grand_credit
        table_data.append(
            ['GRAND TOTAL'] + [''] * (NCOLS - 4)
            + [
                '{:,.2f}'.format(grand_debit),
                '{:,.2f}'.format(grand_credit),
                '{:,.2f}'.format(_grand_close) if _grand_close != 0 else '0',
            ]
        )
        table_styles += [
            ('SPAN',       (0, ri), (NCOLS - 4, ri)),
            ('BACKGROUND', (0, ri), (-1, ri), C_GRAND_BG2),
            ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, ri), (-1, ri), 8),
            ('LINEABOVE',  (0, ri), (-1, ri), 1, colors.HexColor('#64748b')),
        ]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_styles))
        elements.append(table)

        # ── Per-policy summary page ──────────────────────────────────────────
        # Aggregates each (Customer, Project) pair across the same filters
        # used for the main ledger. Rendered on its own page after the detail.
        try:
            conn2 = get_hana_connection(company)
            summary_rows = hana_queries.project_summary_report(
                conn2,
                account_from=account_from or None,
                account_to=account_to   or None,
                from_date=from_date     or None,
                to_date=to_date         or None,
                bp_code=bp_code if bp_code else None,
                project_code=project_code or None,
                trans_type=trans_type   or None,
            )

            # Group by BP and attach per-policy opening + closing balance
            _by_bp = {}
            for _r in summary_rows:
                _bpc = str(_r.get('BPCode') or '')
                _bp = _by_bp.setdefault(_bpc, {
                    'BPCode': _bpc,
                    'BPName': str(_r.get('BPName') or ''),
                    'rows': [],
                })
                _opening = 0.0
                if from_date:
                    try:
                        _opening = hana_queries.project_opening_balance(
                            conn2, _bpc, str(_r.get('ProjectCode') or ''), from_date
                        )
                    except Exception:
                        _opening = 0.0
                _td = float(_r.get('TotalDebit')  or 0)
                _tc = float(_r.get('TotalCredit') or 0)
                _r['Opening'] = _opening
                _r['Balance'] = _opening + _td - _tc
                _bp['rows'].append(_r)
            conn2.close()
        except Exception as _summary_err:
            logger.exception("Error building per-policy summary: %s", _summary_err)
            _by_bp = {}

        if _by_bp:
            elements.append(PageBreak())

            _sum_col_widths = [
                3.40 * inch,   # Policy Name
                0.80 * inch,   # Opening
                0.80 * inch,   # Sales
                0.70 * inch,   # Tax
                0.80 * inch,   # Return
                0.90 * inch,   # Collection
                0.85 * inch,   # Total Debit
                0.85 * inch,   # Total Credit
                0.85 * inch,   # Balance
            ]
            _SUM_NCOLS = len(_sum_col_widths)

            _sum_data = []
            _sum_styles = [
                ('FONTSIZE',      (0, 0), (-1, -1), _fs_td),
                ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING',   (0, 0), (-1, -1), 3),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
                ('TOPPADDING',    (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LINEBELOW',     (0, 0), (-1, -1), 0.25, C_GRID),
                ('ALIGN',         (1, 0), (-1, -1), 'RIGHT'),
            ]

            _sum_headers = [
                'Policy Name', 'Opening', 'Sales', 'Tax', 'Return',
                'Collection', 'Total Debit', 'Total Credit', 'Balance',
            ]
            _sum_data.append(_sum_headers)
            _sri = 0
            _sum_styles += [
                ('BACKGROUND',    (0, _sri), (-1, _sri), C_HEADER_BG),
                ('TEXTCOLOR',     (0, _sri), (-1, _sri), C_HEADER_TXT),
                ('FONTNAME',      (0, _sri), (-1, _sri), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, _sri), (-1, _sri), _fs_th),
                ('ALIGN',         (0, _sri), (-1, _sri), 'CENTER'),
                ('BOTTOMPADDING', (0, _sri), (-1, _sri), 5),
                ('TOPPADDING',    (0, _sri), (-1, _sri), 5),
                ('LINEABOVE',     (0, _sri), (-1, _sri), 0.75, C_GRID),
                ('LINEBELOW',     (0, _sri), (-1, _sri), 0.75, C_GRID),
            ]
            _sri += 1

            def _fmt(v):
                try:
                    f = float(v or 0)
                except (TypeError, ValueError):
                    f = 0.0
                return '{:,.2f}'.format(f) if f else '0'

            for _bpc, _bp in _by_bp.items():
                _bp_label = (
                    f"{_bpc}    {_bp['BPName']}".strip()
                    if _bpc else (_bp['BPName'] or 'No BP')
                )
                _sum_data.append([_bp_label] + [''] * (_SUM_NCOLS - 1))
                _sum_styles += [
                    ('SPAN',         (0, _sri), (_SUM_NCOLS - 1, _sri)),
                    ('BACKGROUND',   (0, _sri), (-1, _sri), C_TERR_BG),
                    ('FONTNAME',     (0, _sri), (-1, _sri), 'Helvetica-Bold'),
                    ('FONTSIZE',     (0, _sri), (-1, _sri), 8),
                    ('LINEABOVE',    (0, _sri), (-1, _sri), 0.5, C_GRID),
                    ('LINEBELOW',    (0, _sri), (-1, _sri), 0.5, C_GRID),
                ]
                _sri += 1

                for _r in _bp['rows']:
                    _pname = (
                        f"{_r.get('ProjectCode','')} {_r.get('ProjectName','')}"
                    ).strip()
                    _sum_data.append([
                        Paragraph(_pname, cell_style),
                        _fmt(_r.get('Opening')),
                        _fmt(_r.get('Sales')),
                        _fmt(_r.get('Tax')),
                        _fmt(_r.get('Return')),
                        _fmt(_r.get('Collection')),
                        _fmt(_r.get('TotalDebit')),
                        _fmt(_r.get('TotalCredit')),
                        _fmt(_r.get('Balance')),
                    ])
                    _sri += 1

            _sum_table = Table(_sum_data, colWidths=_sum_col_widths, repeatRows=1)
            _sum_table.setStyle(TableStyle(_sum_styles))
            elements.append(_sum_table)

        # ── Urdu footer image (flows immediately after the last table) ───────
        if _cfg.footer_image:
            try:
                from reportlab.lib.utils import ImageReader as _ImgReader
                _ir = _ImgReader(_cfg.footer_image.path)
                _iw, _ih = _ir.getSize()
                _img_w = landscape(letter)[0] - 0.6 * inch
                _img_h = _img_w * (_ih / _iw) if _iw else 1.2 * inch
                _img_max_h = 2.0 * inch
                if _img_h > _img_max_h:
                    _img_h = _img_max_h
                    _img_w = _img_h * (_iw / _ih) if _ih else _img_w
                elements.append(RLImage(_cfg.footer_image.path,
                                        width=_img_w, height=_img_h))
            except Exception as _e:
                print(f"[General Ledger] ERROR appending footer image: {_e}")

        # ── Urdu footer (lines from admin settings) ──────────────────────────
        _urdu_lines = _cfg.footer_lines()
        _urdu_reshaped_lines = []

        if _urdu_lines:
            print(f"[General Ledger] ARABIC_AVAILABLE: {ARABIC_AVAILABLE}")
            print(f"[General Ledger] Using font: {_URDU_FONT}")
            
            for _line in _urdu_lines:
                if ARABIC_AVAILABLE:
                    try:
                        if _urdu_reshaper is not None:
                            _reshaped = _urdu_reshaper.reshape(_line)
                        else:
                            _reshaped = arabic_reshaper.reshape(_line)
                        _display  = bidi_display(_reshaped)
                        _urdu_reshaped_lines.append(_display)
                    except Exception as e:
                        _urdu_reshaped_lines.append(_line)
                else:
                    _urdu_reshaped_lines.append(_line)

        # Callback to draw Urdu text footer — invoked by _LastPageOnlyCanvas on
        # the final page only. Skipped entirely when a footer_image is set
        # (the image is appended as a flowable above so it sits right after
        # the table).
        def draw_urdu_footer(canvas):
            """Draw Urdu text footer at the bottom of the last page."""
            if _cfg.footer_image:
                return
            if not _urdu_reshaped_lines:
                return

            page_width    = landscape(letter)[0]
            bottom_margin = 0.4 * inch
            box_width     = page_width - 0.6 * inch
            box_x         = 0.3 * inch

            canvas.saveState()
            try:
                from reportlab.lib.utils import simpleSplit

                _font_size = 10
                _padding   = 8
                _line_gap  = 3
                max_text_w = box_width - (_padding * 2)

                # Pre-split all lines accounting for wrapping
                all_wrapped = []
                for line_text in _urdu_reshaped_lines:
                    wrapped = simpleSplit(line_text, _URDU_FONT, _font_size, max_text_w)
                    all_wrapped.append(wrapped if wrapped else [line_text])

                # Calculate total height needed
                single_line_h = _font_size * 1.4
                y_pos = bottom_margin + 5

                for wrapped_lines in reversed(all_wrapped):
                    box_h = (single_line_h * len(wrapped_lines)) + (_padding * 2)
                    # Draw background box (no border)
                    canvas.setFillColor(colors.HexColor('#f8f9ff'))
                    canvas.rect(box_x, y_pos, box_width, box_h, fill=1, stroke=0)
                    # Draw each wrapped line left-anchored (bidi puts visual-left char first)
                    canvas.setFillColor(colors.HexColor('#1e293b'))
                    canvas.setFont(_URDU_FONT, _font_size)
                    text_x = box_x + _padding
                    text_y = y_pos + box_h - _padding - _font_size
                    for sub_line in wrapped_lines:
                        canvas.drawString(text_x, text_y, sub_line)
                        text_y -= single_line_h
                    y_pos += box_h + _line_gap

            except Exception as e:
                print(f"[General Ledger] ERROR drawing footer: {e}")
                import traceback
                traceback.print_exc()
            finally:
                canvas.restoreState()

        doc.build(elements, canvasmaker=_make_last_page_only_canvas(draw_urdu_footer))
        pdf_buffer.seek(0)
        response.write(pdf_buffer.getvalue())
        return response

    except Exception as e:
        logger.exception("Error exporting PDF")
        return Response({
            'success': False,
            'error': f'Error exporting PDF: {str(e)}'
        }, status=500)


@staff_member_required
def export_ledger_pdf(request):
    """
    Export general ledger to PDF — matches the admin HTML table design.
    Grouped by account with account header, opening balance, transactions,
    subtotal, and grand total rows using the same colour scheme as the UI.
    """
    if not REPORTLAB_AVAILABLE:
        return HttpResponse(
            "PDF export requires reportlab. Install it with: pip install reportlab",
            status=500
        )

    try:
        company      = (request.GET.get('company')      or '').strip()
        account_from = (request.GET.get('account_from') or '').strip()
        account_to   = (request.GET.get('account_to')   or '').strip()
        from_date    = (request.GET.get('from_date')    or '').strip()
        to_date      = (request.GET.get('to_date')      or '').strip()
        bp_code      = (request.GET.get('bp_code')      or '').strip()
        project_code = (request.GET.get('project_code') or '').strip()
        trans_type   = (request.GET.get('trans_type')   or '').strip()

        # ── Fetch data ────────────────────────────────────────────────────────
        if not from_date or not to_date:
            return HttpResponse('from_date and to_date are required for export.', status=400)

        conn = get_hana_connection(company)
        transactions = hana_queries.general_ledger_report(
            conn,
            account_from=account_from or None,
            account_to=account_to or None,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code or None,
            project_code=project_code or None,
            trans_type=trans_type or None,
            limit=50000,
        )

        # Group & calculate balances (same logic as the admin view)
        grouped_raw = group_by_account(transactions)
        accounts_grouped = []
        for account_code, account_data in grouped_raw.items():
            opening_balance = 0.0
            if from_date:
                opening = hana_queries.account_opening_balance(
                    conn, account_code, from_date, bp_code=bp_code or None
                )
                opening_balance = float(opening.get('Balance', 0))
            account_data['transactions'] = calculate_running_balance(account_data['transactions'])
            totals = calculate_totals(account_data['transactions'])
            accounts_grouped.append({
                'Account':        account_data['Account'],
                'AccountName':    account_data['AccountName'],
                'OpeningBalance': opening_balance,
                'transactions':   account_data['transactions'],
                'TotalDebit':     totals['TotalDebit'],
                'TotalCredit':    totals['TotalCredit'],
                'ClosingBalance': opening_balance + totals['Balance'],
            })
        conn.close()

        grand_total = None
        if accounts_grouped:
            grand_total = {
                'TotalDebit':  sum(a['TotalDebit']  for a in accounts_grouped),
                'TotalCredit': sum(a['TotalCredit'] for a in accounts_grouped),
            }
            grand_total['Difference'] = grand_total['TotalDebit'] - grand_total['TotalCredit']

        # ── PDF setup ─────────────────────────────────────────────────────────
        response = HttpResponse(content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = (
            f'attachment; filename="general_ledger_{company}_{timestamp}.pdf"'
        )

        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            topMargin=0.4 * inch, bottomMargin=0.4 * inch,
            leftMargin=0.3 * inch, rightMargin=0.3 * inch,
        )
        elements = []
        styles = getSampleStyleSheet()

        # cell_style defined after _cfg is loaded (below)

        # ── Date range display (computed before header table) ─────────────────
        from_disp = from_date or '01/01/2000'
        to_disp   = to_date   or datetime.now().strftime('%m/%d/%Y')
        try:
            if from_date and len(from_date) == 10:
                d = datetime.strptime(from_date, '%Y-%m-%d')
                from_disp = f"{d.month}/{d.day}/{d.year}"
        except Exception:
            pass
        try:
            if to_date and len(to_date) == 10:
                d = datetime.strptime(to_date, '%Y-%m-%d')
                to_disp = f"{d.month}/{d.day}/{d.year}"
        except Exception:
            pass
        now = datetime.now()
        print_date_str = f"{now.month}/{now.day}/{now.year}   {now.strftime('%I:%M:%S%p')}"

        # ── Resolve company full name ─────────────────────────────────────────
        _company_full_name = company  # fallback to the key passed
        try:
            from FieldAdvisoryService.models import Company as _CompanyModel
            _co = _CompanyModel.objects.filter(name=company).first() \
                  or _CompanyModel.objects.filter(Company_name=company).first()
            if _co:
                _company_full_name = getattr(_co, 'Company_name', None) or company
        except Exception:
            pass

        # ── Load admin-managed settings ────────────────────────────────────
        from .models import LedgerSettings as _LedgerSettings
        _cfg = _LedgerSettings.get()

        _fs_grp  = int(_cfg.font_size_group_name   or 11)
        _fs_ent  = int(_cfg.font_size_company_name  or 13)
        _fs_ttl  = int(_cfg.font_size_report_title  or 10)
        _fs_dt   = int(_cfg.font_size_dates         or 9)
        _fs_td   = int(_cfg.font_size_table_data    or 8)

        cell_style = ParagraphStyle(
            'PDFCell', parent=styles['Normal'],
            fontSize=_fs_td, fontName='Helvetica', leading=_fs_td + 2, alignment=TA_LEFT,
        )

        # ── Header styles ─────────────────────────────────────────────────────
        _h_group = ParagraphStyle('HDRGroup', parent=styles['Normal'],
            fontSize=_fs_grp, fontName='Helvetica-Bold', leading=_fs_grp + 3)
        _h_entity = ParagraphStyle('HDREntity', parent=styles['Normal'],
            fontSize=_fs_ent, fontName='Helvetica-Bold', leading=_fs_ent + 4)
        _h_title = ParagraphStyle('HDRTitle', parent=styles['Normal'],
            fontSize=_fs_ttl, fontName='Helvetica', leading=_fs_ttl + 3)
        _h_dates = ParagraphStyle('HDRDates', parent=styles['Normal'],
            fontSize=_fs_dt, fontName='Helvetica', leading=_fs_dt + 3)
        _h_print = ParagraphStyle('HDRPrint', parent=styles['Normal'],
            fontSize=8, fontName='Helvetica', leading=11, alignment=TA_RIGHT)

        # ── Logo path (from admin-uploaded image or legacy path) ──────────────
        import os as _os
        if _cfg.smart_stamp_image:
            _logo_path = _cfg.smart_stamp_image.path
        else:
            _logo_path = _os.path.join(
                _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                'media', 'Ledger_template', 'smart_stamp.png'
            )

        # ── Page width (needed for nested tables) ────────────────────────────
        _page_w = landscape(letter)[0] - 0.6 * inch   # left+right margin = 0.3*2
        _right_col_w = _page_w * 0.30

        # ── SAP logo: use uploaded image if available, else draw programmatically ──
        _sap_logo = None
        if _cfg.sap_logo_image:
            try:
                _sap_logo = RLImage(_cfg.sap_logo_image.path, width=0.65 * inch, height=0.33 * inch)
            except Exception:
                _sap_logo = None
        if _sap_logo is None:
            try:
                from reportlab.graphics.shapes import Drawing as _RLDrawing, Rect as _RLRect, String as _RLString
                _sap_w, _sap_h = 0.55 * inch, 0.28 * inch
                _sap_d = _RLDrawing(_sap_w, _sap_h)
                _sap_d.add(_RLRect(0, 0, _sap_w, _sap_h,
                                   fillColor=colors.HexColor('#0070f3'),
                                   strokeColor=colors.HexColor('#0070f3')))
                _sap_d.add(_RLString(_sap_w / 2, (_sap_h / 2) - 4, 'SAP',
                                     fontSize=12, fontName='Helvetica-Bold',
                                     fillColor=colors.white, textAnchor='middle'))
                _sap_logo = _sap_d
            except Exception:
                _sap_logo = None

        # ── Right sub-table: [ print-date  |  SAP-logo ] ─────────────────────
        _pdate_para = Paragraph(f'Print Date:   {print_date_str}', _h_print)
        _sap_cell = _sap_logo if _sap_logo is not None else Paragraph('SAP', _h_print)
        _right_top = Table(
            [[_pdate_para, _sap_cell]],
            colWidths=[_right_col_w - 0.65 * inch, 0.65 * inch],
        )
        _right_top.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN',         (0, 0), (0, 0),   'RIGHT'),
            ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ]))

        # ── Build header as 2-column table ───────────────────────────────────
        # Left cell: group name / entity name / report title / date range
        # Right cell: print date + SAP logo (top) + Smart Agriculture stamp (below)
        _left_content = [
            Paragraph(_cfg.group_name or 'Four Brothers Group', _h_group),
            Paragraph((_cfg.company_name or _company_full_name or '').upper(), _h_entity),
            Paragraph(_cfg.report_title or 'General Ledger', _h_title),
            Spacer(1, 2),
            Paragraph(f'From:  {from_disp}     To:  {to_disp}', _h_dates),
        ]

        _right_content = [_right_top]
        if _os.path.exists(_logo_path):
            _stamp = RLImage(_logo_path, width=1.1 * inch, height=1.1 * inch)
            _stamp.hAlign = 'RIGHT'
            _stamp.preserveAspectRatio = True
            _right_content.append(Spacer(1, 4))
            _right_content.append(_stamp)

        _hdr_table = Table(
            [[_left_content, _right_content]],
            colWidths=[_page_w * 0.70, _right_col_w],
        )
        _hdr_table.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ]))
        elements.append(_hdr_table)

        # Divider line below header
        elements.append(Spacer(1, 4))
        _divider = Table([['']], colWidths=[_page_w])
        _divider.setStyle(TableStyle([
            ('LINEBELOW', (0, 0), (-1, -1), 0.75, colors.HexColor('#1e293b')),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(_divider)
        elements.append(Spacer(1, 4))

        # ── Build territory lookup from Dealer model ──────────────────────────
        _bp_territory_map = {}
        try:
            from FieldAdvisoryService.models import Dealer as _Dealer
            for _d in _Dealer.objects.filter(card_code__isnull=False).select_related('region', 'zone', 'territory'):
                _parts = [p for p in [
                    getattr(_d.region,    'name', None),
                    getattr(_d.zone,      'name', None),
                    getattr(_d.territory, 'name', None),
                ] if p]
                _bp_territory_map[_d.card_code] = ' / '.join(_parts)
        except Exception:
            pass

        # ── Group flat transactions by BP ─────────────────────────────────────
        _bp_groups = {}
        _all_txns = [t for acct in accounts_grouped for t in acct['transactions']]
        for _txn in _all_txns:
            _bpc = str(_txn.get('BPCode', '') or '')
            if _bpc not in _bp_groups:
                _bp_groups[_bpc] = {
                    'BPCode': _bpc,
                    'BPName': str(_txn.get('BPName', '') or ''),
                    'Territory': _bp_territory_map.get(_bpc, ''),
                    'transactions': [],
                }
            _bp_groups[_bpc]['transactions'].append(_txn)

        # Recalculate running balance per BP
        for _bpg in _bp_groups.values():
            _bal = 0.0
            for _t in _bpg['transactions']:
                _bal += float(_t.get('Debit', 0) or 0) - float(_t.get('Credit', 0) or 0)
                _t['RunningBalance'] = _bal

        # ── Column widths (landscape letter ≈ 10.4" usable) ──────────────────
        col_widths = [
            0.65 * inch,   # VNo.
            0.80 * inch,   # VDate
            0.68 * inch,   # Type
            1.70 * inch,   # Policy Name (Account Name)
            2.50 * inch,   # Narration (Description)
            0.90 * inch,   # PR NO. (Reference)
            0.90 * inch,   # Debit
            0.90 * inch,   # Credit
            0.90 * inch,   # Balance
        ]
        NCOLS = len(col_widths)

        C_HEADER_BG  = colors.HexColor(_cfg.table_header_bg_color   or '#1e293b')
        C_HEADER_TXT = colors.HexColor(_cfg.table_header_text_color  or '#ffffff')
        C_TERR_BG    = colors.HexColor(_cfg.territory_row_bg_color   or '#f1f5f9')
        C_CLOSE_BG   = colors.HexColor(_cfg.closing_balance_bg_color or '#f1f5f9')
        C_GRAND_BG2  = colors.HexColor(_cfg.grand_total_bg_color     or '#e2e8f0')
        C_GRID       = colors.HexColor(_cfg.grid_color               or '#d1d5db')

        _fs_th = int(_cfg.font_size_table_header or 8)

        table_data   = []
        table_styles = [
            ('FONTSIZE',      (0, 0), (-1, -1), _fs_td),
            ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 3),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
            ('TOPPADDING',    (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            # Horizontal lines only - no vertical column borders
            ('LINEBELOW',     (0, 0), (-1, -1), 0.25, C_GRID),
            ('LINEBEFORE',    (0, 0), (0, -1),  0,    colors.white),
            ('LINEAFTER',     (-1, 0), (-1, -1), 0,   colors.white),
            ('ALIGN',         (6, 0), (8, -1), 'RIGHT'),
        ]

        headers = ['VNo.', 'VDate', 'Type', 'Policy Name', 'Narration', 'PR NO.', 'Debit', 'Credit', 'Balance']
        table_data.append(headers)
        ri = 0
        table_styles += [
            ('BACKGROUND',    (0, ri), (-1, ri), C_HEADER_BG),
            ('TEXTCOLOR',     (0, ri), (-1, ri), C_HEADER_TXT),
            ('FONTNAME',      (0, ri), (-1, ri), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, ri), (-1, ri), _fs_th),
            ('ALIGN',         (0, ri), (-1, ri), 'CENTER'),
            ('BOTTOMPADDING', (0, ri), (-1, ri), 5),
            ('TOPPADDING',    (0, ri), (-1, ri), 5),
            # Top and bottom border around header row
            ('LINEABOVE',     (0, ri), (-1, ri), 0.75, C_GRID),
            ('LINEBELOW',     (0, ri), (-1, ri), 0.75, C_GRID),
        ]
        ri += 1

        # Deduplicate transactions per BP by TransId
        for _bpg in _bp_groups.values():
            _seen_ids = set()
            _deduped  = []
            for _t in _bpg['transactions']:
                _tid = str(_t.get('TransId', '') or '')
                if _tid and _tid in _seen_ids:
                    continue
                if _tid:
                    _seen_ids.add(_tid)
                _deduped.append(_t)
            _bpg['transactions'] = _deduped

        # Recalculate running balance after deduplication
        for _bpg in _bp_groups.values():
            _bal = 0.0
            for _t in _bpg['transactions']:
                _bal += float(_t.get('Debit', 0) or 0) - float(_t.get('Credit', 0) or 0)
                _t['RunningBalance'] = _bal

        # Group BP groups under their territory for display ordering
        _by_territory = {}
        for _bpg in _bp_groups.values():
            _by_territory.setdefault(_bpg['Territory'] or '', []).append(_bpg)

        grand_debit = grand_credit = 0.0

        for _terr_label, _bp_list in _by_territory.items():
            if _terr_label:
                table_data.append([_terr_label] + [''] * (NCOLS - 1))
                table_styles += [
                    ('SPAN',       (0, ri), (NCOLS - 1, ri)),
                    ('BACKGROUND', (0, ri), (-1, ri), C_TERR_BG),
                    ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('FONTSIZE',   (0, ri), (-1, ri), 8),
                    ('LINEABOVE',  (0, ri), (-1, ri), 0.5, C_GRID),
                ]
                ri += 1

            for _bpg in _bp_list:
                _bp_label = (
                    f"{_bpg['BPName']} ({_bpg['BPCode']})"
                    if _bpg['BPCode'] else (_bpg['BPName'] or 'No BP')
                )
                table_data.append([_bp_label] + [''] * (NCOLS - 1))
                table_styles += [
                    ('SPAN',         (0, ri), (NCOLS - 1, ri)),
                    ('FONTNAME',     (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('FONTSIZE',     (0, ri), (-1, ri), 7),
                    ('LINEABOVE',    (0, ri), (-1, ri), 0.5, C_GRID),
                    ('LINEBELOW',    (0, ri), (-1, ri), 0.5, C_GRID),
                    ('BOTTOMPADDING',(0, ri), (-1, ri), 4),
                ]
                ri += 1

                bp_debit = bp_credit = 0.0
                for _txn in _bpg['transactions']:
                    _debit   = float(_txn.get('Debit',  0) or 0)
                    _credit  = float(_txn.get('Credit', 0) or 0)
                    _running = float(_txn.get('RunningBalance', 0) or 0)
                    bp_debit  += _debit
                    bp_credit += _credit

                    _pdate = str(_txn.get('PostingDate', ''))[:10]
                    try:
                        _pd = datetime.strptime(_pdate, '%Y-%m-%d')
                        _pdate = f"{_pd.month}/{_pd.day}/{_pd.year}"
                    except Exception:
                        pass

                    _pol_name  = str(_txn.get('AccountName', '') or '')
                    _narration = str(_txn.get('Description', '') or '')
                    # Suppress narration when it is identical to policy name
                    if _narration.strip() == _pol_name.strip():
                        _narration = ''

                    table_data.append([
                        '',
                        '',
                        str(_txn.get('TransTypeName') or _txn.get('TransType', '') or ''),
                        Paragraph(_pol_name,  cell_style),
                        Paragraph(_narration, cell_style),
                        str(_txn.get('Reference1', '') or ''),
                        '{:,.2f}'.format(_debit)   if _debit   > 0 else '',
                        '{:,.2f}'.format(_credit)  if _credit  > 0 else '',
                        '{:,.2f}'.format(_running) if _running != 0 else '0',
                    ])
                    ri += 1

                # Closing balance per BP
                _bp_close = bp_debit - bp_credit
                table_data.append(
                    ['Closing Balance'] + [''] * (NCOLS - 4)
                    + [
                        '{:,.2f}'.format(bp_debit),
                        '{:,.2f}'.format(bp_credit),
                        '{:,.2f}'.format(_bp_close) if _bp_close != 0 else '0',
                    ]
                )
                table_styles += [
                    ('SPAN',       (0, ri), (NCOLS - 4, ri)),
                    ('BACKGROUND', (0, ri), (-1, ri), C_CLOSE_BG),
                    ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('LINEABOVE',  (0, ri), (-1, ri), 0.5, C_GRID),
                ]
                ri += 1

                grand_debit  += bp_debit
                grand_credit += bp_credit

        # Grand total row
        _grand_close = grand_debit - grand_credit
        table_data.append(
            ['GRAND TOTAL'] + [''] * (NCOLS - 4)
            + [
                '{:,.2f}'.format(grand_debit),
                '{:,.2f}'.format(grand_credit),
                '{:,.2f}'.format(_grand_close) if _grand_close != 0 else '0',
            ]
        )
        table_styles += [
            ('SPAN',       (0, ri), (NCOLS - 4, ri)),
            ('BACKGROUND', (0, ri), (-1, ri), C_GRAND_BG2),
            ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, ri), (-1, ri), 8),
            ('LINEABOVE',  (0, ri), (-1, ri), 1, colors.HexColor('#64748b')),
        ]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_styles))
        elements.append(table)

        # ── Urdu footer image (flows immediately after the table) ────────────
        if _cfg.footer_image:
            try:
                from reportlab.lib.utils import ImageReader as _ImgReader
                _ir = _ImgReader(_cfg.footer_image.path)
                _iw, _ih = _ir.getSize()
                _img_w = landscape(letter)[0] - 0.6 * inch
                _img_h = _img_w * (_ih / _iw) if _iw else 1.2 * inch
                _img_max_h = 2.0 * inch
                if _img_h > _img_max_h:
                    _img_h = _img_max_h
                    _img_w = _img_h * (_iw / _ih) if _ih else _img_w
                elements.append(RLImage(_cfg.footer_image.path,
                                        width=_img_w, height=_img_h))
            except Exception as _e:
                print(f"[General Ledger] ERROR appending footer image: {_e}")

        # ── Urdu footer ───────────────────────────────────────────────────────
        _urdu_lines = _cfg.footer_lines()
        _urdu_reshaped_lines = []

        print(f"[General Ledger] ARABIC_AVAILABLE: {ARABIC_AVAILABLE}")
        print(f"[General Ledger] Using font: {_URDU_FONT}")
        
        for _line in _urdu_lines:
            if ARABIC_AVAILABLE:
                try:
                    _reshaped = arabic_reshaper.reshape(_line)
                    _display  = bidi_display(_reshaped)
                    _urdu_reshaped_lines.append(_display)
                except Exception as e:
                    _urdu_reshaped_lines.append(_line)
            else:
                _urdu_reshaped_lines.append(_line)

        # Callback to draw Urdu text footer — invoked by _LastPageOnlyCanvas on
        # the final page only. Skipped entirely when a footer_image is set
        # (the image is appended as a flowable above so it sits right after
        # the table).
        def draw_urdu_footer(canvas):
            """Draw Urdu text footer at the bottom of the last page."""
            if _cfg.footer_image:
                return
            if not _urdu_reshaped_lines:
                return
            page_width    = landscape(letter)[0]
            bottom_margin = 0.4 * inch
            box_width     = page_width - 0.6 * inch
            box_x         = 0.3 * inch
            canvas.saveState()
            try:
                from reportlab.lib.utils import simpleSplit
                _font_size = 10
                _padding   = 8
                _line_gap  = 3
                max_text_w = box_width - (_padding * 2)
                single_line_h = _font_size * 1.4
                all_wrapped = []
                for line_text in _urdu_reshaped_lines:
                    wrapped = simpleSplit(line_text, _URDU_FONT, _font_size, max_text_w)
                    all_wrapped.append(wrapped if wrapped else [line_text])
                y_pos = bottom_margin + 5
                for wrapped_lines in reversed(all_wrapped):
                    box_h = (single_line_h * len(wrapped_lines)) + (_padding * 2)
                    canvas.setFillColor(colors.HexColor('#f8f9ff'))
                    canvas.rect(box_x, y_pos, box_width, box_h, fill=1, stroke=0)
                    # Draw each wrapped line left-anchored (bidi puts visual-left char first)
                    canvas.setFillColor(colors.HexColor('#1e293b'))
                    canvas.setFont(_URDU_FONT, _font_size)
                    text_x = box_x + _padding
                    text_y = y_pos + box_h - _padding - _font_size
                    for sub_line in wrapped_lines:
                        canvas.drawString(text_x, text_y, sub_line)
                        text_y -= single_line_h
                    y_pos += box_h + _line_gap
            except Exception as e:
                print(f"[General Ledger] ERROR drawing footer: {e}")
                import traceback
                traceback.print_exc()
            finally:
                canvas.restoreState()

        doc.build(elements, canvasmaker=_make_last_page_only_canvas(draw_urdu_footer))
        pdf_buffer.seek(0)
        response.write(pdf_buffer.getvalue())
        return response

    except Exception as e:
        logger.exception("Error exporting PDF")
        return HttpResponse(f"Error exporting PDF: {str(e)}", status=500)


# ─────────────────────────────────────────────────────────────────────────────
# Customer Detail Ledger — PDF export
# ─────────────────────────────────────────────────────────────────────────────

@swagger_auto_schema(
    method='get',
    tags=['General Ledger'],
    operation_summary='Export Customer Detail Ledger to PDF',
    operation_description='Download customer detail ledger as PDF. Shows invoice line-level product details (Product Name, Qty, Price, Policy Discount, Sale Value) for sales/return transactions, and journal-level detail for payments, credit/debit notes. Use user parameter to filter by user\'s assigned customers.',
    manual_parameters=[
        openapi.Parameter('company',      openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: Company database key', required=False),
        openapi.Parameter('account_from', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start of account range'),
        openapi.Parameter('account_to',   openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End of account range'),
        openapi.Parameter('from_date',    openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Start date (YYYY-MM-DD)'),
        openapi.Parameter('to_date',      openapi.IN_QUERY, type=openapi.TYPE_STRING, description='End date (YYYY-MM-DD)'),
        openapi.Parameter('bp_code',      openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Business Partner CardCode. If user parameter is provided, validates that bp_code belongs to that user.'),
        openapi.Parameter('user',         openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Optional: User ID or username. Filters transactions to only user\'s assigned customers. Can be combined with bp_code for validation.'),
        openapi.Parameter('project_code', openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Project Code'),
        openapi.Parameter('trans_type',   openapi.IN_QUERY, type=openapi.TYPE_STRING, description='Transaction Type (e.g., 13, 30)'),
    ],
    responses={
        200: openapi.Response(
            description='PDF file download',
            schema=openapi.Schema(type=openapi.TYPE_FILE)
        ),
        500: 'Error generating PDF file'
    }
)
@api_view(['GET'])
def export_detail_ledger_pdf_api(request):
    """
    GET /api/general-ledger/detail-ledger/export-pdf/

    Export Customer Detail Ledger to PDF. Same header/footer/summary as the
    General Ledger but with invoice line-item columns:
    VNo. | VDate | Product Name | Qty | Price | Policy Discount | Sale Value |
    Type | Policy Name/Narration | PR NO. | Debit | Credit | Balance
    """
    if not REPORTLAB_AVAILABLE:
        return Response({
            'success': False,
            'error': 'PDF export not available. Please install reportlab package.'
        }, status=500)

    try:
        # ── Query parameters ──────────────────────────────────────────────────
        company      = (request.GET.get('company')      or '').strip()
        from_date    = (request.GET.get('from_date')    or '').strip()
        to_date      = (request.GET.get('to_date')      or '').strip()
        bp_code      = (request.GET.get('bp_code')      or '').strip()
        user_param   = (request.GET.get('user')         or '').strip()
        project_code = (request.GET.get('project_code') or '').strip()

        # ── User-based filtering (mobile/dealer) ──────────────────────────────
        if user_param:
            try:
                from accounts.models import User
                from FieldAdvisoryService.models import Dealer

                user_obj = None
                try:
                    user_obj = User.objects.get(id=int(user_param))
                except (ValueError, User.DoesNotExist):
                    user_obj = User.objects.filter(username=user_param).first()

                if not user_obj:
                    return Response({'success': False, 'error': f'User "{user_param}" not found'}, status=404)

                dealer_card_codes = list(
                    Dealer.objects.filter(user=user_obj)
                    .exclude(card_code__isnull=True).exclude(card_code='')
                    .values_list('card_code', flat=True)
                )

                if not dealer_card_codes:
                    bp_code = []
                elif bp_code:
                    if bp_code not in dealer_card_codes:
                        return Response({
                            'success': False,
                            'error': f'Business Partner "{bp_code}" does not belong to user "{user_param}"'
                        }, status=403)
                else:
                    bp_code = dealer_card_codes

            except Exception as e:
                return Response({'success': False, 'error': f'Error processing user filter: {str(e)}'}, status=500)

        # ── Fetch detail ledger data ──────────────────────────────────────────
        conn = get_hana_connection(company)
        transactions = hana_queries.detail_ledger_report(
            conn,
            from_date=from_date or None,
            to_date=to_date or None,
            bp_code=bp_code if bp_code else None,
            project_code=project_code or None,
        )

        # Group transactions by BP code
        _bp_territory_map = {}
        try:
            from FieldAdvisoryService.models import Dealer as _Dealer
            for _d in _Dealer.objects.filter(card_code__isnull=False).select_related('region', 'zone', 'territory'):
                _parts = [p for p in [
                    getattr(_d.region,    'name', None),
                    getattr(_d.zone,      'name', None),
                    getattr(_d.territory, 'name', None),
                ] if p]
                _bp_territory_map[_d.card_code] = ' / '.join(_parts)
        except Exception:
            pass

        _bp_groups = {}
        for _txn in transactions:
            _bpc = str(_txn.get('BPCode', '') or '')
            if _bpc not in _bp_groups:
                _bp_groups[_bpc] = {
                    'BPCode':    _bpc,
                    'BPName':    str(_txn.get('BPName', '') or ''),
                    'Territory': _bp_territory_map.get(_bpc, ''),
                    'transactions': [],
                }
            _bp_groups[_bpc]['transactions'].append(_txn)

        # Calculate per-BP opening balance and running balance
        _bp_opening = {}
        for _bpc in _bp_groups:
            if from_date and _bpc:
                try:
                    _ob = hana_queries.bp_opening_balance(conn, _bpc, from_date)
                    _bp_opening[_bpc] = float(_ob.get('Balance', 0) or 0)
                except Exception:
                    _bp_opening[_bpc] = 0.0
            else:
                _bp_opening[_bpc] = 0.0

        conn.close()

        # ── PDF setup ─────────────────────────────────────────────────────────
        response = HttpResponse(content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = (
            f'attachment; filename="detail_ledger_{company}_{timestamp}.pdf"'
        )

        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            topMargin=0.4 * inch, bottomMargin=0.4 * inch,
            leftMargin=0.3 * inch, rightMargin=0.3 * inch,
        )
        elements = []
        styles = getSampleStyleSheet()

        # ── Date range strings ────────────────────────────────────────────────
        from_disp = from_date or '01/01/2000'
        to_disp   = to_date   or datetime.now().strftime('%m/%d/%Y')
        try:
            if from_date and len(from_date) == 10:
                d = datetime.strptime(from_date, '%Y-%m-%d')
                from_disp = f"{d.month}/{d.day}/{d.year}"
        except Exception:
            pass
        try:
            if to_date and len(to_date) == 10:
                d = datetime.strptime(to_date, '%Y-%m-%d')
                to_disp = f"{d.month}/{d.day}/{d.year}"
        except Exception:
            pass
        now = datetime.now()
        print_date_str = f"{now.month}/{now.day}/{now.year}   {now.strftime('%I:%M:%S%p')}"

        # ── Resolve company full name ─────────────────────────────────────────
        _company_full_name = company
        try:
            from FieldAdvisoryService.models import Company as _CompanyModel
            _co = _CompanyModel.objects.filter(name=company).first() \
                  or _CompanyModel.objects.filter(Company_name=company).first()
            if _co:
                _company_full_name = getattr(_co, 'Company_name', None) or company
        except Exception:
            pass

        # ── Load admin-managed settings ────────────────────────────────────────
        from .models import LedgerSettings as _LedgerSettings
        _cfg = _LedgerSettings.get()

        _fs_grp  = int(_cfg.font_size_group_name   or 11)
        _fs_ent  = int(_cfg.font_size_company_name  or 13)
        _fs_ttl  = int(_cfg.font_size_report_title  or 10)
        _fs_dt   = int(_cfg.font_size_dates         or 9)
        _fs_td   = int(_cfg.font_size_table_data    or 8)

        cell_style = ParagraphStyle(
            'DLCell', parent=styles['Normal'],
            fontSize=_fs_td, fontName='Helvetica', leading=_fs_td + 2, alignment=TA_LEFT,
        )

        # ── Header styles ──────────────────────────────────────────────────────
        _h_group  = ParagraphStyle('DLHDRGroup',  parent=styles['Normal'],
            fontSize=_fs_grp, fontName='Helvetica-Bold', leading=_fs_grp + 3)
        _h_entity = ParagraphStyle('DLHDREntity', parent=styles['Normal'],
            fontSize=_fs_ent, fontName='Helvetica-Bold', leading=_fs_ent + 4)
        _h_title  = ParagraphStyle('DLHDRTitle',  parent=styles['Normal'],
            fontSize=_fs_ttl, fontName='Helvetica',      leading=_fs_ttl + 3)
        _h_dates  = ParagraphStyle('DLHDRDates',  parent=styles['Normal'],
            fontSize=_fs_dt,  fontName='Helvetica',      leading=_fs_dt + 3)
        _h_print  = ParagraphStyle('DLHDRPrint',  parent=styles['Normal'],
            fontSize=8, fontName='Helvetica', leading=11, alignment=TA_RIGHT)

        # ── Logo ───────────────────────────────────────────────────────────────
        import os as _os
        if _cfg.smart_stamp_image:
            _logo_path = _cfg.smart_stamp_image.path
        else:
            _logo_path = _os.path.join(
                _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                'media', 'Ledger_template', 'smart_stamp.png'
            )

        # ── 2-column header ────────────────────────────────────────────────────
        _page_w       = landscape(letter)[0] - 0.6 * inch
        _right_col_w  = _page_w * 0.30

        _sap_logo_dl = None
        if _cfg.sap_logo_image:
            try:
                _sap_logo_dl = RLImage(_cfg.sap_logo_image.path, width=0.65 * inch, height=0.33 * inch)
            except Exception:
                _sap_logo_dl = None
        if _sap_logo_dl is None:
            try:
                from reportlab.graphics.shapes import Drawing as _RLDrawing, Rect as _RLRect, String as _RLString
                _sw, _sh = 0.55 * inch, 0.28 * inch
                _sd = _RLDrawing(_sw, _sh)
                _sd.add(_RLRect(0, 0, _sw, _sh,
                                fillColor=colors.HexColor('#0070f3'),
                                strokeColor=colors.HexColor('#0070f3')))
                _sd.add(_RLString(_sw / 2, (_sh / 2) - 4, 'SAP',
                                  fontSize=12, fontName='Helvetica-Bold',
                                  fillColor=colors.white, textAnchor='middle'))
                _sap_logo_dl = _sd
            except Exception:
                _sap_logo_dl = None

        _pdate_para = Paragraph(f'Print Date:   {print_date_str}', _h_print)
        _sap_cell   = _sap_logo_dl if _sap_logo_dl is not None else Paragraph('SAP', _h_print)
        _right_top  = Table(
            [[_pdate_para, _sap_cell]],
            colWidths=[_right_col_w - 0.65 * inch, 0.65 * inch],
        )
        _right_top.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN',         (0, 0), (0, 0),   'RIGHT'),
            ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ]))

        _left_content = [
            Paragraph(_cfg.group_name or 'Four Brothers Group', _h_group),
            Paragraph((_cfg.company_name or _company_full_name or '').upper(), _h_entity),
            Paragraph('Customer Detail Ledger', _h_title),
        ]
        _right_content = [_right_top]
        if _os.path.exists(_logo_path):
            try:
                from reportlab.lib.utils import ImageReader as _IR
                _ir   = _IR(_logo_path)
                _iw, _ih = _ir.getSize()
                _max  = 1.1 * inch
                if _iw >= _ih:
                    _sw = _max
                    _sh = _max * (_ih / _iw)
                else:
                    _sh = _max
                    _sw = _max * (_iw / _ih)
            except Exception:
                _sw = _sh = 1.1 * inch
            _stamp = RLImage(_logo_path, width=_sw, height=_sh)
            _stamp.hAlign = 'RIGHT'
            _right_content.append(Spacer(1, 4))
            _right_content.append(_stamp)

        _hdr_table = Table(
            [[_left_content, _right_content]],
            colWidths=[_page_w * 0.70, _right_col_w],
        )
        _hdr_table.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 0),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ]))
        elements.append(_hdr_table)
        elements.append(Spacer(1, 4))

        _divider = Table([['']], colWidths=[_page_w])
        _divider.setStyle(TableStyle([
            ('LINEBELOW',     (0, 0), (-1, -1), 0.75, colors.HexColor('#1e293b')),
            ('TOPPADDING',    (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(_divider)
        elements.append(Spacer(1, 4))

        # ── Colours from settings ──────────────────────────────────────────────
        C_HEADER_BG  = colors.HexColor(_cfg.table_header_bg_color   or '#1e293b')
        C_HEADER_TXT = colors.HexColor(_cfg.table_header_text_color  or '#ffffff')
        C_TERR_BG    = colors.HexColor(_cfg.territory_row_bg_color   or '#f1f5f9')
        C_CLOSE_BG   = colors.HexColor(_cfg.closing_balance_bg_color or '#f1f5f9')
        C_GRAND_BG   = colors.HexColor(_cfg.grand_total_bg_color     or '#e2e8f0')
        C_GRID       = colors.HexColor(_cfg.grid_color               or '#d1d5db')

        _fs_th  = int(_cfg.font_size_table_header or 8)
        _fs_row = max(6, _fs_td - 1)   # one point smaller for compact transaction rows

        # ── Column widths (landscape letter ≈ 10.4" usable) ──────────────────
        # VNo | VDate | Product Name | Qty | Price | PolicyDiscount | SaleValue |
        # Type | Policy Name/Narration | PR NO. | Debit | Credit | Balance
        col_widths = [
            0.62 * inch,   # VNo.
            0.65 * inch,   # VDate
            1.30 * inch,   # Product Name
            0.37 * inch,   # Qty
            0.60 * inch,   # Price
            0.65 * inch,   # Policy Discount
            0.65 * inch,   # Sale Value
            0.75 * inch,   # Type
            1.90 * inch,   # Policy Name / Narration
            0.55 * inch,   # PR NO.
            0.80 * inch,   # Debit
            0.80 * inch,   # Credit
            0.80 * inch,   # Balance
        ]
        NCOLS = len(col_widths)

        # Column indices for numeric-right alignment
        _NUM_COLS = [3, 4, 5, 6, 10, 11, 12]   # Qty, Price, Disc, SaleVal, D, C, Bal

        table_data   = []
        table_styles = [
            ('FONTSIZE',      (0, 0), (-1, -1), _fs_row),
            ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING',   (0, 0), (-1, -1), 2),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
            ('TOPPADDING',    (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('LINEBELOW',     (0, 0), (-1, -1), 0.25, C_GRID),
            ('LINEBEFORE',    (0, 0), (0, -1),  0,    colors.white),
            ('LINEAFTER',     (-1, 0), (-1, -1), 0,   colors.white),
        ]
        for _nc in _NUM_COLS:
            table_styles.append(('ALIGN', (_nc, 0), (_nc, -1), 'RIGHT'))

        # Date row — spans all columns, sits right above the column headers
        table_data.append([f'From:  {from_disp}     To:  {to_disp}'] + [''] * (NCOLS - 1))
        ri = 0
        table_styles += [
            ('SPAN',          (0, ri), (NCOLS - 1, ri)),
            ('BACKGROUND',    (0, ri), (-1, ri), colors.white),
            ('FONTNAME',      (0, ri), (-1, ri), 'Helvetica'),
            ('FONTSIZE',      (0, ri), (-1, ri), int(_cfg.font_size_dates or 9)),
            ('ALIGN',         (0, ri), (-1, ri), 'LEFT'),
            ('TOPPADDING',    (0, ri), (-1, ri), 4),
            ('BOTTOMPADDING', (0, ri), (-1, ri), 2),
            ('LINEABOVE',     (0, ri), (-1, ri), 0,    colors.white),
            ('LINEBELOW',     (0, ri), (-1, ri), 0,    colors.white),
        ]
        ri += 1

        headers = [
            'VNo.', 'VDate', 'Product Name', 'Qty', 'Price',
            'Policy\nDiscount', 'Sale Value', 'Type',
            'Policy Name / Narration', 'PR NO.', 'Debit', 'Credit', 'Balance',
        ]
        table_data.append(headers)
        table_styles += [
            ('BACKGROUND',    (0, ri), (-1, ri), C_HEADER_BG),
            ('TEXTCOLOR',     (0, ri), (-1, ri), C_HEADER_TXT),
            ('FONTNAME',      (0, ri), (-1, ri), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, ri), (-1, ri), _fs_th),
            ('ALIGN',         (0, ri), (-1, ri), 'CENTER'),
            ('BOTTOMPADDING', (0, ri), (-1, ri), 5),
            ('TOPPADDING',    (0, ri), (-1, ri), 5),
            ('LINEABOVE',     (0, ri), (-1, ri), 0.75, C_GRID),
            ('LINEBELOW',     (0, ri), (-1, ri), 0.75, C_GRID),
        ]
        ri += 1

        # ── Group BP groups under territory for display ordering ───────────────
        _by_territory = {}
        for _bpg in _bp_groups.values():
            _by_territory.setdefault(_bpg['Territory'] or '', []).append(_bpg)

        def _fmt_num(v, zero_blank=False):
            try:
                f = float(v or 0)
            except (TypeError, ValueError):
                f = 0.0
            if zero_blank and f == 0:
                return ''
            return '{:,.0f}'.format(f) if f == int(f) else '{:,.2f}'.format(f)

        def _fmt_money(v):
            try:
                f = float(v or 0)
            except (TypeError, ValueError):
                f = 0.0
            return '{:,.2f}'.format(f) if f != 0 else ''

        grand_debit = grand_credit = 0.0

        for _terr_label, _bp_list in _by_territory.items():
            if _terr_label:
                table_data.append([_terr_label] + [''] * (NCOLS - 1))
                table_styles += [
                    ('SPAN',       (0, ri), (NCOLS - 1, ri)),
                    ('BACKGROUND', (0, ri), (-1, ri), colors.white),
                    ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('FONTSIZE',   (0, ri), (-1, ri), _fs_row),
                    ('LINEABOVE',  (0, ri), (-1, ri), 0.5, C_GRID),
                ]
                ri += 1

            for _bpg in _bp_list:
                _bpc      = _bpg['BPCode']
                _bp_label = (
                    f"{_bpg['BPName']} ({_bpc})"
                    if _bpc else (_bpg['BPName'] or 'No BP')
                )
                table_data.append([_bp_label] + [''] * (NCOLS - 1))
                table_styles += [
                    ('SPAN',         (0, ri), (NCOLS - 1, ri)),
                    ('FONTNAME',     (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('FONTSIZE',     (0, ri), (-1, ri), _fs_row),
                    ('LINEABOVE',    (0, ri), (-1, ri), 0.5, C_GRID),
                    ('LINEBELOW',    (0, ri), (-1, ri), 0.5, C_GRID),
                    ('BOTTOMPADDING',(0, ri), (-1, ri), 3),
                ]
                ri += 1

                # Opening balance — used for running balance calc but NOT shown as a row
                _ob = _bp_opening.get(_bpc, 0.0)

                # Period transactions with running balance
                _running = _ob
                bp_debit = bp_credit = 0.0

                _row_style = ParagraphStyle(
                    'DLRowCell', parent=cell_style, fontSize=_fs_row, leading=_fs_row + 2,
                )

                for _txn in _bpg['transactions']:
                    _debit  = float(_txn.get('Debit',  0) or 0)
                    _credit = float(_txn.get('Credit', 0) or 0)
                    _running += _debit - _credit
                    bp_debit  += _debit
                    bp_credit += _credit

                    _pdate = str(_txn.get('PostingDate', ''))[:10]
                    try:
                        _pd = datetime.strptime(_pdate, '%Y-%m-%d')
                        _pdate = f"{_pd.month}/{_pd.day}/{_pd.year}"
                    except Exception:
                        pass

                    _item_name = str(_txn.get('ItemName', '') or '')
                    _qty       = _txn.get('Qty', 0) or 0
                    _price     = _txn.get('Price', 0) or 0
                    _disc      = _txn.get('PolicyDiscount', 0) or 0
                    _sale_val  = _txn.get('SaleValue', 0) or 0
                    _type_name = str(_txn.get('TransTypeName', '') or '')
                    _pol_narr  = str(_txn.get('PolicyNarration', '') or '')
                    _ref1      = str(_txn.get('Reference1', '') or '')

                    # For SIV/SRV rows PolicyNarration is blank — fall back to project name
                    if not _pol_narr:
                        _proj_name = str(_txn.get('ProjectName', '') or '')
                        _proj_code = str(_txn.get('ProjectCode', '') or '')
                        if _proj_name:
                            _pol_narr = f"{_proj_code} {_proj_name}".strip()

                    table_data.append([
                        str(_txn.get('TransId', '') or ''),
                        _pdate,
                        Paragraph(_item_name, _row_style),
                        _fmt_num(_qty, zero_blank=True)      if _item_name else '',
                        _fmt_num(_price, zero_blank=True)    if _item_name else '',
                        _fmt_num(_disc, zero_blank=True)     if _item_name else '',
                        _fmt_num(_sale_val, zero_blank=True) if _item_name else '',
                        Paragraph(_type_name, _row_style),
                        Paragraph(_pol_narr, _row_style),
                        Paragraph(_ref1, _row_style),
                        _fmt_money(_debit)  if _debit  > 0 else '',
                        _fmt_money(_credit) if _credit > 0 else '',
                        '{:,.2f}'.format(_running) if _running != 0 else '0',
                    ])
                    ri += 1

                # Closing balance per BP
                _bp_close = _ob + bp_debit - bp_credit
                table_data.append(
                    ['Closing Balance'] + [''] * (NCOLS - 4)
                    + [
                        _fmt_money(bp_debit),
                        _fmt_money(bp_credit),
                        '{:,.2f}'.format(_bp_close) if _bp_close != 0 else '0',
                    ]
                )
                table_styles += [
                    ('SPAN',       (0, ri), (NCOLS - 4, ri)),
                    ('BACKGROUND', (0, ri), (-1, ri), C_CLOSE_BG),
                    ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
                    ('LINEABOVE',  (0, ri), (-1, ri), 0.5, C_GRID),
                ]
                ri += 1

                grand_debit  += bp_debit
                grand_credit += bp_credit

        # Grand total row
        _grand_close = grand_debit - grand_credit
        table_data.append(
            ['GRAND TOTAL'] + [''] * (NCOLS - 4)
            + [
                _fmt_money(grand_debit),
                _fmt_money(grand_credit),
                '{:,.2f}'.format(_grand_close) if _grand_close != 0 else '0',
            ]
        )
        table_styles += [
            ('SPAN',       (0, ri), (NCOLS - 4, ri)),
            ('BACKGROUND', (0, ri), (-1, ri), C_GRAND_BG),
            ('FONTNAME',   (0, ri), (-1, ri), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, ri), (-1, ri), 8),
            ('LINEABOVE',  (0, ri), (-1, ri), 1, colors.HexColor('#64748b')),
        ]

        table = Table(table_data, colWidths=col_widths, repeatRows=2)
        table.setStyle(TableStyle(table_styles))
        elements.append(table)

        # ── Per-policy summary page (reuses the same project_summary_report) ───
        try:
            conn2 = get_hana_connection(company)
            summary_rows = hana_queries.project_summary_report(
                conn2,
                from_date=from_date or None,
                to_date=to_date     or None,
                bp_code=bp_code if bp_code else None,
                project_code=project_code or None,
            )

            _by_bp = {}
            for _r in summary_rows:
                _bpc2 = str(_r.get('BPCode') or '')
                _bp2  = _by_bp.setdefault(_bpc2, {
                    'BPCode': _bpc2,
                    'BPName': str(_r.get('BPName') or ''),
                    'rows': [],
                })
                _opening2 = 0.0
                if from_date:
                    try:
                        _opening2 = hana_queries.project_opening_balance(
                            conn2, _bpc2, str(_r.get('ProjectCode') or ''), from_date
                        )
                    except Exception:
                        _opening2 = 0.0
                _td2 = float(_r.get('TotalDebit')  or 0)
                _tc2 = float(_r.get('TotalCredit') or 0)
                _r['Opening'] = _opening2
                _r['Balance'] = _opening2 + _td2 - _tc2
                _bp2['rows'].append(_r)
            conn2.close()
        except Exception as _se:
            logger.exception("Detail ledger: error building policy summary: %s", _se)
            _by_bp = {}

        if _by_bp:
            elements.append(PageBreak())

            _sum_col_widths = [
                3.40 * inch,
                0.80 * inch,
                0.80 * inch,
                0.70 * inch,
                0.80 * inch,
                0.90 * inch,
                0.85 * inch,
                0.85 * inch,
                0.85 * inch,
            ]
            _SUM_NCOLS = len(_sum_col_widths)

            _sum_data   = []
            _sum_styles = [
                ('FONTSIZE',      (0, 0), (-1, -1), _fs_td),
                ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING',   (0, 0), (-1, -1), 3),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
                ('TOPPADDING',    (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LINEBELOW',     (0, 0), (-1, -1), 0.25, C_GRID),
                ('ALIGN',         (1, 0), (-1, -1), 'RIGHT'),
            ]

            # ── 2-row summary header ──────────────────────────────────────────
            # Row 0: Policy Name | Opening | Sales | Tax | Return | Collection | Total(span6-7) | Balance
            # Row 1: (spanned)   | (spanned)| ...  | ... | ...    | (spanned)  | Debit | Credit | (spanned)
            _sum_data.append(['Policy Name', 'Opening', 'Sales', 'Tax', 'Return', 'Collection', 'Total', '', 'Balance'])
            _sum_data.append(['', '', '', '', '', '', 'Debit', 'Credit', ''])
            _sri = 0   # row index of first header row
            _sum_styles += [
                # Background + text for both header rows
                ('BACKGROUND',    (0, _sri), (-1, _sri + 1), C_HEADER_BG),
                ('TEXTCOLOR',     (0, _sri), (-1, _sri + 1), C_HEADER_TXT),
                ('FONTNAME',      (0, _sri), (-1, _sri + 1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, _sri), (-1, _sri + 1), _fs_th),
                ('ALIGN',         (0, _sri), (-1, _sri + 1), 'CENTER'),
                ('VALIGN',        (0, _sri), (-1, _sri + 1), 'MIDDLE'),
                ('TOPPADDING',    (0, _sri), (-1, _sri + 1), 3),
                ('BOTTOMPADDING', (0, _sri), (-1, _sri + 1), 3),
                ('LINEABOVE',     (0, _sri),      (-1, _sri),      0.75, C_GRID),
                ('LINEBELOW',     (0, _sri + 1),  (-1, _sri + 1),  0.75, C_GRID),
                # Row-spans: cols 0-5 and col 8 span both header rows
                ('SPAN', (0, _sri), (0, _sri + 1)),
                ('SPAN', (1, _sri), (1, _sri + 1)),
                ('SPAN', (2, _sri), (2, _sri + 1)),
                ('SPAN', (3, _sri), (3, _sri + 1)),
                ('SPAN', (4, _sri), (4, _sri + 1)),
                ('SPAN', (5, _sri), (5, _sri + 1)),
                ('SPAN', (8, _sri), (8, _sri + 1)),
                # Col-span: "Total" across cols 6-7 in row 0
                ('SPAN', (6, _sri), (7, _sri)),
                # Box border only around Total+Debit+Credit group and Balance
                ('BOX',       (6, _sri), (7, _sri + 1), 1,   C_HEADER_TXT),
                ('LINEBELOW', (6, _sri), (7, _sri),     0.5, C_HEADER_TXT),
                ('BOX',       (8, _sri), (8, _sri + 1), 1,   C_HEADER_TXT),
            ]
            _sri += 2   # data rows start after both header rows

            def _sfmt(v):
                try:
                    f = float(v or 0)
                except (TypeError, ValueError):
                    f = 0.0
                return '{:,.2f}'.format(f) if f else '0'

            for _bpc2, _bp2 in _by_bp.items():
                _bp_label2 = (
                    f"{_bpc2}    {_bp2['BPName']}".strip()
                    if _bpc2 else (_bp2['BPName'] or 'No BP')
                )
                _sum_data.append([_bp_label2] + [''] * (_SUM_NCOLS - 1))
                _sum_styles += [
                    ('SPAN',         (0, _sri), (_SUM_NCOLS - 1, _sri)),
                    ('BACKGROUND',   (0, _sri), (-1, _sri), C_TERR_BG),
                    ('FONTNAME',     (0, _sri), (-1, _sri), 'Helvetica-Bold'),
                    ('FONTSIZE',     (0, _sri), (-1, _sri), 8),
                    ('LINEABOVE',    (0, _sri), (-1, _sri), 0.5, C_GRID),
                    ('LINEBELOW',    (0, _sri), (-1, _sri), 0.5, C_GRID),
                ]
                _sri += 1
                for _r in _bp2['rows']:
                    _pname = (
                        f"{_r.get('ProjectCode','')} {_r.get('ProjectName','')}"
                    ).strip()
                    _sum_data.append([
                        Paragraph(_pname, cell_style),
                        _sfmt(_r.get('Opening')),
                        _sfmt(_r.get('Sales')),
                        _sfmt(_r.get('Tax')),
                        _sfmt(_r.get('Return')),
                        _sfmt(_r.get('Collection')),
                        _sfmt(_r.get('TotalDebit')),
                        _sfmt(_r.get('TotalCredit')),
                        _sfmt(_r.get('Balance')),
                    ])
                    _sri += 1

            _sum_table = Table(_sum_data, colWidths=_sum_col_widths, repeatRows=2)
            _sum_table.setStyle(TableStyle(_sum_styles))
            elements.append(_sum_table)

        # ── Urdu footer (shared logic with GL) ────────────────────────────────
        if _cfg.footer_image:
            try:
                from reportlab.lib.utils import ImageReader as _ImgReader
                _ir  = _ImgReader(_cfg.footer_image.path)
                _iw, _ih = _ir.getSize()
                _img_w = landscape(letter)[0] - 0.6 * inch
                _img_h = _img_w * (_ih / _iw) if _iw else 1.2 * inch
                _img_max_h = 2.0 * inch
                if _img_h > _img_max_h:
                    _img_h = _img_max_h
                    _img_w = _img_h * (_iw / _ih) if _ih else _img_w
                elements.append(RLImage(_cfg.footer_image.path, width=_img_w, height=_img_h))
            except Exception:
                pass

        _urdu_lines = _cfg.footer_lines()
        _urdu_reshaped_dl = []
        if _urdu_lines:
            for _line in _urdu_lines:
                if ARABIC_AVAILABLE:
                    try:
                        _res = (_urdu_reshaper.reshape(_line)
                                if _urdu_reshaper is not None
                                else arabic_reshaper.reshape(_line))
                        _urdu_reshaped_dl.append(bidi_display(_res))
                    except Exception:
                        _urdu_reshaped_dl.append(_line)
                else:
                    _urdu_reshaped_dl.append(_line)

        def _draw_dl_footer(canvas):
            if _cfg.footer_image or not _urdu_reshaped_dl:
                return
            page_width    = landscape(letter)[0]
            bottom_margin = 0.4 * inch
            box_width     = page_width - 0.6 * inch
            box_x         = 0.3 * inch
            canvas.saveState()
            try:
                from reportlab.lib.utils import simpleSplit
                _font_size  = 10
                _padding    = 8
                _line_gap   = 3
                max_text_w  = box_width - (_padding * 2)
                all_wrapped = []
                for _lt in _urdu_reshaped_dl:
                    _wr = simpleSplit(_lt, _URDU_FONT, _font_size, max_text_w)
                    all_wrapped.append(_wr if _wr else [_lt])
                single_line_h = _font_size * 1.4
                y_pos = bottom_margin + 5
                for _wl in reversed(all_wrapped):
                    box_h = (single_line_h * len(_wl)) + (_padding * 2)
                    canvas.setFillColor(colors.HexColor('#f8f9ff'))
                    canvas.rect(box_x, y_pos, box_width, box_h, fill=1, stroke=0)
                    canvas.setFillColor(colors.HexColor('#1e293b'))
                    canvas.setFont(_URDU_FONT, _font_size)
                    text_x = box_x + _padding
                    text_y = y_pos + box_h - _padding - _font_size
                    for _sl in _wl:
                        canvas.drawString(text_x, text_y, _sl)
                        text_y -= single_line_h
                    y_pos += box_h + _line_gap
            except Exception:
                pass
            finally:
                canvas.restoreState()

        doc.build(elements, canvasmaker=_make_last_page_only_canvas(_draw_dl_footer))
        pdf_buffer.seek(0)
        response.write(pdf_buffer.getvalue())
        return response

    except Exception as e:
        logger.exception("Error exporting Detail Ledger PDF")
        return Response({
            'success': False,
            'error': f'Error exporting Detail Ledger PDF: {str(e)}'
        }, status=500)
