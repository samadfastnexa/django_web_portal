from io import BytesIO
from django.test import TestCase
from django.utils import timezone

from ..models import Trial, TrialTreatment, Product
from ..exports import build_trials_workbook, build_trials_pdf


class ExportBuildersTest(TestCase):
    def setUp(self):
        # Minimal dataset
        self.trial1 = Trial.objects.create(
            station='Station A',
            trial_name='Trial 1',
            location_area='Area 1',
            crop_variety='Wheat',
            application_date=timezone.now().date(),
            design_replicates='Replicated',
            water_volume_used='120 L/A',
            temp_min_c=11.0,
            temp_max_c=21.0,
            humidity_min_percent=57,
            humidity_max_percent=74,
            wind_velocity_kmh=10.0,
            rainfall='Nil',
            created_at=timezone.now(),
        )
        self.product = Product.objects.create(name='Prod X', brand='Brand X', active_ingredient='AI')
        TrialTreatment.objects.create(
            trial=self.trial1,
            label='T1',
            product=self.product,
            crop_stage_soil='Loamy soil',
            pest_stage_start='Early infestation',
            crop_safety_stress_rating=5,
            details='Observed 40% control after 7 days',
            growth_improvement_type='Improved tillering',
            best_dose='100 ml',
            others='Notes',
        )

    def test_build_trials_workbook_basic(self):
        trials = Trial.objects.all()
        treatments = TrialTreatment.objects.select_related('trial', 'product').all()
        output = build_trials_workbook(trials, treatments)
        self.assertIsInstance(output, BytesIO)
        self.assertGreater(len(output.getvalue()), 0)

        # Inspect workbook styling
        import openpyxl
        wb = openpyxl.load_workbook(output)
        self.assertIn('Trials', wb.sheetnames)
        self.assertIn('Treatments', wb.sheetnames)

        ws_trials = wb['Trials']
        # Header should be bold with a background fill
        first_cell = ws_trials['A1']
        self.assertTrue(first_cell.font.bold)
        self.assertEqual(first_cell.fill.fill_type, 'solid')
        # Color may be ARGB; tolerate both representations
        self.assertTrue(
            str(first_cell.fill.start_color.value).lower().endswith('bcd3e0')
        )

    def test_build_trials_pdf_chunking_and_wrapping(self):
        # Add enough treatments to trigger chunking/page breaks
        for i in range(40):
            TrialTreatment.objects.create(
                trial=self.trial1,
                label=f'T{i+2}',
                product=self.product,
                crop_stage_soil='Loamy soil, adequate moisture',
                pest_stage_start='Aphids at early infestation',
                crop_safety_stress_rating=5,
                details='A long description that should wrap correctly within the table cell.',
                growth_improvement_type='Improved tillering',
                best_dose='100 ml',
                others='More notes',
            )
        trials = Trial.objects.all()
        treatments = TrialTreatment.objects.select_related('trial', 'product').all()
        buffer = build_trials_pdf(trials, treatments)
        self.assertIsInstance(buffer, BytesIO)
        # PDF should not be empty and should be at least a few KB
        self.assertGreater(len(buffer.getvalue()), 1024)