from io import BytesIO

# XLSX
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

# PDF
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
except Exception:
    SimpleDocTemplate = None

LARGE_DATASET_THRESHOLD = 10000  # rows across sheets


def _apply_common_ws_styles(ws, header_fill_hex, freeze='A2'):
    # Freeze header row
    ws.freeze_panes = freeze
    # Auto filter on header row
    ws.auto_filter.ref = ws.dimensions
    # Header style
    header_font = Font(name='Calibri', bold=True, color='000000')
    header_fill = PatternFill(start_color=header_fill_hex.replace('#', ''), end_color=header_fill_hex.replace('#', ''), fill_type='solid')
    border = Border(left=Side(style='thin', color='809EAE'),
                    right=Side(style='thin', color='809EAE'),
                    top=Side(style='thin', color='809EAE'),
                    bottom=Side(style='thin', color='809EAE'))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Set row heights for readability
    ws.row_dimensions[1].height = 22


def _set_column_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _to_excel_dt(value):
    """Ensure date/datetime is timezone-naive for Excel.

    openpyxl rejects timezone-aware datetimes. Dates are fine as-is.
    """
    if value is None:
        return None
    import datetime
    if isinstance(value, datetime.datetime):
        # Drop tzinfo; assume value is already in desired timezone
        return value.replace(tzinfo=None)
    return value


def build_trials_workbook(trials_qs, treatments_qs):
    """Create styled XLSX workbook for Trials and Treatments.

    Applies fonts, borders, fills, widths, number formats and alignment.
    Switches to write-only mode for very large datasets.
    """
    if Workbook is None:
        raise RuntimeError('openpyxl is not installed')

    total_rows = trials_qs.count() + treatments_qs.count()
    wb = Workbook(write_only=total_rows > LARGE_DATASET_THRESHOLD)

    ws_trials = wb.create_sheet('Trials') if wb.write_only else wb.active
    if not wb.write_only:
        ws_trials.title = 'Trials'

    # Vertical layout: label/value rows per trial
    # Configure column widths for portrait-friendly layout
    if not wb.write_only:
        _set_column_widths(ws_trials, [24, 64])

    # Predefine styles for label/value when not in write-only mode
    label_fill = PatternFill(start_color='DBE8EF', end_color='DBE8EF', fill_type='solid')
    border = Border(left=Side(style='thin', color='809EAE'),
                    right=Side(style='thin', color='809EAE'),
                    top=Side(style='thin', color='809EAE'),
                    bottom=Side(style='thin', color='809EAE'))

    current_row = 1
    for tr in trials_qs.order_by('station', 'trial_name'):
        title_text = f"Trial — {tr.trial_name} ({tr.station})"
        if wb.write_only:
            ws_trials.append([title_text, ''])
        else:
            ws_trials.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            c = ws_trials.cell(row=current_row, column=1, value=title_text)
            c.font = Font(name='Calibri', bold=True, size=12)
            c.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        def add_row(label, value, num_fmt=None):
            nonlocal current_row
            if wb.write_only:
                ws_trials.append([label, value])
            else:
                lc = ws_trials.cell(row=current_row, column=1, value=label)
                vc = ws_trials.cell(row=current_row, column=2, value=value)
                lc.font = Font(name='Calibri', bold=True)
                lc.fill = label_fill
                lc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                vc.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                lc.border = border
                vc.border = border
                if num_fmt:
                    vc.number_format = num_fmt
            current_row += 1

        add_row('Location/Area', tr.location_area)
        add_row('Crop/Variety', tr.crop_variety)
        add_row('Application Date', _to_excel_dt(tr.application_date), num_fmt='YYYY-MM-DD')
        add_row('Design/Replicates', tr.design_replicates)
        add_row('Water Volume Used', tr.water_volume_used)
        add_row('Previous Sprays', tr.previous_sprays or '')
        add_row('Temp Min (°C)', float(tr.temp_min_c) if tr.temp_min_c is not None else None, num_fmt='0.0')
        add_row('Temp Max (°C)', float(tr.temp_max_c) if tr.temp_max_c is not None else None, num_fmt='0.0')
        add_row('Humidity Min (%)', tr.humidity_min_percent, num_fmt='0')
        add_row('Humidity Max (%)', tr.humidity_max_percent, num_fmt='0')
        add_row('Wind Velocity (km/h)', float(tr.wind_velocity_kmh) if tr.wind_velocity_kmh is not None else None, num_fmt='0.0')
        add_row('Rainfall', tr.rainfall)
        add_row('Created At', _to_excel_dt(tr.created_at), num_fmt='YYYY-MM-DD HH:MM:SS')

        # spacer row between trials
        if wb.write_only:
            ws_trials.append(['', ''])
        current_row += 1

    # Treatments sheet
    ws_treat = wb.create_sheet('Treatments')
    t_headers = [
        'Station', 'Trial Name', 'Treatment Label', 'Product',
        'Crop Stage/Soil', 'Pest Stage Start', 'Safety/Stress (1-9)',
        'Details', 'Growth Improvement', 'Best Dose', 'Others'
    ]
    ws_treat.append(t_headers)

    for t in treatments_qs.order_by('trial__station', 'trial__trial_name', 'label'):
        ws_treat.append([
            t.trial.station,
            t.trial.trial_name,
            t.label,
            t.product.name if t.product else '',
            t.crop_stage_soil or '',
            t.pest_stage_start or '',
            t.crop_safety_stress_rating if t.crop_safety_stress_rating is not None else '',
            t.details or '',
            t.growth_improvement_type or '',
            t.best_dose or '',
            t.others or '',
        ])

    if not wb.write_only:
        _apply_common_ws_styles(ws_treat, '#c4e7ae')
        _set_column_widths(ws_treat, [16, 16, 14, 18, 30, 26, 12, 36, 24, 14, 20])
        for row in ws_treat.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_trials_pdf(trials_qs, treatments_qs):
    """Create styled PDF with page-safe tables.

    Uses Paragraphs for wrapping and splits tables across pages to avoid cutoff.
    """
    if SimpleDocTemplate is None:
        raise RuntimeError('reportlab is not installed')

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
        title='Trials Report'
    )
    styles = getSampleStyleSheet()
    normal = styles['BodyText']
    normal.fontSize = 8.5
    # Encourage word wrapping for long English text
    try:
        normal.wordWrap = 'LTR'
    except Exception:
        pass

    elements = [Paragraph('Trials Report', styles['Title']), Spacer(1, 10)]

    # Trials summary section — vertical field/value table per trial
    elements.append(Paragraph('Trials Summary', styles['Heading2']))
    elements.append(Spacer(1, 6))

    # helper for scaling widths to page
    def _scale(width_factors):
        total = sum(width_factors)
        return [doc.width * (w / total) for w in width_factors]

    label_style = styles['BodyText']
    label_style.fontSize = 8.5
    label_style.wordWrap = 'LTR'

    for tr in trials_qs.order_by('station', 'trial_name'):
        rows = [
            [Paragraph('Station', label_style), Paragraph(tr.station, normal)],
            [Paragraph('Trial Name', label_style), Paragraph(tr.trial_name, normal)],
            [Paragraph('Location/Area', label_style), Paragraph(tr.location_area, normal)],
            [Paragraph('Crop/Variety', label_style), Paragraph(tr.crop_variety, normal)],
            [Paragraph('Application Date', label_style), Paragraph(tr.application_date.strftime('%Y-%m-%d') if tr.application_date else '', normal)],
            [Paragraph('Design/Replicates', label_style), Paragraph(tr.design_replicates, normal)],
            [Paragraph('Water Vol.', label_style), Paragraph(tr.water_volume_used, normal)],
            [Paragraph('Prev. Sprays', label_style), Paragraph(tr.previous_sprays or '', normal)],
            [Paragraph('Temp Min (°C)', label_style), Paragraph(str(tr.temp_min_c), normal)],
            [Paragraph('Temp Max (°C)', label_style), Paragraph(str(tr.temp_max_c), normal)],
            [Paragraph('Humidity Min (%)', label_style), Paragraph(str(tr.humidity_min_percent), normal)],
            [Paragraph('Humidity Max (%)', label_style), Paragraph(str(tr.humidity_max_percent), normal)],
            [Paragraph('Wind km/h', label_style), Paragraph(str(tr.wind_velocity_kmh), normal)],
            [Paragraph('Rainfall', label_style), Paragraph(tr.rainfall, normal)],
        ]
        vtable = Table(rows, colWidths=_scale([0.28, 0.72]))
        vtable.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#dbe8ef')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#809eae')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(vtable)
        elements.append(Spacer(1, 10))

    # Treatments section — vertical field/value table per treatment
    elements.append(Paragraph('Treatments', styles['Heading2']))
    elements.append(Spacer(1, 6))

    per_page = 10
    processed = 0
    for t in treatments_qs.order_by('trial__station', 'trial__trial_name', 'label'):
        # Heading per treatment block
        elements.append(Paragraph(
            f"Treatment — {t.label} ({t.trial.trial_name}, {t.trial.station})",
            styles['Heading3']
        ))

        t_rows = [
            [Paragraph('Product', label_style), Paragraph((t.product.name if t.product else ''), normal)],
            [Paragraph('Crop Stage/Soil', label_style), Paragraph(t.crop_stage_soil or '', normal)],
            [Paragraph('Pest Stage Start', label_style), Paragraph(t.pest_stage_start or '', normal)],
            [Paragraph('Safety/Stress (1-9)', label_style), Paragraph('' if t.crop_safety_stress_rating is None else str(t.crop_safety_stress_rating), normal)],
            [Paragraph('Details', label_style), Paragraph(t.details or '', normal)],
            [Paragraph('Growth Improvement', label_style), Paragraph(t.growth_improvement_type or '', normal)],
            [Paragraph('Best Dose', label_style), Paragraph(t.best_dose or '', normal)],
            [Paragraph('Others', label_style), Paragraph(t.others or '', normal)],
        ]

        t_vtable = Table(t_rows, colWidths=_scale([0.28, 0.72]))
        t_vtable.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#c4e7ae')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#8ecb6b')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(t_vtable)
        elements.append(Spacer(1, 8))

        processed += 1
        if processed % per_page == 0:
            elements.append(PageBreak())

    doc.build(elements)
    buffer.seek(0)
    return buffer